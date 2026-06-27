import os
import django
import pandas as pd
import io
from contextlib import redirect_stdout

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'elanelsystem.settings.prod')
django.setup()

from users.models import Sucursal, Usuario
from liquidacion.models import Liquidacion
from liquidacion.views import recalcular_liquidacion_data

campaigns_2025 = [
    "Enero 2025", "Febrero 2025", "Marzo 2025", "Abril 2025",
    "Mayo 2025", "Junio 2025", "Julio 2025", "Agosto 2025",
    "Septiembre 2025", "Octubre 2025", "Noviembre 2025", "Diciembre 2025"
]

sucursales = list(Sucursal.objects.all())
excel_path = "../liquidaciones_2025.xlsx"

# To collect agency-month totals
summary_data = {camp: {suc.pseudonimo: 0 for suc in sucursales} for camp in campaigns_2025}

print("Iniciando generación de liquidaciones de 2025...")

agency_dfs = {}

for suc in sucursales:
    print(f"Procesando sucursal: {suc.pseudonimo}...")
    suc_data = []
    
    for camp in campaigns_2025:
        try:
            with redirect_stdout(io.StringIO()):
                colaboradores_list, total_comisiones = recalcular_liquidacion_data(None, camp, suc.id)
            summary_data[camp][suc.pseudonimo] = total_comisiones
            
            for col in colaboradores_list:
                info = col.get("info_total_de_comision", {})
                suc_data.append({
                    "Campaña": camp,
                    "Nombre": col.get("nombre"),
                    "DNI": col.get("dni"),
                    "Rol": col.get("tipo_colaborador"),
                    "Neto Liquidado": col.get("comisionTotal", 0),
                    "Comisión Bruta": info.get("comision_bruta", 0),
                    "Asegurado (Diferencia)": info.get("asegurado", 0),
                    "Descuento Total": col.get("descuentoTotal", 0)
                })
        except Exception as e:
            print(f"  Error en campaña {camp} - sucursal {suc.pseudonimo}: {e}")
            
    if suc_data:
        agency_dfs[suc.pseudonimo] = pd.DataFrame(suc_data)

# Summary DataFrame
summary_rows = []
for camp in campaigns_2025:
    row = {"Mes": camp}
    total_mes = 0
    for suc in sucursales:
        val = summary_data[camp][suc.pseudonimo]
        row[suc.pseudonimo] = val
        total_mes += val
    row["Total General"] = total_mes
    summary_rows.append(row)

df_summary = pd.DataFrame(summary_rows)

# Let's write using openpyxl
print("Escribiendo datos iniciales a Excel...")
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name="Resumen General", index=False)
    for suc_name, df_detail in agency_dfs.items():
        sheet_name = suc_name.replace("/", "-").replace(":", "-")[:30]
        df_detail.to_excel(writer, sheet_name=sheet_name, index=False)

# Let's style it with openpyxl post-write
print("Aplicando estilos profesionales...")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook(excel_path)

# Styles
font_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
font_total = Font(name="Calibri", size=11, bold=True)
fill_total = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
border_thin = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)
border_double = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)

for name in wb.sheetnames:
    ws = wb[name]
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Format Header
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = font_title
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Format cells
    if name == "Resumen General":
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')
            ws.cell(row=r, column=1).border = border_thin
            for c in range(2, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border_thin
                
        # Total Row
        ws.cell(row=max_row + 1, column=1, value="TOTAL GENERAL 2025").font = font_total
        ws.cell(row=max_row + 1, column=1).fill = fill_total
        ws.cell(row=max_row + 1, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=max_row + 1, column=1).border = border_double
        
        for c in range(2, max_col + 1):
            col_letter = get_column_letter(c)
            cell = ws.cell(row=max_row + 1, column=c)
            cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
            cell.font = font_total
            cell.fill = fill_total
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border_double
            
    else:
        for r in range(2, max_row + 1):
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = border_thin
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='left')
            for c in range(5, 9):
                cell = ws.cell(row=r, column=c)
                cell.number_format = '$#,##0'
                cell.alignment = Alignment(horizontal='right')
                cell.border = border_thin
                
        # Total Row
        ws.cell(row=max_row + 1, column=4, value="TOTAL:").font = font_total
        ws.cell(row=max_row + 1, column=4).alignment = Alignment(horizontal='right')
        ws.cell(row=max_row + 1, column=4).border = border_double
        ws.cell(row=max_row + 1, column=4).fill = fill_total
        
        # fill other cells in total label range
        for c in range(1, 4):
            ws.cell(row=max_row + 1, column=c).border = border_double
            ws.cell(row=max_row + 1, column=c).fill = fill_total
            
        for c in range(5, 9):
            col_letter = get_column_letter(c)
            cell = ws.cell(row=max_row + 1, column=c)
            cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
            cell.font = font_total
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='right')
            cell.border = border_double
            cell.fill = fill_total
            
    # Auto-adjust column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format == '$#,##0' and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.0f}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(excel_path)
print(f"Reporte de liquidaciones 2025 generado y formateado con éxito en: {os.path.abspath(excel_path)}")
