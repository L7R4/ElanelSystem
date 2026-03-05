from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from products.models import Products, Plan


def _normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ü": "u", "ñ": "n",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = s.replace("\n", " ")
    s = " ".join(s.split())
    s = s.replace(" ", "_")
    return s


def _to_int(value, field_name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"Campo requerido vacío: {field_name}")
    if isinstance(value, bool):
        raise ValueError(f"Valor inválido para {field_name}: {value}")
    try:
        if isinstance(value, str):
            value = value.replace(".", "").replace(",", ".").strip()
        return int(float(value))
    except Exception as e:
        raise ValueError(f"No pude convertir {field_name}='{value}' a int") from e


def _to_title_case(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    words = text.split()
    return " ".join(word[:1].upper() + word[1:] for word in words)


def _sheet_is_product_sheet(header_map: dict) -> bool:
    return "producto" in header_map and "valor_nominal" in header_map


def _infer_tipo_de_producto(ws) -> str:
    title = ws.title.strip().lower()

    if "moto" in title:
        return "Moto"
    if "combo" in title:
        return "Combo"
    if "solucion" in title or "solución" in title:
        return "Solucion"
    if "default" in title:
        return "Default"

    sample_texts = []
    for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 8), values_only=True):
        if not row:
            continue
        first_cell = row[0]
        if first_cell:
            sample_texts.append(str(first_cell).strip().lower())

    joined = " ".join(sample_texts)

    if "solucion" in joined or "solución" in joined:
        return "Solucion"
    if "combo" in joined:
        return "Combo"

    return "Moto"


class Command(BaseCommand):
    help = "Importa productos desde un Excel y los asocia a un Plan activo por valor_nominal."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            "-f",
            dest="file_path",
            help="Ruta al archivo .xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula sin guardar en DB.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options.get("file_path")
        dry_run = options.get("dry_run")

        if not file_path:
            file_path = input("Ruta del Excel (.xlsx): ").strip()

        if not file_path.lower().endswith(".xlsx"):
            raise CommandError("El archivo debe ser .xlsx")

        try:
            wb = load_workbook(filename=file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"No pude abrir el Excel: {e}")

        created = 0
        linked = 0
        unlinked = 0
        errors = 0
        skipped_sheets = 0

        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                skipped_sheets += 1
                continue

            header = rows[0]
            header_map = {}

            for idx, col_name in enumerate(header):
                key = _normalize_header(col_name)
                if key:
                    header_map[key] = idx

            if not _sheet_is_product_sheet(header_map):
                skipped_sheets += 1
                continue

            idx_producto = header_map["producto"]
            idx_valor = header_map["valor_nominal"]

            tipo_de_producto = _infer_tipo_de_producto(ws)

            for i, row in enumerate(rows[1:], start=2):
                if row is None or all(v is None or v == "" for v in row):
                    continue

                try:
                    nombre = row[idx_producto]
                    valor_nominal = _to_int(row[idx_valor], "Valor Nominal")

                    if nombre is None or str(nombre).strip() == "":
                        raise ValueError("Nombre de producto vacío")

                    nombre = _to_title_case(nombre)

                    plan = (
                        Plan.objects
                        .filter(valor_nominal=valor_nominal, activo=True)
                        .order_by("-id")
                        .first()
                    )

                    Products.objects.create(
                        tipodePlan="Estandar",
                        tipo_de_producto=tipo_de_producto,
                        nombre=nombre,
                        plan=plan,
                        activo=True,
                    )

                    created += 1

                    if plan:
                        linked += 1
                    else:
                        unlinked += 1

                except Exception as e:
                    errors += 1
                    self.stderr.write(
                        self.style.ERROR(
                            f"Hoja '{ws.title}' fila {i}: {e}"
                        )
                    )

        if dry_run:
            transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f"Import terminado. created={created}, linked={linked}, "
            f"unlinked={unlinked}, skipped_sheets={skipped_sheets}, errors={errors}"
            + (" (DRY RUN)" if dry_run else "")
        ))