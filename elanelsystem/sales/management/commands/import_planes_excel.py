from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
from products.models import Plan


def _normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u",
        "ü": "u", "ñ": "n",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = s.replace("\n", " ")
    s = " ".join(s.split())
    s = s.replace(" ", "_")
    return s


def _to_int(value, field_name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"Campo requerido vacío: {field_name}")
    if isinstance(value, bool):
        raise ValueError(f"Valor inválido para {field_name}: {value}")
    try:
        if isinstance(value, str):
            value = value.replace(".", "").replace(",", ".").strip()
        return int(float(value))
    except Exception as e:
        raise ValueError(f"No pude convertir {field_name}='{value}' a int") from e


def _to_percent_factor(value, field_name: str) -> float:
    """
    Convierte:
      - "15%"    -> 0.15
      - "14.99%" -> 0.1499
      - 15       -> 0.15
      - 14.99    -> 0.1499
      - 0.15     -> 0.15
    Y redondea a 4 decimales.
    """
    if value is None or value == "":
        return 0.0

    try:
        if isinstance(value, str):
            raw = value.strip().replace(",", ".")
            if raw.endswith("%"):
                raw = raw[:-1].strip()
                num = Decimal(raw) / Decimal("100")
            else:
                num = Decimal(raw)
        else:
            num = Decimal(str(value))

        if num > 1:
            num = num / Decimal("100")

        num = num.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

        return float(num)

    except Exception as e:
        raise ValueError(f"No pude convertir {field_name}='{value}' a porcentaje") from e


class Command(BaseCommand):
    help = "Importa planes (Plan) desde un Excel, creando siempre nuevos registros."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            "-f",
            dest="file_path",
            help="Ruta al archivo .xlsx",
        )
        parser.add_argument(
            "--sheet",
            "-s",
            dest="sheet_name",
            default=None,
            help="Nombre de la hoja (opcional). Si no, usa la activa.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula sin guardar en DB.",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options.get("file_path")
        sheet_name = options.get("sheet_name")
        dry_run = options.get("dry_run")

        if not file_path:
            file_path = input("Ruta del Excel (.xlsx): ").strip()

        if not file_path.lower().endswith(".xlsx"):
            raise CommandError("El archivo debe ser .xlsx")

        try:
            wb = load_workbook(filename=file_path, data_only=True)
        except Exception as e:
            raise CommandError(f"No pude abrir el Excel: {e}")

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise CommandError(
                    f"La hoja '{sheet_name}' no existe. Disponibles: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("La hoja está vacía.")

        header = rows[0]
        header_map = {}
        for idx, col_name in enumerate(header):
            key = _normalize_header(col_name)
            if key:
                header_map[key] = idx

        required = {
            "valor_nominal": ["valor_nominal", "valor_nominal_", "valor_nominal__"],
            "suscripcion": ["suscripcion", "suscripcion_", "suscripcion__"],
            "porcentaje_24": ["porcentaje_24", "porcentaje24"],
            "porcentaje_30": ["porcentaje_30", "porcentaje30"],
            "porcentaje_48": ["porcentaje_48", "porcentaje48"],
            "porcentaje_60": ["porcentaje_60", "porcentaje60"],
        }

        def pick_index(aliases):
            for a in aliases:
                if a in header_map:
                    return header_map[a]
            return None

        idx_valor = pick_index(required["valor_nominal"])
        idx_susc = pick_index(required["suscripcion"])
        idx_24 = pick_index(required["porcentaje_24"])
        idx_30 = pick_index(required["porcentaje_30"])
        idx_48 = pick_index(required["porcentaje_48"])
        idx_60 = pick_index(required["porcentaje_60"])

        missing = []
        if idx_valor is None: missing.append("Valor Nominal")
        if idx_susc is None: missing.append("Suscripción")
        if idx_24 is None: missing.append("Porcentaje 24")
        if idx_30 is None: missing.append("Porcentaje 30")
        if idx_48 is None: missing.append("Porcentaje 48")
        if idx_60 is None: missing.append("Porcentaje 60")

        if missing:
            raise CommandError(
                "Faltan columnas requeridas en el Excel: " + ", ".join(missing)
            )

        created = 0
        errors = 0

        for i, row in enumerate(rows[1:], start=2):
            if row is None or all(v is None or v == "" for v in row):
                continue

            try:
                valor_nominal = _to_int(row[idx_valor], "Valor Nominal")
                suscripcion = _to_int(row[idx_susc], "Suscripción")

                c24 = _to_percent_factor(row[idx_24], "Porcentaje 24")
                c30 = _to_percent_factor(row[idx_30], "Porcentaje 30")
                c48 = _to_percent_factor(row[idx_48], "Porcentaje 48")
                c60 = _to_percent_factor(row[idx_60], "Porcentaje 60")

                Plan.objects.create(
                    valor_nominal=valor_nominal,
                    suscripcion=suscripcion,
                    primer_cuota=suscripcion,
                    activo=True,
                    c24_porcentage=c24,
                    c30_porcentage=c30,
                    c48_porcentage=c48,
                    c60_porcentage=c60,
                )

                created += 1

            except Exception as e:
                errors += 1
                self.stderr.write(self.style.ERROR(f"Fila {i}: {e}"))

        if dry_run:
            transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f"Import terminado. created={created}, errors={errors}"
            + (" (DRY RUN)" if dry_run else "")
        ))