# Toda la campaña (todas las sucursales con participantes)
# python manage.py export_comisiones --campania "Julio 2025"

# # Solo algunas sucursales por ID
# python manage.py export_comisiones --campania "Julio 2025" --sucursal_id 1 3 7

# # Con archivo de ajustes (mismo formato que guardas en sesión)
# python manage.py export_comisiones --campania "Julio 2025" --ajustes_json ./ajustes.json

# # Elegir ruta de salida
# python manage.py export_comisiones --campania "Julio 2025" --output ./exports/liquidaciones_julio_2025.xlsx

# # Incluir también sucursales sin colaboradores en la campaña
# python manage.py export_comisiones --campania "Julio 2025" --incluir_vacias


from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from users.models import Sucursal  # MISMO APP de tus utils/models
from liquidacion.utils import get_comision_total  # usa tu lógica existente
from users.utils import obtener_usuarios_segun_campana, snapshot_usuario_by_campana

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, NamedStyle, Border, Side


def safe_sheet_title(name: str) -> str:
    """
    Limpia título de hoja para Excel y lo recorta a 31 chars.
    """
    name = re.sub(r'[\[\]\*\?\/\\\:]', '-', name)
    return (name or "Hoja")[:31]


def autosize_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        col = column_cells[0].column if hasattr(column_cells[0], "column") else column_cells[0].column_letter
        col_letter = get_column_letter(col)
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            max_length = max(max_length, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_length + 2), 45)


def build_styles(wb: Workbook):
    currency = NamedStyle(name="ars_currency")
    currency.number_format = '#,##0'  # entero sin decimales; cambiá a '#,##0.00' si querés 2 decimales
    wb.add_named_style(currency)

    header = NamedStyle(name="header")
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="4F81BD")
    header.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="D9D9D9")
    header.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(header)

    total = NamedStyle(name="total")
    total.font = Font(bold=True)
    total.fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(border_style="thin", color="BFBFBF")
    total.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(total)

    sub = NamedStyle(name="sub")
    sub.font = Font(bold=True)
    sub.fill = PatternFill("solid", fgColor="FFF2CC")
    sub.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb.add_named_style(sub)

class Command(BaseCommand):
    help = "Exporta comisiones por campaña a Excel, una hoja por Sucursal, manteniendo la lógica de participantes y cálculo."

    def add_arguments(self, parser):
        parser.add_argument(
            "--campania",
            required=True,
            help='Campaña en formato "Mes Año" (ej. "Julio 2025").'
        )
        parser.add_argument(
            "--sucursal_id",
            nargs="*",
            type=int,
            help="(Opcional) Uno o varios IDs de sucursal para limitar el export."
        )
        parser.add_argument(
            "--output",
            help="Ruta del archivo .xlsx de salida. Por defecto: ./export_comisiones_{campania}.xlsx"
        )
        parser.add_argument(
            "--ajustes_json",
            help="(Opcional) Ruta a un JSON con ajustes de comisión (mismo formato que guardas en sesión)."
        )
        parser.add_argument(
            "--incluir_vacias",
            action="store_true",
            help="Incluye sucursales sin colaboradores en la campaña (por defecto se omiten)."
        )

    @transaction.atomic
    def handle(self, *args, **options):
        campania = options["campania"]
        suc_ids = options.get("sucursal_id") or []
        ajustes_path = options.get("ajustes_json")

        # Cargar ajustes (opcional)
        ajustes = []
        if ajustes_path:
            p = Path(ajustes_path)
            if not p.exists():
                raise CommandError(f"No existe el archivo de ajustes: {ajustes_path}")
            try:
                ajustes = json.loads(p.read_text(encoding="utf-8"))
                if not isinstance(ajustes, list):
                    raise ValueError("El JSON de ajustes debe ser una lista de dicts.")
            except Exception as e:
                raise CommandError(f"Error leyendo ajustes JSON: {e}")

        # Selección de sucursales
        suc_qs = Sucursal.objects.all().order_by("pseudonimo")
        if suc_ids:
            suc_qs = suc_qs.filter(id__in=suc_ids)

        # Crear Workbook
        wb = Workbook()
        wb.remove(wb.active)  # quitamos la hoja por defecto
        build_styles(wb)

        resumen_rows = []  # (suc_nombre, hoja, addr_total_neta)

        # Columnas de la grilla
        headers = [
            "Nombre",
            "Rango (Campaña)",
            "DNI",
            "Comisión Bruta",
            # "Descuentos",
            "Comisión Neta",
        ]
        # índices de columnas (1-based)
        COL_NOMBRE = 1
        COL_RANGO = 2
        COL_DNI = 3
        COL_BRUTA = 4
        # COL_DESC = 5 # <-- si se reactiva Descuentos, este pasa a ser 5 y NETA 6
        COL_NETA = 5

        procesadas = 0

        for suc in suc_qs:
            # Participantes con la función OFICIAL
            colaboradores = obtener_usuarios_segun_campana(campania, suc)

            # filtramos administrativos/superuser como en tu vista
            # colaboradores = colaboradores.exclude(
            #     rango__in=["Admin", "Administrativa", "Administrativo"]
            # ).exclude(is_superuser=True)

            colaboradores = [c for c in colaboradores if c.rango not in ["Admin","Administrativa","Administrativo"] and c.is_superuser == False]

            # if not colaboradores.exists() and not options["incluir_vacias"]:
            #     continue

            sheet_title = safe_sheet_title(suc.pseudonimo or f"Sucursal {suc.id}")
            ws = wb.create_sheet(title=sheet_title)

            # Título
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(row=1, column=1, value=f"Campaña: {campania}  |  Sucursal: {suc.pseudonimo} (ID {suc.id})").style = "header"
            ws.row_dimensions[1].height = 24

            # Encabezados
            for i, h in enumerate(headers, start=1):
                c = ws.cell(row=2, column=i, value=h)
                c.style = "header"

            current_row = 3

            for user in colaboradores:
                # Ajustes del usuario para esta (campaña, sucursal) si se pasaron
                ajustes_usuario = []
                if ajustes:
                    ajustes_usuario = [
                        a for a in ajustes
                        if int(a.get("user_id", -1)) == user.pk
                        and a.get("campania") == campania
                        and str(a.get("agencia")) == str(suc.id)
                    ]

                # Cálculo OFICIAL
                data = get_comision_total(user, campania, suc, ajustes_usuario)

                # Rango según snapshot de campaña (respeta tu lógica)
                try:
                    snap = snapshot_usuario_by_campana(user, campania)
                    rango_camp = snap[0].rango if snap else (user.rango or "")
                except Exception:
                    rango_camp = user.rango or ""

                # Escribir fila
                ws.cell(row=current_row, column=COL_NOMBRE, value=getattr(user, "nombre", "")).alignment = Alignment(wrap_text=True)
                ws.cell(row=current_row, column=COL_RANGO, value=str(rango_camp))
                ws.cell(row=current_row, column=COL_DNI, value=getattr(user, "dni", ""))

                c_bruta = ws.cell(row=current_row, column=COL_BRUTA, value=int(data.get("comision_bruta", 0)))
                c_bruta.style = "ars_currency"
                # c_desc = ws.cell(row=current_row, column=COL_DESC, value=int(data.get("descuentoTotal", 0)))
                # c_desc.style = "ars_currency"

                c_neta = ws.cell(row=current_row, column=COL_NETA, value=int(data.get("comision_total", 0)))
                
                c_neta.style = "ars_currency"

                current_row += 1

            last_data_row = current_row - 1

            # Subtotales por rol (SUMIF) – sin acentos/espacios exactos, usa contiene
            # Nota: SUMIF en inglés se traduce al idioma de Excel del usuario automáticamente.
            if last_data_row >= 3:
                # Subtotales por rol (sobre Comisión Neta)
                ws.cell(row=current_row, column=COL_NOMBRE, value="Subtotal Vendedores").style = "sub"
                c_sub_v = ws.cell(
                    row=current_row, column=COL_NETA,
                    value=f'=SUMIF({get_column_letter(COL_RANGO)}3:{get_column_letter(COL_RANGO)}{last_data_row},"*Vendedor*",{get_column_letter(COL_NETA)}3:{get_column_letter(COL_NETA)}{last_data_row})'
                )
                c_sub_v.style = "ars_currency"
                current_row += 1

                ws.cell(row=current_row, column=COL_NOMBRE, value="Subtotal Supervisores").style = "sub"
                c_sub_s = ws.cell(
                    row=current_row, column=COL_NETA,
                    value=f'=SUMIF({get_column_letter(COL_RANGO)}3:{get_column_letter(COL_RANGO)}{last_data_row},"*Supervisor*",{get_column_letter(COL_NETA)}3:{get_column_letter(COL_NETA)}{last_data_row})'
                )
                c_sub_s.style = "ars_currency"
                current_row += 1

                ws.cell(row=current_row, column=COL_NOMBRE, value="Subtotal Gerentes").style = "sub"
                c_sub_g = ws.cell(
                    row=current_row, column=COL_NETA,
                    value=f'=SUMIF({get_column_letter(COL_RANGO)}3:{get_column_letter(COL_RANGO)}{last_data_row},"*Gerente*",{get_column_letter(COL_NETA)}3:{get_column_letter(COL_NETA)}{last_data_row})'
                )
                c_sub_g.style = "ars_currency"
                current_row += 1

                # TOTAL sucursal
                ws.cell(row=current_row, column=COL_NOMBRE, value="TOTAL (Comisión Neta)").style = "total"
                total_neta_cell = ws.cell(
                    row=current_row, column=COL_NETA,
                    value=f'=SUM({get_column_letter(COL_NETA)}3:{get_column_letter(COL_NETA)}{last_data_row})'
                )
                total_neta_cell.style = "ars_currency"

                total_bruta_cell = ws.cell(
                    row=current_row, column=COL_BRUTA,
                    value=f'=SUM({get_column_letter(COL_BRUTA)}3:{get_column_letter(COL_BRUTA)}{last_data_row})'
                )
                total_bruta_cell.style = "ars_currency"

                # Totales de columnas comentadas (si reactivás descuentos, descomentá esta parte y ajustá índices)
                # total_desc_cell = ws.cell(
                #     row=current_row, column=COL_DESC,
                #     value=f'=SUM({get_column_letter(COL_DESC)}3:{get_column_letter(COL_DESC)}{last_data_row})'
                # )
                # total_desc_cell.style = "ars_currency"

                # Guardar referencia para Resumen (toma el total de Neta)
                resumen_rows.append((
                    suc.pseudonimo,
                    sheet_title,
                    f"'{sheet_title}'!{get_column_letter(COL_NETA)}{current_row}",
                ))

            autosize_columns(ws)
            procesadas += 1

        # Hoja Resumen
        wsr = wb.create_sheet(title="Resumen")
        wsr.cell(row=1, column=1, value=f"Campaña: {campania}").style = "header"
        wsr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        wsr.row_dimensions[1].height = 24

        wsr.cell(row=2, column=1, value="Sucursal").style = "header"
        wsr.cell(row=2, column=2, value="Total Comisión Neta").style = "header"
        wsr.cell(row=2, column=3, value="Hoja").style = "header"

        r = 3
        for nombre, hoja, addr in resumen_rows:
            wsr.cell(row=r, column=1, value=nombre)
            c_total = wsr.cell(row=r, column=2, value=f"={addr}")
            c_total.number_format = '#,##0'
            c_total.style = "ars_currency"
            wsr.cell(row=r, column=3, value=hoja)
            r += 1

        if resumen_rows:
            wsr.cell(row=r, column=1, value="TOTAL CAMPAÑA").style = "total"
            c_gran_total = wsr.cell(row=r, column=2, value=f'=SUM(B3:B{r-1})')
            c_gran_total.style = "ars_currency"

        autosize_columns(wsr)

        # Nombre de salida
        default_name = f"export_comisiones_{campania.replace(' ', '_')}.xlsx"
        out_path = Path(options.get("output") or default_name).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)

        self.stdout.write(self.style.SUCCESS(
            f"Exportación OK. Archivo generado: {out_path}"
        ))
        if not procesadas:
            self.stdout.write(self.style.WARNING(
                "No se generó ninguna hoja (¿no hay colaboradores en esa campaña/sucursales?)."
            ))


