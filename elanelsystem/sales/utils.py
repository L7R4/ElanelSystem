import random
import string
from django.template.loader import get_template
from weasyprint import HTML,CSS
import elanelsystem.settings as settings
import os
import pandas as pd
import logging
logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG)


import datetime
from django.contrib.auth.models import Group
from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.db.models import Max
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db import transaction
from users.models import Cliente, Usuario
from django.templatetags.static import static
from django.db.models.functions import Replace, Lower
from django.db.models import Value

from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from dateutil.relativedelta import relativedelta
from num2words import num2words
from configuracion.models import Configuracion
from products.models import Products, Plan

#region Funciones para exportar PDFs
def printPDFBaja(data,url,productoName):
    template = get_template("pdfForBaja.html")
    html_template = template.render(data)
    css_url = os.path.join(settings.BASE_DIR, "static/css/pdfForBaja.css")
    if not os.path.exists(settings.PDF_STORAGE_DIR):
        os.makedirs(settings.PDF_STORAGE_DIR)
    HTML(string=html_template,base_url=url).write_pdf(target=productoName, stylesheets = [CSS(css_url)])
    

def printPDFtitularidad(data,url,productoName):
    template = get_template("pdfForTitu.html")
    context = data
    html_template = template.render(context)
    css_url = os.path.join(settings.BASE_DIR, "static/css/pdfForTitu.css")
    if not os.path.exists(settings.PDF_STORAGE_DIR):
        os.makedirs(settings.PDF_STORAGE_DIR)
    pdf = HTML(string=html_template,base_url=url).write_pdf(target=productoName, stylesheets = [CSS(css_url)])
    # print(pdf)
    return pdf

def printPDFarqueo(data,url,productoName):
    template = get_template("pdfForArqueo.html")
    context = data
    html_template = template.render(context)
    css_url = os.path.join(settings.BASE_DIR, "static/css/pdfArqueo.css")
    if not os.path.exists(settings.PDF_STORAGE_DIR):
        os.makedirs(settings.PDF_STORAGE_DIR)
    pdf = HTML(string=html_template,base_url=url).write_pdf(target=productoName, stylesheets = [CSS(css_url)])
    # print(pdf)
    return pdf

def printPDFinforme(data,url,productoName):
    template = get_template("pdfForInforme.html")
    context = data
    html_template = template.render(context)
    css_url = os.path.join(settings.BASE_DIR, "static/css/pdfInforme.css")
    if not os.path.exists(settings.PDF_STORAGE_DIR):
        os.makedirs(settings.PDF_STORAGE_DIR)
    pdf = HTML(string=html_template,base_url=url).write_pdf(target=productoName, stylesheets = [CSS(css_url)])
    # print(pdf)
    return pdf

def printPDFinformePostVenta(data,url,productoName):
    template = get_template("pdfPostVentaInforme.html")
    context = data
    html_template = template.render(context)
    css_url = os.path.join(settings.BASE_DIR, "static/css/pdfPostVentaInforme.css")
    if not os.path.exists(settings.PDF_STORAGE_DIR):
        os.makedirs(settings.PDF_STORAGE_DIR)
    pdf = HTML(string=html_template,base_url=url).write_pdf(target=productoName, stylesheets = [CSS(css_url)])
    # print(pdf)
    return pdf


def sendEmailPDF(email,pdf_path,sujeto):
        try:
            subject = sujeto
            from_email =  settings.EMAIL_HOST_USER
            to_email = [email]

            email = EmailMultiAlternatives(subject, '', from_email, to_email)
            email.attach_alternative("Se adjunta el PDF.", "text/plain")
            
            # Adjuntar el PDF al correo electrónico
            email.attach_file(pdf_path)

            # Enviar el correo electrónico
            email.send()
            response_data = {'success': True}
            print("weps se envio")
        except Exception as e:
            response_data = {'success': False, 'error_message': str(e)}
            print(e)

        return response_data
#endregion ---------------------------------------------


#region Funciones para exportar .XLSs
def exportar_excel(data):
    information = data["data"]
    tipo_information = data["tipo"]
    fecha = datetime.datetime.now().strftime("%d-%m-%Y %H-%M")

    # Crear un nuevo libro de trabajo
    wb = Workbook()

    # Seleccionar la primera hoja (automáticamente creada)
    ws = wb.active
    ws.title = "Datos"

    # Estilo para los encabezados
    bold_font = Font(color="FFFFFF", size=13, name="Berlin Sans FB")  # Letra blanca, más grande, y fuente específica
    yellow_fill = PatternFill(start_color="1753ED", end_color="1753ED", fill_type="solid")
    center_alignment = Alignment(horizontal='center')
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Insertar imagen antes de la fila 1
    image_path = os.path.join(settings.BASE_DIR, "static/images/logoElanelPDF.png")
    img = Image(image_path)  # Cargar la imagen desde la ruta proporcionada
    img.width, img.height = 830, 90  # Ajustar tamaño de la imagen
    ws.add_image(img, "A1")  # Insertar imagen en la celda A1

    # Espaciar la primera fila para la imagen
    ws.row_dimensions[1].height = 130

    # Iniciar los encabezados en la fila 3
    if information:
        encabezados = [v['verbose_name'] for v in information[0].values()]
        ws.append([])  # Fila vacía para mantener la separación
        ws.append(encabezados)  # Agregar encabezados en la fila 3

        # Aplicar estilos a los encabezados (fila 3)
        for cell in ws[2]:
            cell.font = bold_font
            cell.fill = yellow_fill
            cell.alignment = center_alignment
            cell.border = thin_border

        # Agregar datos de cada diccionario a partir de la fila 4
        for item in information:
            fila = [v['data'] for v in item.values()]
            ws.append(fila)

        # Establecer la fuente general para todo el Excel
        berlin_font = Font(name="Berlin Sans FB")
        for row in ws.iter_rows(min_row=3):  # Desde la fila 4
            for cell in row:
                cell.font = berlin_font

    # Ajustar el ancho de las columnas basado en los encabezados (fila 3)
    for idx, cell in enumerate(ws[3], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = len(str(cell.value) if cell.value is not None else "") + 15


    # Configurar la respuesta como archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="datos_{tipo_information}_{fecha}.xlsx"'

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response

#endregion 


#region Funciones para obtener campañas
def getCampaniaActual():
    list_mesesStr = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes_actual = datetime.datetime.now().month
    anio_actual = datetime.datetime.now().year
    return f'{list_mesesStr[mes_actual-1]} {anio_actual}'


def getAllCampaniaOfYear():
    list_mesesStr = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    anio_actual = datetime.datetime.now().year
    meses_formato = [f'{mes} {anio_actual}' for mes in list_mesesStr]
    return meses_formato

# Obtener la campania de acuerdo a la fecha que se pasa por parametro
def getCampaniaByFecha(fecha):
    list_mesesStr = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    mes = fecha.month
    anio = fecha.year
    return f'{list_mesesStr[mes-1]} {anio}'

# Para obtener todas las compañas desde el inicio de la empresa (2021)
def getTodasCampaniasDesdeInicio():
    list_mesesStr = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    anio_actual = datetime.datetime.now().year
    anio_inicio = 2021
    campanias = []
    
    # Iterar desde el año actual hasta el año de inicio (descendente)
    for anio in range(anio_actual, anio_inicio - 1, -1):
        # Para cada año, agregar las campañas en orden
        campanias.extend([f'{mes} {anio}' for mes in list_mesesStr])
    
    return campanias


# """
# Retorna las campañas disponibles segun las cantidad de dias.
# Ya que si hay mas de 5 dias de diferencia entre la campaña actual con la anterior no se
# habilita la campaña actual, pero sino, si.
# """
def getCampanasDisponibles():
    campaniasDelAño = getAllCampaniaOfYear()
    campaniaActual = getCampaniaActual()
    campaniaAnterior = campaniasDelAño[campaniasDelAño.index(campaniaActual) - 1]


    fechaActual = datetime.datetime.now()
    ultimo_dia_mes_pasado = datetime.datetime.now().replace(day=1) - relativedelta(days=1)
    diferencia_dias = (fechaActual - ultimo_dia_mes_pasado).days
    if(diferencia_dias > 5): # Si la diferencia de dias es mayor a 5 dias, no se puede asignar la campania porque ya paso el tiempo limite para dar de alta una venta en la campania anterior
        return [campaniaActual]
    else:
        return [campaniaActual,campaniaAnterior]

#endregion


#region Funciones para formatear monedas
def formatear_moneda_sin_centavos(valor):
    try:
        valor = float(valor)
        return f"{valor:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "-"
    
def formatear_moneda_con_centavos(valor):
    try:
        valor = float(valor)
        return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "-"

def convertir_moneda_a_texto(cantidad):
    # Eliminar símbolos y puntos
    cantidad = str(cantidad)
    cantidad_limpia = cantidad.replace('$', '').replace('.', '').replace(',', '')
    # Convertir a entero
    numero = int(cantidad_limpia)
    # Convertir número a palabras en español
    texto = num2words(numero, lang='es')
    # Retornar el texto con la palabra "pesos"
    return f"{texto} pesos".capitalize()

# endregion


#region Data Structures ----------------------------------------------------------

def getInfoBaseCannon(venta, cuota):
    cliente = Cliente.objects.get(id=venta.nro_cliente.id)
    return {
        'cuota': {'data': cuota["cuota"], 'verbose_name': 'Cuota'},
        'nro_operacion': {'data': cuota["nro_operacion"], 'verbose_name': 'N° Operación'},
        'modalidad': {'data': venta.modalidad, 'verbose_name': 'Modalidad'},
        'campania': {'data': venta.campania, 'verbose_name': 'Campaña de venta'},
        'contratos': {'data': venta.cantidadContratos, 'verbose_name': 'Cantidad de Contratos'},
        "fecha_inscripcion": {'data': venta.fecha, 'verbose_name': 'Fecha Inscripción'},
        'nombre_del_cliente': {'data': cliente.nombre, 'verbose_name': 'Nombre del Cliente'},
        'nro_del_cliente': {'data': cliente.nro_cliente, 'verbose_name': 'N° del Cliente'},
        'agencia': {'data': venta.agencia.pseudonimo, 'verbose_name': 'Agencia'},
        'tipo_mov': {'data': "ingreso", 'verbose_name': 'Tipo de Movimiento'},
        'fecha_de_vencimiento': {'data': cuota['fechaDeVencimiento'], 'verbose_name': 'Fecha de Vencimiento'},
        'descuento': {'data': cuota['descuento'], 'verbose_name': 'Descuento'},
        'estado': {'data': cuota['status'], 'verbose_name': 'Estado'},
        'dias_de_mora': {'data': cuota['diasRetraso'], 'verbose_name': 'Días de Mora'},
        'cuota_comercial': {'data': cuota.get('total', 0), 'verbose_name': 'Cuota comercial'},
        'total_final': {'data': cuota.get('totalFinal', 0), 'verbose_name': 'Total Final'},
        'interes_por_mora': {'data': cuota.get('interesPorMora', 0), 'verbose_name': 'Interés por Mora'},
        'vendedor': {'data': venta.vendedor.nombre, 'verbose_name': 'Vendedor'},
        'supervisor': {'data': venta.supervisor.nombre, 'verbose_name': 'Supervisor'},

    }


def dataStructureCannons(sucursal=None):
    from sales.models import Ventas, MetodoPago, CuentaCobranza
    from elanelsystem.views import convertirValoresALista
    from elanelsystem.utils import formatear_dd_mm_yyyy_h_m
    
    if sucursal:
        ventas = get_ventasBySucursal(sucursal)
    else:
        ventas = Ventas.objects.all()

    cuotas_data = []

    for i in range(int(ventas.count())):
        venta = ventas[i]

        for k in range(len(venta.cuotas)):
            cuota = venta.cuotas[k]
            cuotaStatusCleaned = cuota["status"].replace(" ", "").lower()
            if cuotaStatusCleaned in ["pagado", "parcial", "atrasado"]:
                for pago in cuota["pagos"]:
                    metodo_pago_obj = get_or_create_metodo_pago(pago['metodoPago'])
                    cobrador_obj = get_or_create_cobrador(pago["cobrador"])
                    dias_diferencia = ""
                    if pago["fecha"] and venta.fecha:
                        try:
                            # Parsear fecha de pago
                            fecha_pago = datetime.datetime.strptime(formatear_dd_mm_yyyy_h_m(pago["fecha"]), "%d/%m/%Y %H:%M").date()

                            # Parsear fecha de inscripción (venta.fecha)
                            fecha_inscripcion = datetime.datetime.strptime(formatear_dd_mm_yyyy_h_m(venta.fecha), "%d/%m/%Y %H:%M").date()

                            # Calcular diferencia
                            dias_diferencia = (fecha_pago - fecha_inscripcion).days
                        except ValueError as e:
                            print(f"Error al parsear fechas: {e}")
                            dias_diferencia = ""

                    mov = {**getInfoBaseCannon(venta, cuota), **{
                        'metodoPago': {'data': str(metodo_pago_obj.id), 'verbose_name': 'Método de Pago'},
                        'metodoPagoAlias': {'data': metodo_pago_obj.alias, 'verbose_name': 'Método de Pago'},
                        'monto': {'data': pago['monto'], 'verbose_name': 'Monto Pagado'},
                        'montoFormated': {'data': formatear_moneda_sin_centavos(pago['monto']), 'verbose_name': 'Monto Pagado'},
                        'fecha': {'data': pago['fecha'], 'verbose_name': 'Fecha de Pago'},
                        'cobrador': {'data': str(cobrador_obj.id), 'verbose_name': 'Cobrador'},
                        'cobradorAlias': {'data': cobrador_obj.alias, 'verbose_name': 'Cobrador'},
                        'campaniaPago': {'data': pago["campaniaPago"], 'verbose_name': 'Campaña de pago'},
                        'dias_de_diferencia': {'data': dias_diferencia,'verbose_name': 'Días de diferencia'}

                    }}
                    cuotas_data.append(mov)
            else:
                mov = {**getInfoBaseCannon(venta, cuota), **{
                    'metodoPago': {'data': "", 'verbose_name': 'Método de Pago'},
                    'monto': {'data': "", 'verbose_name': 'Monto Pagado'},
                    'fecha': {'data': "", 'verbose_name': 'Fecha de Pago'},
                    'cobrador': {'data': "", 'verbose_name': 'Cobrador'},
                }}
                cuotas_data.append(mov)

    # cuotas_data.reverse()
    # Obtener las cuotas con nro_operacion == 597
    # w = [cuota for cuota in cuotas_data if cuota['nro_operacion']['data'] == "597"]
    # print(f"\n\n\n {type(w[0]['metodoPago']['data'])} \n\n\n")
    return cuotas_data


def dataStructureVentas(sucursal=None):
    from sales.models import Ventas

    if sucursal:
        ventas = get_ventasBySucursal(sucursal)
    else:
        ventas = Ventas.objects.all()
        
    ventasList = []
    for i in range(int(ventas.count())):
        venta = ventas[i]
        dineroEntregado = getDineroEntregado(venta.cuotas)
        ventaDict = {
            'nro_operacion': {'data': venta.nro_operacion, 'verbose_name': 'N° ope'},
            'fecha': {'data': venta.fecha, 'verbose_name': 'Fecha de inscripción'},
            'nro_cliente': {'data': venta.nro_cliente.nro_cliente, 'verbose_name': 'N° cliente'},
            'nombre_de_cliente': {'data': venta.nro_cliente.nombre, 'verbose_name': 'Nombre del cliente'},
            'agencia': {'data': venta.agencia.pseudonimo, 'verbose_name': 'Agencia'},
            'nro_cuotas': {'data': venta.nro_cuotas, 'verbose_name': 'N° cuotas'},
            'id': {'data': venta.id, 'verbose_name': 'ID'},
            'campania': {'data': venta.campania, 'verbose_name': 'Campaña'},

            'importe': {'data': venta.importe, 'verbose_name': 'Importe'},
            'importe_formated': {'data': f"${formatear_moneda_sin_centavos(venta.importe)}", 'verbose_name': 'Importe'},

            'paquete': {'data': venta.paquete, 'verbose_name': 'Paquete'},
            
            'primer_cuota': {'data': venta.primer_cuota, 'verbose_name': 'Primer cuota'},
            'primer_cuota_formated': {'data': f"${formatear_moneda_sin_centavos(venta.primer_cuota)}", 'verbose_name': 'Primer cuota'},
            
            'suscripcion': {'data': venta.anticipo, 'verbose_name': 'Suscripción'},
            'suscripcion_formated': {'data': f"${formatear_moneda_sin_centavos(venta.anticipo)}", 'verbose_name': 'Suscripción'},

            'cuota_comercial': {'data': venta.importe_x_cuota, 'verbose_name': 'Cuota comercial'},
            'cuota_comercial_formated': {'data': f"${formatear_moneda_sin_centavos(venta.importe_x_cuota)}", 'verbose_name': 'Cuota comercial'},
            
            'interes_generado': {'data': venta.intereses_generados, 'verbose_name': 'Interés generado'},
            'interes_generado_formated': {'data': f"${formatear_moneda_sin_centavos(venta.intereses_generados)}", 'verbose_name': 'Interés generado'},
            
            'total_a_pagar': {'data': venta.total_a_pagar, 'verbose_name': 'Total a pagar'},
            'total_a_pagar_formated': {'data': f"${formatear_moneda_sin_centavos(venta.total_a_pagar)}", 'verbose_name': 'Total a pagar'},

            'dinero_entregado': {'data': dineroEntregado, 'verbose_name': 'Dinero entregado'},
            'dinero_entregado_formated': {'data': f"${formatear_moneda_sin_centavos(dineroEntregado)}", 'verbose_name': 'Dinero entregado'},
            
            'dinero_restante': {'data': venta.total_a_pagar - dineroEntregado, 'verbose_name': 'Dinero restante'},
            'dinero_restante_formated': {'data': f"${formatear_moneda_sin_centavos(venta.total_a_pagar - dineroEntregado)}", 'verbose_name': 'Dinero restante'},
            
            "cantidad_chances":{'data': len(venta.cantidadContratos), 'verbose_name': 'Cantidad de chances'},
            "total_por_contrato":{'data': round(venta.importe / len(venta.cantidadContratos),0), 'verbose_name': 'Importe por contrato'},

            'vendedor': {'data': venta.vendedor.nombre, 'verbose_name': 'Vendedor'},
            'producto': {'data': venta.producto.nombre, 'verbose_name': 'Producto'},
            'supervisor': {'data': venta.supervisor.nombre, 'verbose_name': 'Supervisor'},
        }

        # Agregar el campo si la venta esta AUDITADA  
        if len(venta.auditoria) > 0:   
            last_auditoria = venta.auditoria[-1]
            ventaDict["auditada"] = {'data': "Si", 'verbose_name': 'Auditada'}
            ventaDict["ultima_auditoria"] = {'data': last_auditoria["grade"], 'verbose_name': 'Última auditoría'}

        # Agregar el campo si la venta esta ADJUDICADA
        if(venta.adjudicado["status"]):
            ventaDict["adjudicado"] = {'data': "Si", 'verbose_name': 'Adjudicado'}
            ventaDict["tipo_de_adjudicacion"] = {'data': venta.adjudicado["tipo"], 'verbose_name': 'Tipo de adjudicación'}

        # Agregar el campo si la venta esta DE BAJA
        if(venta.deBaja["status"]):
            ventaDict["de_baja"] = {'data': "Si", 'verbose_name': 'De baja'}
            ventaDict["motivo_de_baja"] = {'data': venta.deBaja["motivo"] , 'verbose_name': 'Motivo de baja'}
            ventaDict["responsable"] = {'data': venta.deBaja["responsable"], 'verbose_name': 'Responsable'}
        
        # Agregar el campo si la venta esta SUSPENDIDA
        if(venta.suspendida):
            ventaDict["suspendida"] = {'data': "Si", 'verbose_name': 'Suspendida'}

        ventaDict['observaciones'] = {'data': venta.observaciones, 'verbose_name': 'Observaciones'}

        ventasList.append(ventaDict)

    return ventasList


def dataStructureMovimientosExternos(sucursal=None):
    from sales.models import MovimientoExterno
    from elanelsystem.views import convertirValoresALista

    movs_externos = ""
    
    if sucursal:
        listaAgencias = convertirValoresALista({"agencia":sucursal})["agencia"]
        # print(f"Listas de agencias: {listaAgencias}")
        movs_externos = MovimientoExterno.objects.filter(agencia__pseudonimo__in=listaAgencias)
        # print(f"MOvimientos: {movs_externos[0]}")
    
    else:
        movs_externos = MovimientoExterno.objects.all()
        
    movsExternosList = []
    for movs_externo in movs_externos:
        movsExternoDict = {
            "tipoIdentificacion": {'data': movs_externo.tipoIdentificacion, 'verbose_name': 'Tipo Identificación'},
            "nroIdentificacion": {'data': movs_externo.nroIdentificacion, 'verbose_name': 'N° Identificación'},
            "tipoComprobante": {'data': movs_externo.tipoComprobante, 'verbose_name': 'Tipo Comprobante'},
            "nroComprobante": {'data': movs_externo.nroComprobante, 'verbose_name': 'N° Comprobante'},
            "denominacion": {'data': movs_externo.denominacion, 'verbose_name': 'Denominación'},
            "tipoMoneda": {'data': movs_externo.tipoMoneda, 'verbose_name': 'Tipo Moneda'},
            "tipo_mov": {'data': movs_externo.movimiento, 'verbose_name': 'Tipo Movimiento'},
            "monto": {'data': movs_externo.dinero, 'verbose_name': 'Monto'},
            'montoFormated': {'data': formatear_moneda_sin_centavos(movs_externo.dinero), 'verbose_name': 'Monto Pagado'},
            
            "metodoPago": {'data': str(movs_externo.metodoPago.id), 'verbose_name': 'Método de Pago'},
            "metodoPagoAlias": {'data': movs_externo.metodoPago.alias, 'verbose_name': 'Método de Pago'},
            "agencia": {'data': movs_externo.agencia.pseudonimo, 'verbose_name': 'Sucursal'},
            "ente": {'data': movs_externo.ente.alias, 'verbose_name': 'Ente recaudador'},
            "fecha": {'data': movs_externo.fecha, 'verbose_name': 'Fecha'},
            "campania": {'data': movs_externo.campania, 'verbose_name': 'Campaña'},
            "concepto": {'data': movs_externo.concepto, 'verbose_name': 'Concepto'},
            "observaciones": {'data': movs_externo.observaciones, 'verbose_name': 'Observaciones'},
            "premio": {'data':  "Si" if movs_externo.premio else "No", 'verbose_name': 'Premio'},
            "adelanto": {'data': "Si" if movs_externo.adelanto else "No", 'verbose_name': 'Adelanto'},
        }
        
        movsExternosList.append(movsExternoDict)
    print(movsExternosList)
    return movsExternosList


def dataStructureMoviemientosYCannons(sucursal):
    structureCannons = dataStructureCannons(sucursal)
    structureMovsExternos = dataStructureMovimientosExternos(sucursal)
    structureMovimientos_Generales = structureCannons + structureMovsExternos

    # Logica para colocar un ID de indice a cada movimiento
    id_cont = 0
    for mov in structureMovimientos_Generales:
        id_cont += 1
        mov["id_cont"] = id_cont

    return structureMovimientos_Generales


def dataStructureClientes(sucursal=None):
    from users.models import Cliente
    from elanelsystem.views import convertirValoresALista

    clientes = ""
    
    if sucursal:
        listaAgencias = convertirValoresALista({"agencia":sucursal})["agencia"]
        clientes = Cliente.objects.filter(agencia_registrada__pseudonimo__in=listaAgencias)
    
    else:
        clientes = Cliente.objects.all()
        
    clienteList = []
    for cliente in clientes:
        clienteDict = {
            "nro_cliente": {'data': cliente.nro_cliente, 'verbose_name': 'N° Cli'},
            "nombre": {'data': cliente.nombre, 'verbose_name': 'Nombre'},
            "dni": {'data': cliente.dni, 'verbose_name': 'DNI'},
            "domic": {'data': cliente.domic, 'verbose_name': 'Domicilio'},
            "loc": {'data': cliente.loc, 'verbose_name': 'Localidd'},
            "prov": {'data': cliente.prov, 'verbose_name': 'Provincia'},
            "cod_postal": {'data': cliente.cod_postal, 'verbose_name': 'Codigo Postal'},
            "tel": {'data': cliente.tel, 'verbose_name': 'Telefono'},
            "fec_nacimiento": {'data': cliente.fec_nacimiento, 'verbose_name': 'Fecha de Nacimiento'},
            "agencia": {'data': cliente.agencia_registrada.pseudonimo, 'verbose_name': 'Agencia registrada'},
            "estado_civil": {'data': cliente.estado_civil, 'verbose_name': 'Estado civil'},
            "ocupacion": {'data': cliente.ocupacion, 'verbose_name': 'Ocupacion'},
        }
        
        clienteList.append(clienteDict)
    
    return clienteList
#endregion


#region Other functions

def get_ventasBySucursal(sucursal):
    from sales.models import Ventas
    from elanelsystem.views import convertirValoresALista

    if sucursal == "":
        return Ventas.objects.all()
    else:
        listaAgencias = convertirValoresALista({"agencia": sucursal})["agencia"]
        return Ventas.objects.filter(agencia__id__in=listaAgencias)


def deleteFieldsInDataStructures(lista_dicts, campos_a_eliminar):
    # Iterar sobre cada diccionario en la lista
    for item in lista_dicts:
        # Eliminar los campos especificados si existen en el diccionario
        for campo in campos_a_eliminar:
            if campo in item:
                del item[campo]
    return lista_dicts


def getEstadoVenta(venta):
    if(venta.deBaja["status"]):
        return f"De baja por {venta.deBaja['motivo']}"
    elif(venta.suspendida):
        return "Suspendida"
    elif(venta.adjudicado["status"]):
        return f"Adjudicada por {venta.adjudicado['tipo']}"
    else:
        return "Activo"
    

def getEstadoVenta2(venta):
    if(venta.deBaja["status"]):
        return {'status': 'Baja', 'motivo': venta.deBaja['motivo']}
    elif(venta.suspendida):
        return {'status': 'Suspendida', 'motivo': ""}
    elif(venta.adjudicado["status"]):
        return {'status': 'Adjudicada', 'motivo': ""}
    else:
        return {'status': 'Activa', 'motivo': ""}

def getCuotasPagadasSinCredito(cuotas):
    from sales.models import PagoCannon

    cuotasPagadasSinCredito = []
    for cuota in cuotas:
        pagos = cuota["pagos"]
        metodoDePagoDeCuota = [PagoCannon.objects.get(id=pago).metodo_pago.alias for pago in pagos] # Obtiene los metodos de pago de la cuota
        # Me aseguro que si se aplico un credito a una cuota y no fue el unico metodo de pago, se tome como ultima cuota porque quiere decir que el cliente abono un monto de la cuota
        if("Credito" in metodoDePagoDeCuota and len(metodoDePagoDeCuota) != 1):
            cuotasPagadasSinCredito.append(cuota)
        # Si la cuota no tiene credito, se toma como ultima cuota
        elif ("Credito" not in metodoDePagoDeCuota):
            cuotasPagadasSinCredito.append(cuota)
    return cuotasPagadasSinCredito


def bloquer_desbloquear_cuotas(cuotas):
    from sales.models import PagoCannon

    nuevas_cuotas = []
    for i in range(0,len(cuotas)):
        cuota = cuotas[i]
        if not (i == 0):
            pagos = cuota["pagos"]
            metodoDePagoDeCuota = [PagoCannon.objects.get(id=pago).metodo_pago.alias for pago in pagos] # Obtiene los metodos de pago de la cuota
            if(cuotas[i-1]["status"] == "Pagado"):
                cuota["bloqueada"] = False
            elif("Credito" in metodoDePagoDeCuota):
                cuota["bloqueada"] = False
            else:
                cuota["bloqueada"] = True
        nuevas_cuotas.append(cuota)
    return nuevas_cuotas


def getDineroEntregado(cuotas):
    dineroEntregado = 0
    for cuota in cuotas:
        pagos = cuota["pagos"]
        if(len(pagos)>0):
            for pago in pagos:
                dineroEntregado += pago["monto"]
    return dineroEntregado


def obtenerStatusAuditoria(venta): # Devuelve el estado de la ultima auditoria

    #Verifica si la lista de auditorías está vacía
    if len(venta.auditoria) == 0:
        return {"statusText": "Pendiente", "statusIcon": static(f"/images/icons/operationSuspendido.svg")}  # No auditada
    
    # Obtiene la última auditoría
    ultima_auditoria = venta.auditoria[-1]
    
    # Verifica el estado de la última auditoría
    if ultima_auditoria.get("grade") is True:
        return {"statusText": "Aprobada", "statusIcon": static(f"images/icons/operationActivo.svg")}  # No auditada

    elif ultima_auditoria.get("grade") is False:
        return {"statusText": "Desaprobada", "statusIcon": static(f"images/icons/operationBaja.svg")}  # No auditada

#endregion


#region Para enviar correos electrónicos

def send_html_email(subject, template, context, from_email, to_email):
    """
    Envía un correo electrónico en formato HTML.
    """
    html_message = render_to_string(template, context)  # Renderizar el template con contexto
    plain_message = strip_tags(html_message)  # Generar versión en texto plano

    email = EmailMessage(subject, html_message, from_email, [to_email])
    email.content_subtype = "html"  # Asegurar que se envía como HTML
    email.send()
#endregion


#region Funciones para la importacion en excel
def reportar_nans(df, cols, id_field = 'id_venta'):
    """
    Recorre las columnas `cols` de `df` y devuelve una lista de dicts
    con {'index', 'id_venta', 'column'} para cada NaN.
    """
    errores = []
    for col in cols:
        # detectamos NaN reales o cadenas vacías tras strip()
        mask = df[col].isna() | df[col].astype(str).str.strip().eq('')
        for idx in df[mask].index:
            errores.append({
                'index': idx,
                'id_venta': df.at[idx, id_field],
                'column': col,
                'valor': df.at[idx, col]
            })
    return errores


def preprocesar_excel_ventas(file_path):
    from elanelsystem.utils import obtenerCampaña_atraves_fecha,formatar_fecha

    # Leer la hoja del archivo Excel
    df_res = pd.read_excel(file_path, sheet_name="RESUMEN")
    df_est = pd.read_excel(file_path, sheet_name="ESTADOS")
    
    # Renombrar las columnas
    df_res.columns = [
        col.strip().lower().replace(" ", "_").replace("-", "_").replace(".", "_").replace("/", "_")
        for col in df_res.columns
    ]

    df_est.columns = [
        col.strip().lower().replace(" ", "_").replace("-", "_").replace(".", "_").replace("/", "_")
        for col in df_est.columns
    ]

    campos_int_res = ['importe','tasa_de_inte','id_venta','cod_cli','producto','vendedor','superv']
    errores_res = reportar_nans(df_res, campos_int_res, id_field='id_venta')
    if errores_res:
        # Aquí podrías loguearlo, devolverlo en la respuesta JSON, o lanzar excepción
        raise ValueError(f"⁉️ Errores en RESUMEN antes de conversión ⁉️:\n{errores_res}")

    df_res['id_venta'] = df_res['id_venta'].astype(str)
    df_res['cod_cli'] = df_res['cod_cli'].astype(str)
    df_res['importe'] = df_res['importe'].astype(int)
    df_res['modalidad'] = df_res['modalidad'].astype(str)
    df_res['tasa_de_inte'] = df_res['tasa_de_inte'].astype(float)
    df_res['fecha_incripcion'] = df_res['fecha_incripcion'].astype(str)
    df_res['paq'] = df_res['paq'].fillna('').map(lambda x: 'Basico' if x=='BASE' else x.capitalize())
    df_res['vendedor_raw'] = (df_res['vendedor'].fillna('').str.title().str.replace(r'\s+', ' ', regex=True).str.strip())
    df_res['superv_raw'] = (df_res['superv'].fillna('').str.title().str.replace(r'\s+', ' ', regex=True).str.strip())
    # Y para facilitar el *groupby* (que no admite raw con espacios):
    df_res['vendedor_key'] = df_res['vendedor_raw'].str.lower().str.replace(' ', '')
    df_res['superv_key']  = df_res['superv_raw'].str.lower().str.replace(' ', '')
    
    df_res['producto_raw'] = (df_res['producto'].fillna('').astype(str).str.strip().str.replace(r'\s+', ' ', regex=True).str.title())
    df_res['producto_key'] = (df_res['producto_raw'].str.lower().str.replace(' ', '', regex=False))

    df_res['comentarios__observaciones'] = df_res['comentarios__observaciones'].fillna('')

    campos_int_est = ['importe_cuotas', 'cuotas','estado','fecha_venc']
    errores_est = reportar_nans(df_est, campos_int_est, id_field='id_venta')
    if errores_est:
        raise ValueError(f"⁉️ Errores en ESTADOS antes de conversión ⁉️:\n{errores_est}")

    # Preparamos ESTADOS
    df_est['id_venta']     = df_est['id_venta'].astype(int)
    df_est['importe_cuotas']= df_est['importe_cuotas']\
        .replace('[\$,]', '', regex=True).astype(int)
    df_est['cuota_num']    = df_est['cuotas']\
        .str.extract(r'(\d+)').astype(int)
    df_est['estado_norm']  = df_est['estado'].str.title()
    df_est['fecha_de_venc']= df_est['fecha_venc'].astype(str).apply(lambda d: formatar_fecha(d))
    df_est['fecha_de_pago']= df_est['fecha_de_pago'].astype(str).apply(lambda d: formatar_fecha(d))
    df_est['campania_pago']= df_est['fecha_de_pago'].apply(lambda d: obtenerCampaña_atraves_fecha(formatar_fecha(d)) if formatar_fecha(d) else "")
    
    return df_res, df_est


def normalize_key(s):
    """Minúsculas y sin espacios para comparar claves."""
    return ''.join(s.lower().split())


def get_or_create_product_from_import(raw_name, importe):
    """
    1) Normaliza raw_name (minúsculas, sin espacios) → norm_raw.
    2) Busca un Products con nombre_norm = Lower(Replace(nombre, ' ', '')) == norm_raw.
    3) Si no existe, maneja casos especiales o genéricos creando el producto asociado al Plan.
    """
    norm_raw = normalize_key(raw_name)

    # 1) Intento encontrar raw_name en BD (ignorando mayúsculas y espacios)
    prod = Products.objects.annotate(
        nombre_norm=Lower(Replace('nombre', Value(' '), Value('')))
    ).filter(nombre_norm=norm_raw).first()
    if prod:
        return prod

    # 2) Casos especiales: Solucion Dineraria / Valor nominal
    if norm_raw in ('soluciondineraria', 'valornominal', 'valoresnominales','solucionesdinerarias', 'solucionesdineraria'):
        price_name = f'${int(importe)}'
        norm_price = normalize_key(price_name)

        prod = Products.objects.annotate(
            nombre_norm=Lower(Replace('nombre', Value(' '), Value('')))
        ).filter(nombre_norm=norm_price).first()
        if prod:
            return prod

        # Necesitamos un Plan existente con ese valor_nominal
        plan = Plan.objects.filter(valor_nominal=int(importe)).first()
        if not plan:
            raise ValueError(f"No existe un Plan con valor_nominal={int(importe)}")

        with transaction.atomic():
            prod = Products.objects.create(
                nombre           = price_name,
                tipo_de_producto = 'Solucion',
                tipodePlan       = 'Estandar',
                plan             = plan
            )
        return prod

    # 3) Caso genérico: crear con raw_name y Plan(valor_nominal=importe)
    plan = Plan.objects.filter(valor_nominal=int(importe)).first()
    if not plan:
        raise ValueError(f"No existe un Plan con valor_nominal={int(importe)}")

    with transaction.atomic():
        prod = Products.objects.create(
            nombre           = raw_name,
            tipo_de_producto = 'Moto',
            tipodePlan       = 'Estandar',
            plan             = plan
        )
    return prod


def get_or_create_usuario_from_import(raw_name, tipo, sucursal_obj):
    """
    1) Busca un Usuario por nombre (ignorando mayúsculas y espacios).
    2) Si existe, se asegura de añadirlo a la sucursal y lo devuelve.
    3) Si no existe, genera:
       - un DNI aleatorio de 9 dígitos único,
       - email = <nombre_normalizado>@gmail.com (añade sufijo si hace falta para ser único),
       - password = ese DNI,
       - rango = 'Vendedor' o 'Supervisor' según `tipo`,
       - lo guarda y lo añade a sucursales.
    """
    norm = normalize_key(raw_name)

    # 1) Intento encontrarlo en BD
    user = Usuario.objects.annotate(
        nombre_norm=Lower(Replace('nombre', Value(' '), Value('')))
    ).filter(nombre_norm=norm,sucursales=sucursal_obj.pk).first()

    if user:
        if not user.sucursales.filter(pk=sucursal_obj.pk).exists():
            user.sucursales.add(sucursal_obj)
        return user

    # 2) Generar un DNI aleatorio único de 9 dígitos
    while True:
        dni = ''.join(random.choices(string.digits, k=9))
        if not Usuario.objects.filter(dni=dni).exists():
            break

    # 3) Email basado en nombre, garantizando unicidad
    base = f"{norm}@gmail.com"
    email = base
    suffix = 1
    while Usuario.objects.filter(email=email).exists():
        email = f"{norm}{suffix}@gmail.com"
        suffix += 1

    # 4) Rango según tipo
    rango = 'Vendedor' if tipo == 'vendedor' else 'Supervisor'

    # 5) Crear el usuario
    with transaction.atomic():
        user = Usuario.objects.create_user(
            email=email,
            nombre=raw_name,
            dni=dni,
            rango=rango,
            password=dni
        )
        # Añadir la sucursal
        user.sucursales.add(sucursal_obj)
        user.suspendido = True
        user.groups.add(Group.objects.filter(name=rango.capitalize()).first())
        user.save()
    return user


def get_or_create_metodo_pago(metodo_str):
    """
    Devuelve un objeto MetodoPago (creándolo si no existe).
    metodo_str puede ser un ID (string con dígitos) o un alias (por ejemplo, "Efectivo").
    Retorna un objeto MetodoPago válido.
    """
    from sales.models import MetodoPago

    # 1) Intentar parsear como entero para usarlo como ID
    try:
        metodo_id = int(metodo_str)
        # Buscar si existe
        metodo_pago = MetodoPago.objects.get(id=metodo_id)
        return metodo_pago
    except ValueError:
        # No es un entero, pasamos a buscarlo por alias
        pass
    except Exception:
        pass

    # 2) Buscarlo por alias (o nombre)
    alias = metodo_str.strip().lower()  # " Efectivo "
    metodo_pago_qs = MetodoPago.objects.filter(alias__iexact=alias)
    if metodo_pago_qs.exists():
        return metodo_pago_qs.first()

    # 3) No existe => lo creamos
    metodo_pago = MetodoPago.objects.create(
        alias=alias,
    )
    return metodo_pago


def get_or_create_cobrador(cobrador_str):
    """
    Recibe un string que puede ser un ID ('1', '3') o un alias (por ej. 'Juan').
    Devuelve un objeto CuentaCobranza (creándolo si no existe).
    """
    from sales.models import CuentaCobranza

    # 1) Intentar parsearlo como un entero para usarlo como ID
    try:
        cobrador_id = int(cobrador_str)
        return CuentaCobranza.objects.get(id=cobrador_id)
    except ValueError:
        # No es un entero, será un alias
        pass
    except Exception:
        pass

    # 2) Buscarlo por alias (ignorando mayúsculas/minúsculas con __iexact)
    alias = cobrador_str.strip().lower()
    qs = CuentaCobranza.objects.filter(alias__iexact=alias)
    if qs.exists():
        return qs.first()

    # 3) Si no existe, se crea
    # Ajusta los campos según tu modelo (alias, nombre, etc.)
    return CuentaCobranza.objects.create(
        alias=alias
    )

#endregion