import asyncio
from decimal import Decimal
import signal
import tempfile
import time
import unicodedata
from django.utils.dateparse import parse_datetime, parse_date
from django.core.cache import cache
import multiprocessing as mp
from django.forms import ValidationError
from django.shortcuts import get_object_or_404, render, redirect
from django.http import FileResponse, HttpResponse, HttpResponseForbidden, JsonResponse, HttpResponseBadRequest
from django.urls import reverse_lazy
from django.views import generic
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth import get_user_model
from django.db import transaction,connection

from sales.services.dolar import get_dolar_blue_y_oficial_async
from sales.services.arqueo import BILLETES, compute_saldo_diario_caja, get_allowed_agencias, today_fragment_ddmmyyyy, validate_agencia_allowed
from sales.services.caja_query import _format_money, _get_allowed_agencia_ids, _normalize_fecha_fragment, _parse_int, _parse_int_list, _row_from_externo, _row_from_pago
from sales.exports.arqueo import export_arqueo_pdf

from .mixins import TestLogin
from .models import ArqueoCaja, Auditoria, MetodoPago, Ventas,CoeficientesListadePrecios,MovimientoExterno,CuentaCobranza
from users.models import Cliente, Sucursal,Usuario
from .models import Ventas,PagoCannon
from products.models import Products,Plan
from types import SimpleNamespace
import datetime
from django.views.decorators.vary import vary_on_headers
import os,re
from django.db.models.functions import Coalesce, Substr, Concat
from django.db.models import (
    Q, F, Count, Sum, OuterRef, Subquery, IntegerField, Value, Prefetch, Func, CharField
,
)
from django.utils.timezone import make_aware
from django.utils.timezone import localtime
import json
from django.shortcuts import reverse
from dateutil.relativedelta import relativedelta
import locale
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .utils import *
from collections import defaultdict
import elanelsystem.settings as settings
from elanelsystem.views import filterMainManage, convertirValoresALista
from django.forms.models import model_to_dict
from django.templatetags.static import static
from django.forms.models import model_to_dict
from users.models import Usuario
from sales.models import Ventas
from django.http import JsonResponse
from django.views.decorators.cache import cache_control,never_cache
from django.utils.decorators import method_decorator
from elanelsystem.utils import *
from django.contrib.auth.decorators import login_required
from django.db.models import Count

import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.http import HttpResponse
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sales.exports.caja import export_caja

#region EndPoint graficos de torta ---------------------------------------------------
class GraficosDashboard(TestLogin, PermissionRequiredMixin, generic.View):
    """Vista para mostrar el dashboard de gráficos de torta"""
    template_name = 'graficos_dashboard.html'
    permission_required = "sales.my_ver_graficos"
    
    def get(self, request, *args, **kwargs):
        lista_dict_agencias = [sucursal for sucursal in Sucursal.objects.all()]  
        context = {
            'agencias': lista_dict_agencias
        }
        return render(request, self.template_name, context)

# Mantener la función para compatibilidad con URLs existentes
def graficos_dashboard(request):
    """Vista para mostrar el dashboard de gráficos de torta"""
    return render(request, 'graficos_dashboard.html')

@require_GET
@csrf_exempt
def exportar_ventas_excel(request):
    import datetime as dt

    # ---- Filtros ----
    fecha_inicial    = request.GET.get('fecha_inicial')  # YYYY-MM-DD
    fecha_final      = request.GET.get('fecha_final')    # YYYY-MM-DD
    colaborador      = request.GET.get('colaborador')
    tipo_colaborador = request.GET.get('tipo_colaborador', 'vendedor')  # o 'supervisor'
    agencia          = request.GET.get('agencia', '')

    # ---- Base + filtros en DB ----
    qs = get_ventasBySucursal(agencia)

    if colaborador:
        if tipo_colaborador == 'supervisor':
            qs = qs.filter(supervisor__nombre__icontains=colaborador)
        else:
            qs = qs.filter(vendedor__nombre__icontains=colaborador)

    # Agregamos el total recaudado por venta (excluyendo cuota 0)
    

    # ---- Manejo de fechas (en Python, como pediste) ----
    f_ini = f_fin = None
    try:
        if fecha_inicial:
            f_ini = dt.datetime.strptime(fecha_inicial, "%Y-%m-%d")
        if fecha_final:
            f_fin = dt.datetime.strptime(fecha_final, "%Y-%m-%d")
    except ValueError:
        f_ini = f_fin = None

    # ---- Construcción de filas (sin N+1 y en streaming) ----
    rows = []
    for v in qs.iterator(chunk_size=1000):
        # Parseo y filtro de fecha
        try:
            fecha_v = dt.datetime.strptime((v.fecha or "")[:10], "%d/%m/%Y")
        except Exception:
            continue
        if f_ini and fecha_v < f_ini:
            continue
        if f_fin and fecha_v > f_fin:
            continue

        # Cantidad de contratos (si lista vacía/None => 1)
        contratos = v.cantidadContratos or []
        try:
            cant_contratos = len(contratos) if len(contratos) > 0 else 1
        except Exception:
            cant_contratos = 1

        # Importe total (ya multiplicado por contratos en tu modelo) e individual
        importe_total = float(v.importe or 0.0)
        importe_individual = importe_total / float(cant_contratos)


        rows.append({
            'Agencia': v.agencia.pseudonimo,
            'Cliente': v.nro_cliente.nombre,
            'N° Operación': v.nro_operacion,
            'Fecha': v.fecha,
            'Cantidad de contratos': cant_contratos,
            'Importe individual': importe_individual,
            'Monto Facturado': importe_total,
            'Vendedor': v.vendedor.nombre if v.vendedor else 'Sin vendedor',
            'Supervisor': v.supervisor.nombre if v.supervisor else 'Sin supervisor',
            'Producto': v.producto.nombre if v.producto else 'Sin producto',
            'Paquete': v.paquete,
            'Campaña': v.campania,
        })

    # ---- DataFrame ----
    df = pd.DataFrame(rows)

    # ---- Totales (según columnas) ----
    total_ventas         = len(df)
    total_facturacion    = float(df['Monto Facturado'].sum()) if 'Monto Facturado' in df else 0.0

    # ---- Excel ----
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet = 'Ventas'
        df.to_excel(writer, sheet_name=sheet, index=False)

        wb = writer.book
        ws = writer.sheets[sheet]

        # Estilos de encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Formato moneda en columnas relevantes
        money_cols = ['Importe individual', 'Monto Facturado']
        for col_name in money_cols:
            if col_name in df.columns:
                cidx = list(df.columns).index(col_name) + 1  # 1-based
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=cidx).number_format = '#,##0.00'

        # Ajuste de anchos
        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Totales al pie
        last_row = ws.max_row + 2
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        total_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def _cell(r, c, val):
            ws.cell(row=r, column=c, value=val).font = total_font
            ws.cell(row=r, column=c).fill = total_fill
            ws.cell(row=r, column=c).border = total_border

        _cell(last_row, 1, "TOTALES")
        _cell(last_row, 2, f"Total Ventas: {total_ventas}")
        _cell(last_row, 4, f"Total Facturación: ${total_facturacion:,.2f}")

    output.seek(0)
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="ventas_{fecha_inicial}_{fecha_final}.xlsx"'
    return resp

@require_GET
@csrf_exempt
def ventas_analytics_api(request):
    import datetime as dt
    """
    - Fechas: se mantienen en Python (DD/MM/YYYY en Ventas.fecha).
    - Filtros en DB: agencia, colaborador (vendedor/supervisor).
    - Evita N+1 con select_related/only (ver utils).
    - Cuenta "items" por mes = cantidad de contratos, no cantidad de ventas.
    - total_facturacion = importe * cantidadContratos.
    - total_recaudacion = SUM(PagoCannon.monto) (no se multiplica).
    """
    fecha_inicial    = request.GET.get('fecha_inicial')  # YYYY-MM-DD
    fecha_final      = request.GET.get('fecha_final')    # YYYY-MM-DD
    colaborador      = request.GET.get('colaborador')
    tipo_colaborador = request.GET.get('tipo_colaborador', 'vendedor')  # o 'supervisor'
    agencia          = request.GET.get('agencia', '')

    # Base + filtros DB
    qs = get_ventasBySucursal(agencia)
    if colaborador:
        if tipo_colaborador == 'supervisor':
            qs = qs.filter(supervisor__nombre__icontains=colaborador)
        else:
            qs = qs.filter(vendedor__nombre__icontains=colaborador)

    # Fechas en Python (se mantiene)
    f_ini = f_fin = None
    try:
        if fecha_inicial:
            f_ini = dt.datetime.strptime(fecha_inicial, "%Y-%m-%d")
        if fecha_final:
            f_fin = dt.datetime.strptime(fecha_final, "%Y-%m-%d")
    except ValueError:
        f_ini = f_fin = None

    from collections import Counter
    ventas_por_mes = Counter()
    total_facturacion = 0.0
    ventas_ids_filtradas = []

    for v in qs.iterator(chunk_size=1000):
        # Parseo de fecha "DD/MM/YYYY ..."
        try:
            fecha = dt.datetime.strptime((v.fecha or "")[:10], "%d/%m/%Y")
        except Exception:
            continue
        if f_ini and fecha < f_ini:
            continue
        if f_fin and fecha > f_fin:
            continue

        # Cantidad de contratos (si None o vacío => 1)
        contratos = v.cantidadContratos or []
        try:
            contracts_count = len(contratos) if len(contratos) > 0 else 1
        except Exception:
            contracts_count = 1

        # Conteo por mes = contratos (no ventas)
        key = fecha.strftime("%Y-%m")
        ventas_por_mes[key] += contracts_count

        # Facturación total = importe * cantidadContratos
        total_facturacion += float(v.importe)

        ventas_ids_filtradas.append(v.id)

    

    labels = sorted(ventas_por_mes.keys())
    data = [int(ventas_por_mes[l]) for l in labels]  # cantidad de contratos por mes
    total = sum(data)

    return JsonResponse({
        "labels": labels,
        "data": data,                           # contratos por mes
        "total": total,                         # total de contratos
        "total_facturacion": total_facturacion, # importe * contratos
    })

@require_GET
@csrf_exempt
def exportar_pagos_cannon_excel(request):
    import datetime as dt
    # ---- Parámetros ----
    fecha_inicial = request.GET.get('fecha_inicial')  # YYYY-MM-DD
    fecha_final   = request.GET.get('fecha_final')    # YYYY-MM-DD
    cobrador      = request.GET.get('cobrador')
    metodo_pago   = request.GET.get('metodo_pago')
    nro_cuota     = request.GET.get('nro_cuota')
    cliente       = request.GET.get('cliente')
    agencia       = request.GET.get('agencia', '')
    monto         = request.GET.get('monto')
    monto_op      = request.GET.get('monto_op')  # 'gt' | 'lt'

    # ---- Base queryset optimizado + filtros en DB ----
    qs = get_pagosCannonBySucursal(agencia)

    if cobrador:
        qs = qs.filter(cobrador__alias__icontains=cobrador)

    if metodo_pago:
        qs = qs.filter(metodo_pago__alias__icontains=metodo_pago)

    if nro_cuota:
        try:
            qs = qs.filter(nro_cuota=int(nro_cuota))
        except (TypeError, ValueError):
            pass

    if cliente:
        qs = qs.filter(venta__nro_cliente__nombre__icontains=cliente)

    # Nota: dejo el filtro por fecha en Python (como pediste)
    f_ini = f_fin = None
    try:
        if fecha_inicial:
            f_ini = dt.datetime.strptime(fecha_inicial, "%Y-%m-%d")
        if fecha_final:
            f_fin = dt.datetime.strptime(fecha_final, "%Y-%m-%d")
    except ValueError:
        f_ini = f_fin = None

    # Si querés filtrar por monto en DB, tendrías que decidir si es total o unitario.
    # Mantengo el filtro de monto en Python para poder aplicar sobre "monto unitario" si hiciera falta.

    # ---- Armar rows para Excel (streaming con iterator) ----
    rows = []
    for pago in qs.iterator(chunk_size=1000):
        # 1) Fecha (DD/MM/YYYY...) -> datetime para chequear rango
        try:
            fecha_pago = dt.datetime.strptime((pago.fecha or "")[:10], "%d/%m/%Y")
        except Exception:
            # Si no parsea, lo excluimos (mismo criterio que tu código)
            continue

        if f_ini and fecha_pago < f_ini:
            continue
        if f_fin and fecha_pago > f_fin:
            continue

        # 2) Cantidad de contratos
        contratos = pago.venta.cantidadContratos or []
        try:
            cant_contratos = len(contratos) if len(contratos) > 0 else 1
        except Exception:
            cant_contratos = 1

        # 3) Monto individual y total de la cuota
        # Asunción por tu lógica previa: pago.monto es TOTAL de la cuota
        monto_total_cuota = float(pago.monto)
        monto_individual = monto_total_cuota / float(cant_contratos)

        # 4) Filtro por monto (si se envía). Aquí lo aplico sobre el TOTAL (igual que tenías).
        # Si preferís filtrarlo por "monto individual", cambia monto_val vs monto_individual.
        if monto:
            try:
                monto_val = float(monto)
                if monto_op == 'gt' and not (monto_total_cuota >= monto_val):
                    continue
                if monto_op == 'lt' and not (monto_total_cuota <= monto_val):
                    continue
            except ValueError:
                pass

        rows.append({
            'Agencia':   pago.venta.agencia.pseudonimo if getattr(pago.venta, "agencia", None) else 'Sin agencia',
            'Cliente':   pago.venta.nro_cliente.nombre if getattr(pago.venta, "nro_cliente", None) else 'Sin cliente',
            'N° Operación': pago.venta.nro_operacion if pago.venta else 'Sin venta',
            'Fecha Pago': pago.fecha,
            'N° Cuota':  pago.nro_cuota,
            # columnas nuevas
            'Cantidad de contratos': cant_contratos,
            'Monto individual': monto_individual,
            'Monto total de la cuota': monto_total_cuota,
            # existentes
            'Método de Pago': pago.metodo_pago.alias if pago.metodo_pago else 'Sin método',
            'Cobrador': pago.cobrador.alias if pago.cobrador else 'Sin cobrador',
            'Producto': pago.venta.producto.nombre if getattr(pago.venta, "producto", None) else 'Sin producto',
            'Paquete':  pago.venta.paquete if pago.venta else 'Sin paquete',
        })

    # ---- DataFrame ----
    df = pd.DataFrame(rows)

    # ---- Excel (estilos + totales) ----
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet = 'Pagos_Cannon'
        df.to_excel(writer, sheet_name=sheet, index=False)

        wb = writer.book
        ws = writer.sheets[sheet]

        # Header styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Format currency columns (si existen)
        # Nota: openpyxl usa códigos tipo "#,##0.00"
        currency_cols = ['Monto individual', 'Monto total de la cuota']
        for col_name in currency_cols:
            if col_name in df.columns:
                col_idx = list(df.columns).index(col_name) + 1  # 1-based
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = '#,##0.00'

        # Ajuste de ancho
        for column in ws.columns:
            max_len = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # Totales
        total_pagos = len(df)
        total_recaudacion = 0.0
        if 'Monto total de la cuota' in df.columns:
            total_recaudacion = float(df['Monto total de la cuota'].sum())

        last_row = ws.max_row + 2
        total_font = Font(bold=True, size=12)
        total_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
        total_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        ws.cell(row=last_row, column=1, value="TOTALES").font = total_font
        ws.cell(row=last_row, column=1).fill = total_fill
        ws.cell(row=last_row, column=1).border = total_border

        ws.cell(row=last_row, column=2, value=f"Total Pagos: {total_pagos}").font = total_font
        ws.cell(row=last_row, column=2).fill = total_fill
        ws.cell(row=last_row, column=2).border = total_border

        ws.cell(row=last_row, column=4, value=f"Total Recaudación: ${total_recaudacion:,.2f}").font = total_font
        ws.cell(row=last_row, column=4).fill = total_fill
        ws.cell(row=last_row, column=4).border = total_border

    output.seek(0)
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="pagos_cannon_{fecha_inicial}_{fecha_final}.xlsx"'
    return resp

@require_GET
@csrf_exempt
def pagos_cannon_analytics_api(request):
    """
    Mantiene el manejo de fechas en Python (DD/MM/YYYY en PagoCannon.fecha).
    Optimiza: filtros en DB (salvo monto por contrato), elimina listas grandes,
    y hace el conteo por mes sin replicar filas.
    """
    from collections import Counter
    import datetime as dt

    # --- params ---
    fecha_inicial = request.GET.get('fecha_inicial')  # YYYY-MM-DD
    fecha_final   = request.GET.get('fecha_final')    # YYYY-MM-DD
    cobrador      = request.GET.get('cobrador')
    metodo_pago   = request.GET.get('metodo_pago')
    nro_cuota     = request.GET.get('nro_cuota')
    cliente       = request.GET.get('cliente')
    agencia       = request.GET.get('agencia', '')
    monto         = request.GET.get('monto')
    monto_op      = request.GET.get('monto_op')  # 'gt' o 'lt'

    # --- base queryset (ya optimizado en utils) + filtros en DB ---
    qs = get_pagosCannonBySucursal(agencia)

    if cobrador:
        qs = qs.filter(cobrador__alias__icontains=cobrador)

    if metodo_pago:
        qs = qs.filter(metodo_pago__alias__icontains=metodo_pago)

    if nro_cuota:
        s = nro_cuota.strip()
        if "-" in s:  # formato "1-60"
            a, b = s.split("-", 1)
            a = a.strip()
            b = b.strip()
            if a:
                a = int(a)
                qs = qs.filter(nro_cuota__gte=a)
            if b:
                b = int(b)
                qs = qs.filter(nro_cuota__lte=b)
        else:
            # exacta
            n = int(s)
            qs = qs.filter(nro_cuota=n)
        pass

    if cliente:
        qs = qs.filter(venta__nro_cliente__nombre__icontains=cliente)

    # --- parseo de fechas (manteniendo tu forma) ---
    # (No se filtra por fecha en DB: respetamos tu parseo DD/MM/YYYY aquí)
    f_ini = f_fin = None
    try:
        if fecha_inicial:
            f_ini = dt.datetime.strptime(fecha_inicial, "%Y-%m-%d")
        if fecha_final:
            f_fin = dt.datetime.strptime(fecha_final, "%Y-%m-%d")
    except ValueError:
        f_ini = f_fin = None

    # --- recorrido eficiente (sin armar pagos_dict gigante) ---
    pagos_por_mes = Counter()
    total_recaudacion = 0.0

    # stream de resultados para no cargar todo en memoria
    for pago in qs.iterator(chunk_size=1000):
        # 1) fecha del pago (DD/MM/YYYY...)
        try:
            fecha_pago = dt.datetime.strptime((pago.fecha or "")[:10], "%d/%m/%Y")
        except Exception:
            # si no se puede parsear, ignoramos ese registro (como hacías)
            continue

        # 2) filtro por rango de fechas (en Python, como pediste)
        if f_ini and fecha_pago < f_ini:
            continue
        if f_fin and fecha_pago > f_fin:
            continue

        # 3) cantidad de contratos (si null/[]/0 => 1)
        contratos = pago.venta.cantidadContratos or []
        contracts_count = len(contratos)

        # 4) monto unitario por contrato (para filtrar por monto_op)
        monto_unit = float(pago.monto) / float(contracts_count)

        if monto:
            try:
                monto_val = float(monto)
                if monto_op == 'gt' and not (monto_unit >= monto_val):
                    continue
                if monto_op == 'lt' and not (monto_unit <= monto_val):
                    continue
            except ValueError:
                pass

        # 5) acumular por mes y total recaudación
        key = fecha_pago.strftime("%Y-%m")     # etiqueta mensual
        pagos_por_mes[key] += contracts_count  # sumamos cantidad de "items" del mes
        total_recaudacion += float(pago.monto) # suma original (equivale a sumar unitarios replicados)

    # --- serialización final (igual que tu salida) ---
    labels = sorted(pagos_por_mes.keys())
    data = [int(pagos_por_mes[l]) for l in labels]
    total = sum(data)

    return JsonResponse({
        "labels": labels,
        "data": data,
        "total": total,
        "total_recaudacion": total_recaudacion
    })

def graficos_pagos_cannon(request):
    from users.models import Sucursal
    agencias = Sucursal.objects.all()
    context = {
        'agencias': agencias,
    }
    return render(request, 'graficos_cannon.html', context)

@require_GET
@csrf_exempt
def graficos(request):
    from collections import Counter
    import datetime
    from django.db.models import Q
    
    # Filtros
    mes = request.GET.get('mes')
    anio = request.GET.get('anio')
    dia = request.GET.get('dia')
    colaborador = request.GET.get('colaborador')
    tipo_colaborador = request.GET.get('tipo_colaborador', 'vendedor')  # vendedor o supervisor

    ventas = Ventas.objects.all()
    
    # Filtrar por año
    if anio:
        ventas = ventas.filter(fecha__regex=rf"{anio}$|/{anio} ")
    # Filtrar por mes
    if mes:
        ventas = ventas.filter(fecha__regex=rf"/{mes}/")
    # Filtrar por día
    if dia:
        ventas = ventas.filter(fecha__startswith=f"{dia}/")
    # Filtrar por colaborador
    if colaborador:
        if tipo_colaborador == 'supervisor':
            ventas = ventas.filter(supervisor__nombre__icontains=colaborador)
        else:
            ventas = ventas.filter(vendedor__nombre__icontains=colaborador)

    # Agrupar por mes
    ventas_por_mes = Counter()
    for v in ventas:
        try:
            fecha = datetime.datetime.strptime(v.fecha[:10], "%d/%m/%Y")
            key = fecha.strftime("%Y-%m")
            ventas_por_mes[key] += 1
        except Exception:
            continue
    labels = sorted(ventas_por_mes.keys())
    data = [ventas_por_mes[l] for l in labels]
    total = sum(data)
    return JsonResponse({
        "labels": labels,
        "data": data,
        "total": total
    })

class DashboardColaboradores(TestLogin, generic.View):
    template_name = 'dashboard_colaboradores.html'

    def get(self, request, *args, **kwargs):
        from users.models import Sucursal
        context = {
            'agencias': Sucursal.objects.all()
        }
        return render(request, self.template_name, context)

def autocomplete_colaborador(request):
    query = request.GET.get('query', '')
    if len(query) < 2:
        return JsonResponse([], safe=False)
    
    from users.models import Usuario
    from django.db.models import Q
    
    # Buscar colaboradores por nombre que contenga la query
    colaboradores = Usuario.objects.filter(
        Q(nombre__icontains=query) & 
        Q(rango__in=['Vendedor', 'Supervisor'])
    )[:10]  # Limitar a 10 resultados
    
    results = []
    for colaborador in colaboradores:
        results.append({
        'id': colaborador.id,
        'nombre': colaborador.nombre,
        'rango': colaborador.rango,
        'sucursales': [s.pseudonimo for s in colaborador.sucursales.all()],
        'suspendido': colaborador.suspendido,  # <-- Agrega esto
    })
    
    return JsonResponse(results, safe=False)

@require_GET
@csrf_exempt
def ventas_colaborador(request):
    from collections import Counter
    import datetime as dt

    if not request.GET:
        return JsonResponse({"status": False, "message": "Faltan parámetros"}, status=400)

    colaborador_id    = request.GET.get('colaborador_id')
    fecha_inicial_str = request.GET.get('fecha_inicial')   # 'YYYY-MM-DD'
    fecha_final_str   = request.GET.get('fecha_final')     # 'YYYY-MM-DD'

    if not (colaborador_id and fecha_inicial_str and fecha_final_str):
        return JsonResponse({"status": False, "message": "Completa todo, flaco 😅"}, status=400)

    try:
        f_ini = dt.datetime.strptime(fecha_inicial_str, "%Y-%m-%d").date()
        f_fin = dt.datetime.strptime(fecha_final_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"status": False, "message": "Fechas inválidas"}, status=400)

    user = get_object_or_404(Usuario, id=colaborador_id)

    if user.rango == "Vendedor":
        qs = Ventas.objects.filter(vendedor=user)

        ids_en_rango = []
        for row in qs.values('id', 'fecha').iterator(chunk_size=1000):
            try:
                d = dt.datetime.strptime((row['fecha'] or '')[:10], '%d/%m/%Y').date()
            except Exception:
                continue
            if (not f_ini or d >= f_ini) and (not f_fin or d <= f_fin):
                ids_en_rango.append(row['id'])

        ventas = qs.filter(id__in=ids_en_rango).select_related('vendedor').only(
            'id', 'fecha', 'importe', 'cantidadContratos',
            'vendedor__id', 'vendedor__nombre',
        )

        periodo_counter = Counter()  # contratos por periodo (YYYY-MM)
        total_facturacion = 0.0

        for v in ventas.iterator(chunk_size=1000):
            # periodo
            try:
                fecha = dt.datetime.strptime((v.fecha or '')[:10], '%d/%m/%Y').date()
            except Exception:
                continue
            periodo = f"{fecha.year}-{fecha.month:02d}"

            # contratos (si None o [] => 1)
            contratos = v.cantidadContratos or []
            try:
                contracts_count = len(contratos) if len(contratos) > 0 else 1
            except Exception:
                contracts_count = 1

            # acumular por periodo y facturación
            periodo_counter[periodo] += contracts_count
            total_facturacion += float(v.importe or 0)

        labels = sorted(periodo_counter.keys())
        data = [int(periodo_counter[k]) for k in labels]
        total_contratos = sum(data)  


        print(
            {
            "status": True,
            "labels": labels,
            "data": data,                              # contratos por periodo
            "total_facturacion": total_facturacion,    # suma de importes (ya multiplicado por contratos)
            "cantidad_total_ventas": total_contratos,  # ventas == contratos
            }
        )
        return JsonResponse({
            "status": True,
            "labels": labels,
            "data": data,                              # contratos por periodo
            "total_facturacion": total_facturacion,    # suma de importes (ya multiplicado por contratos)
            "cantidad_total_ventas": total_contratos,  # ventas == contratos
        })


        
    elif user.rango == "Supervisor":
        qs = Ventas.objects.filter(supervisor=user)
        # Filtrado por fecha (fecha es string DD/MM/YYYY ...)
        ids_en_rango = []
        for row in qs.values('id', 'fecha').iterator(chunk_size=1000):
            try:
                d = dt.datetime.strptime((row['fecha'] or '')[:10], '%d/%m/%Y').date()
            except Exception:
                continue
            if (not f_ini or d >= f_ini) and (not f_fin or d <= f_fin):
                ids_en_rango.append(row['id'])

        ventas = qs.filter(id__in=ids_en_rango).select_related('vendedor').only(
            'id', 'fecha', 'importe', 'cantidadContratos',
            'vendedor__id', 'vendedor__nombre', 'vendedor__suspendido'
        )

        periodo_counter = Counter()  # contratos por periodo (YYYY-MM)
        total_facturacion = 0.0
        vendedores = {}  # {vid: {id, nombre, suspendido, cantidad_ventas(=contratos), total_importe}}

        for v in ventas.iterator(chunk_size=1000):
            # periodo
            try:
                fecha = dt.datetime.strptime((v.fecha or '')[:10], '%d/%m/%Y').date()
            except Exception:
                continue
            periodo = f"{fecha.year}-{fecha.month:02d}"

            # contratos (si None o [] => 1)
            contratos = v.cantidadContratos or []
            try:
                contracts_count = len(contratos) if len(contratos) > 0 else 1
            except Exception:
                contracts_count = 1

            # acumular por periodo y facturación
            periodo_counter[periodo] += contracts_count
            total_facturacion += float(v.importe or 0)

            # acumular por vendedor (ventas = contratos)
            if v.vendedor_id:
                bucket = vendedores.setdefault(v.vendedor_id, {
                    "vendedor__id": v.vendedor_id,
                    "vendedor__nombre": getattr(v.vendedor, 'nombre', 'Sin nombre'),
                    "vendedor__suspendido": bool(getattr(v.vendedor, 'suspendido', False)),
                    "cantidad_ventas": 0,   # aquí guardamos contratos
                    "total_importe": 0.0,
                })
                bucket["cantidad_ventas"] += contracts_count
                bucket["total_importe"] += float(v.importe or 0)

        labels = sorted(periodo_counter.keys())
        data = [int(periodo_counter[k]) for k in labels]
        total_contratos = sum(data)  # == cantidad_total_ventas

        vendedores_data = sorted(vendedores.values(), key=lambda x: x["vendedor__nombre"] or "")

        print(
            {
            "status": True,
            "labels": labels,
            "data": data,                              # contratos por periodo
            "total_facturacion": total_facturacion,    # suma de importes (ya multiplicado por contratos)
            "cantidad_total_ventas": total_contratos,  # ventas == contratos
            "vendedores": vendedores_data,             # cantidad_ventas = contratos por vendedor
            }
        )
        return JsonResponse({
            "status": True,
            "labels": labels,
            "data": data,                              # contratos por periodo
            "total_facturacion": total_facturacion,    # suma de importes (ya multiplicado por contratos)
            "cantidad_total_ventas": total_contratos,  # ventas == contratos
            "vendedores": vendedores_data,             # cantidad_ventas = contratos por vendedor
        })
    else:
        return JsonResponse({"status": False, "message": "Unicamente vendedores y supervisores"}, status=400)

#endregion ----------------------------------------


class Resumen(TestLogin,PermissionRequiredMixin,generic.View):
    permission_required = "sales.my_ver_resumen"
    # login_url = "/ventas/caja/"
    template_name = 'resumen.html'

    def get(self,request,*args,**kwargs):
        ventas = Ventas.objects.all()
        context = {
            "ventas" : ventas,
        }
        # print(context)
        return render(request, self.template_name, context)
    
    def handle_no_permission(self):
        return redirect("users:list_customers")


#region Ventas - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - 
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch') # Para no guardar el cache 
class CrearVenta(TestLogin,generic.DetailView):
    model = Cliente
    template_name = "create_sale.html"

    def get(self,request,*args, **kwargs):
        self.object = self.get_object()

        products = Products.objects.filter(activo=True).order_by('nombre')

        campaniasDisponibles = getCampanasDisponibles()

        #endregion
        context ={
            "object": self.object,
            'products': products, 
            'agencias': request.user.sucursales.all(), 
            'agenciaActual': request.user.sucursales.all()[0], 
            'campaniasDisponibles': campaniasDisponibles, 
        }

        return render(request,self.template_name,context)
    

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()          # Cliente único de la URL /cliente/<pk>/
        form = json.loads(request.body)
        errors = {}
        sale = Ventas()

        # === Cliente: NO vuelvas a consultar por dni/nro_cliente
        sale.nro_cliente = self.object           # asumiendo FK a Cliente

        # === Producto: ahora recibís ID
        producto_id = form.get("producto")
        try:
            sale.producto = Products.objects.get(id=int(producto_id))
        except (TypeError, ValueError, Products.DoesNotExist):
            errors["producto"] = "Producto inválido."

        # === Agencia por ID (ya lo hacías así)
        agencia_id = form.get("agencia")
        try:
            sale.agencia = Sucursal.objects.get(id=int(agencia_id))
        except (TypeError, ValueError, Sucursal.DoesNotExist):
            errors["agencia"] = "Agencia inválida."

        # === Vendedor por ID (evitá nombre__iexact)
        vendedor_id = form.get("vendedor")
        try:
            sale.vendedor = Usuario.objects.get(id=int(vendedor_id), suspendido=False)
        except (TypeError, ValueError, Usuario.DoesNotExist):
            errors["vendedor"] = "Vendedor inválido."

        # === Supervisor por ID y rango
        supervisor_id = form.get("supervisor")
        try:
            sale.supervisor = Usuario.objects.get(id=int(supervisor_id), suspendido=False, rango="Supervisor")
        except (TypeError, ValueError, Usuario.DoesNotExist):
            errors["supervisor"] = "Supervisor inválido."

        # === Chances (nro_contrato_X / nro_orden_X)
        chances = []
        i = 1
        while f"nro_contrato_{i}" in form:
            nc = form.get(f"nro_contrato_{i}")
            no = form.get(f"nro_orden_{i}")
            if not (nc and nc.isdigit()):
                errors[f"nro_contrato_{i}"] = "Debe contener solo números."
            if not (no and no.isdigit()):
                errors[f"nro_orden_{i}"] = "Debe contener solo números."
            if nc and no and nc.isdigit() and no.isdigit():
                chances.append({"nro_contrato": nc, "nro_orden": no})
            i += 1
        sale.cantidadContratos = chances

        # === Conversión segura de numéricos
        def to_int(key, default=0):
            v = form.get(key)
            if v in (None, ""): return default
            try: return int(v)
            except (TypeError, ValueError):
                errors[key] = "Número inválido."
                return default

        def to_float(key, default=0.0):
            v = form.get(key)
            if v in (None, ""): return default
            try: return float(v)
            except (TypeError, ValueError):
                errors[key] = "Número inválido."
                return default

        sale.modalidad = form.get("modalidad") or ""
        sale.importe = to_int("importe")
        sale.primer_cuota = to_int("primer_cuota")
        sale.anticipo = to_int("anticipo")
        sale.tasa_interes = to_float("tasa_interes")
        sale.intereses_generados = to_int("intereses_generados")
        sale.importe_x_cuota = to_int("importe_x_cuota")
        sale.total_a_pagar = to_int("total_a_pagar")
        sale.nro_cuotas = to_int("nro_cuotas")

        # === Fecha: venís formateando ddMMYYYY desde el datepicker
        sale.fecha = form.get("fecha") + " 00:00"

        sale.tipo_producto = form.get("tipo_producto") or ""
        sale.paquete = form.get("paquete") or ""
        sale.campania = form.get("campania") or ""
        sale.observaciones = form.get("observaciones") or ""

        # Si ya hay errores de validación previa, devolvé ahora
        if errors:
            return JsonResponse({"success": False, "errors": errors})

        # Validación de modelo
        try:
            sale.full_clean()
        except ValidationError as e:
            errors.update(e.message_dict)

        if errors:
            return JsonResponse({"success": False, "errors": errors})

        # Guardá primero para tener PK antes de crear cuotas relacionadas
        sale.save()

        # Si crearCuotas() necesita el sale.id, hacelo después del save
        sale.crearCuotas()
        sale.setDefaultFields()  # si setea flags por defecto en el propio sale, podrías llamar antes o después según tu implementación
        # sale.save(update_fields=None)  # por si setDefaultFields modificó algo

        return JsonResponse({
            "success": True,
            "urlRedirect": reverse("users:cuentaUser", args=[sale.nro_cliente.pk])
        })

class VentasDetalles(generic.View):
    template_name = "detallesVentas.html"

    def get(self,request,*args, **kwargs):
        customers = Cliente.objects.all()
        products = Products.objects.all()
        campanias = getTodasCampaniasDesdeInicio()
        sucursales = Sucursal.objects.all()
        usuarios = Usuario.objects.all()
        
        ventas = Ventas.objects.all()
        ventas = [{
            'id': venta.pk,
            'cliente': venta.nro_cliente,
            'nro_operacion': venta.nro_operacion,
            'contratos': [contrato["nro_orden"] for contrato in venta.cantidadContratos],
            'modalidad': venta.modalidad,
            'nro_cuotas': venta.nro_cuotas,
            'agencia': venta.agencia.pseudonimo,
            'campania': venta.campania,
            'importe': venta.importe,
            'tasa_interes': venta.tasa_interes,
            'intereses_generados': venta.intereses_generados,
            'total_a_pagar': venta.total_a_pagar,
            'importe_x_cuota': venta.importe_x_cuota,
            'fecha': venta.fecha,
            'producto': venta.producto.nombre,
            'paquete': venta.paquete,
            'vendedor': venta.vendedor.nombre if venta.vendedor else "",
            'supervisor': venta.supervisor.nombre if venta.supervisor else "",

        } for venta in ventas]
        
        context ={
            'customers': customers, 
            'products': products, 
            'usuarios': usuarios,
            'agencias': sucursales, 
            'campanias': campanias,
            'ventas': ventas
        }

        return render(request,self.template_name,context)
    

def importVentas(request):
    if request.method == "POST":
        start_time = time.time()
        archivo_excel = request.FILES['file']
        agencia = request.POST.get('agencia')
        fs = FileSystemStorage()
        filename = fs.save(archivo_excel.name, archivo_excel)
        file_path = fs.path(filename)
        cantidad_nuevas_ventas = 0

        try:
            df_res, df_est = preprocesar_excel_ventas(file_path)

            sucursal_obj = Sucursal.objects.get(id=agencia)
            print("🔎 ???????????")

             # 1) Preparo el set de contratos ya importados
            todosLosContratosDict = obtener_todos_los_contratos(sucursal_obj)
            set_contratos = {
                str(ct['nro_contrato']) for ct in todosLosContratosDict
            }

            clientes = {
                c.nro_cliente: c
                for c in Cliente.objects.filter(agencia_registrada=sucursal_obj)
            }

            ventas_to_create = []
            grupos = df_res.groupby(
                ['cod_cli','fecha_incripcion','producto_key','paq','vendedor_key','superv_key']
            )
            
            print("🔎 Grupos totales detectados:", grupos.ngroups)
            
            for keys, group in grupos:
                cod_cli, fecha_incripcion, producto, paq, vendedor, superv = keys
                
                # Si el cliente no existe → saltamos
                if cod_cli not in clientes:
                    print(f"|\nNo existe el cliente{ group['cod_cli'].iloc[0]}\n|")
                    continue

                cantidad_chances   = len(group)
                importe_sum= int(group['importe'].sum())
                tasa_int_sum  = float(group['tasa_de_inte'].sum())
                # if cod_cli == "cli_049":
                #     print("========================================================")
                #     print(keys)
                #     print("========================================================")
                #     print(group)
                #     break
                # Lista de contratos / órdenes
                contratos = group[['contrato','nro_de_orden']]\
                    .apply(lambda r: {
                        'nro_contrato': str(int(r['contrato'])),
                        'nro_orden'   : str(int(r['nro_de_orden']))
                    }, axis=1)\
                    .tolist()
                
                id_venta_unica = int(group['id_venta'].iloc[0])

                # — 4) Skip si YA existe cualquiera de esos contratos
                contratos_nros = { c['nro_contrato'] for c in contratos }
                duplicados = contratos_nros & set_contratos
                if duplicados:
                    print(f"❌  Grupo con id_venta={int(group['id_venta'].iloc[0])} SE SALTA porque ya existe(n) contrato(s):"
                          f" {duplicados}.  Grupo completo: {contratos_nros}")
                    continue


                raw_vendedor = group['vendedor_raw'].iloc[0]
                vendedor_obj = get_or_create_usuario_from_import(
                    raw_name    = raw_vendedor,
                    tipo        = 'vendedor',
                    sucursal_obj= sucursal_obj
                )

                raw_superv  = group['superv_raw'].iloc[0]
                supervisor_obj = get_or_create_usuario_from_import(
                    raw_name    = raw_superv,
                    tipo        = 'supervisor',
                    sucursal_obj= sucursal_obj
                )

               # obtén el producto y su plan
                try:
                    producto_raw = group['producto_raw'].iloc[0]
                    producto_obj = get_or_create_product_from_import(producto_raw,group['importe'].iloc[0])
                except ValueError as e:
                    print(f"Error al obtener el producto: {e}")
                    continue
                
                plan = producto_obj.plan if producto_obj else None

                # ------------------------------------------------------------------
                # aquí construimos las cuotas AGREGADAS de todas las chances
                cuotas_agg = build_aggregated_cuotas(
                    id_venta = id_venta_unica,
                    df_est    = df_est,
                    n_chances = cantidad_chances,
                    plan = plan
                )
                print(f"Finalizando grupo de {group['cod_cli'].iloc[0]} - Contratos: {contratos}")
                # ------------------------------------------------------------------
                
                ventas_to_create.append(Ventas(
                    nro_cliente        = clientes[cod_cli],
                    agencia            = sucursal_obj,
                    modalidad          = 'Mensual',
                    nro_cuotas         = len(cuotas_agg)-1,
                    nro_operacion      = id_venta_unica,
                    campania           = obtenerCampaña_atraves_fecha(formatar_fecha(fecha_incripcion)),
                    suspendida         = False,
                    importe            = importe_sum,
                    tasa_interes       = round(tasa_int_sum, 2),
                    primer_cuota       = cuotas_agg[1]['total'] if len(cuotas_agg)>1 else 0,
                    anticipo           = cuotas_agg[0]['total'],
                    intereses_generados= sum(q['total'] for q in cuotas_agg[1:]) - importe_sum,
                    importe_x_cuota    = cuotas_agg[2]['total'] if len(cuotas_agg)>2 else 0,
                    total_a_pagar      = sum(q['total'] for q in cuotas_agg),
                    fecha              = formatar_fecha(fecha_incripcion, with_time=True),
                    producto           = producto_obj,
                    paquete            = paq,
                    vendedor           = vendedor_obj,
                    supervisor         = supervisor_obj,
                    observaciones      = group['comentarios__observaciones'].iloc[0],
                    cantidadContratos  = contratos,
                    cuotas             = cuotas_agg,
                ))
            print(f"Finalizando grupo de {group['cod_cli'].iloc[0]} 2")
            
            
            with transaction.atomic():
                ventas_created = Ventas.objects.bulk_create(ventas_to_create)
                 # 1) Preparo lista de PagoCannon por crear
                
                print(f"✅ CONTINUANDO CON LA CREACION CUOTAS...")
                pagos_to_create = []
                for venta in ventas_created:
                    for cuota_dict in venta.cuotas:
                        nro = int(cuota_dict['cuota'].split()[-1])
                        for pago_data in cuota_dict.get('pagos', []):
                            print(f"ID de metodo de pago: {pago_data['metodoPago']}\n ID de cobrador: {pago_data['cobrador']}\n")
                            pagos_to_create.append(
                                PagoCannon(
                                    venta=venta,
                                    nro_cuota = nro,
                                    monto = pago_data['monto'],
                                    metodo_pago= MetodoPago.objects.get(id=pago_data['metodoPago']),
                                    cobrador = CuentaCobranza.objects.get(id=pago_data['cobrador']),
                                    responsable_pago = None,
                                    fecha = pago_data['fecha'],
                                    campana_de_pago = pago_data['campaniaPago'],
                                )
                            )
                
                print(f"✅ {len(pagos_to_create)} PAGOS POR CREAR")
                 # — Genero N recibos de golpe, empezando justo donde la seq quedó
                count = len(pagos_to_create)
                if count:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT nextval('recibo_seq') FROM generate_series(1, %s)",
                            [count]
                        )
                        seq_vals = [row[0] for row in cursor.fetchall()]

                    # Asigno cada nro_recibo antes del bulk_create
                    for pago_obj, seq in zip(pagos_to_create, seq_vals):
                        pago_obj.nro_recibo = f"RC-{seq:06d}"

                pagos_created = PagoCannon.objects.bulk_create(pagos_to_create)
                print(f"✅ {len(pagos_created)} PAGOS CREADOS")
                # 1) Agrupa los pagos recién creados por (venta_id, nro_cuota)
                pagos_por_venta_cuota = defaultdict(list)
                for pago in pagos_created:
                    key = (pago.venta.id, pago.nro_cuota)
                    pagos_por_venta_cuota[key].append(pago.id)

                print(f"✅ Pasó el agrupamiento de pagos")
                ventas_map = { v.id: v for v in ventas_created}
                ventas_para_actualizar = []
                # 3) Recorre cada grupo y actualiza la cuota correspondiente
                for (venta_id, nro_cuota), pago_ids in pagos_por_venta_cuota.items():
                    venta = ventas_map[venta_id]
                    # como construiste la lista en orden, cuota = índice
                    cuota_dict = venta.cuotas[nro_cuota]
                    cuota_dict['pagos'] = pago_ids
                    ventas_para_actualizar.append(venta)
                print(f"✅ Pasó el agrupamiento de ventas")

                for venta in ventas_para_actualizar:
                    print(f"|\nVenta {venta.nro_operacion} Cliente {venta.nro_cliente.nombre}")
                    # recalculo vencimientos y suspensión (si hace falta)
                    venta.testVencimientoCuotas()
                    print(f"✅ Pasó test de vencimiento de cuotas")
                    
                    venta.suspenderOperacion()
                    print(f"✅ Pasó la verificacion de suspension de la operacion")

                    venta.cuotas = bloquer_desbloquear_cuotas(venta.cuotas)
                    if(int(venta.nro_operacion) == 2342):
                        print(venta.cuotas)
                    print(f"✅ Pasó el bloqueo o desbloqueo de cuotas")

                    venta.setDefaultFields()
                print(f"✅ Pasó el recalculo de vencimientos y suspensión")

                Ventas.objects.bulk_update(ventas_para_actualizar,['cuotas', 'adjudicado', 'suspendida', 'deBaja'])
                print(f"✅  == CREACION DE CUOTAS Y PAGOS CON EXITO == ")


            cantidad_nuevas_ventas = len(ventas_created)
            # Ahora aplicamos bloqueos, defaults y señales a cada venta recién creada

            elapsed = time.time() - start_time
            print(f"✅ == {len(ventas_created)} VENTAS IMPORTADAS CON EXITO ==")
            print(f"⏱️ Tiempo total de importación: {elapsed:.2f} segundos")

            
            fs.delete(filename)
            iconMessage = "/static/images/icons/checkMark.svg"
            message = f"Datos importados correctamente. Se agregaron {cantidad_nuevas_ventas} nuevas ventas"
            return JsonResponse({"message": message, "iconMessage": iconMessage, "status": True})

        except Exception as e:
            print(f"Error al importar: {e}")
            iconMessage = "/static/images/icons/error_icon.svg"
            message = "Error al procesar el archivo"
            return JsonResponse({"message": message, "iconMessage": iconMessage, "status": False})

    return render(request, 'importar_cuotas.html')


def build_aggregated_cuotas(id_venta,df_est,n_chances,plan):
    """
    Para la venta identificada por id_venta (una sola chance),
    agrupa sus filas de ESTADOS en cuotas 0..max, y multiplica
    los importes por n_chances.

    - id_venta: el entero de una sola chance
    - df_est: DataFrame preprocesado de ESTADOS
    - n_chances: cuántas chances totales tiene la venta

    Devuelve lista de dicts [{'cuota','status','total','pagos',...}, ...]
    """
    # Filtramos SOLO la chance que nos interesa
    sub = df_est[df_est['id_venta'] == id_venta]
    if sub.empty:
        return []
    # print(f"Sub dataframe:\n{sub}")
    

    max_q = int(sub['cuota_num'].max())
    # print(f"Max cuota: {max_q}")
    cuotas = []

    for q in range(max_q + 1):
        grp = sub[sub['cuota_num'] == q]
        if grp.empty:
            continue

        # Solo la primera fila de esa cuota
        r = grp.iloc[0]

         # Importe que llegó por fila en el Excel, multiplicado
        excel_amount = int(r['importe_cuotas']) * n_chances

        # Definir importe "oficial" según q
        if q == 0:
            official = plan.suscripcion * n_chances
        elif q == 1:
            official = plan.primer_cuota * n_chances
        else:
            official = excel_amount

        # Descuento solo si excel < official
        descuento_monto = max(0, official - excel_amount)

        # Estado: 'Pagado' o 'Pendiente'
        status = str(r["estado_norm"]).title() if r["estado_norm"] not in ["Vencido", "BAJA","vencido","baja"] else "Vencido"
        # Total de la cuota: importe_cuotas * n_chances
        # total_q = int(r['importe_cuotas']) * n_chances

        # Fecha de vencimiento y de pago con hora forzada
        fecha_venc = formatar_fecha(r['fecha_de_venc'], with_time=True)
        fecha_pago = formatar_fecha(r['fecha_de_pago'], with_time=True)

        # Construimos pagos: uno solo si está pagada
        pagos = []
        if status == 'Pagado':
            pagos = [{
                'monto': excel_amount,
                'metodoPago': get_or_create_metodo_pago(r["medio_de_pago"].title()).id,
                'fecha': fecha_pago,
                'cobrador': get_or_create_cobrador(r["cobrador"].capitalize()).id,
                'campaniaPago': r['campania_pago']
            }]

        cuotas.append({
            'cuota': f'Cuota {q}',
            'nro_operacion': id_venta,
            'status': status,
            'total': official,
            'descuento': {'autorizado': f"{'Gerente de la sucursal' if status == 'Pagado' else ''}", 'monto': descuento_monto},
            'bloqueada': False,
            'fechaDeVencimiento': fecha_venc,
            'diasRetraso': 0,
            "interesPorMora": 0,
            "totalFinal": 0,
            'pagos': pagos,
            'autorizada_para_anular': False,
        })

    return cuotas



def requestVendedores_Supervisores(request):
    payload = json.loads(request.body or '{}')

    sucursal_id = payload.get("agencia") or ""
    sucursal = Sucursal.objects.filter(id=int(sucursal_id)).first() if sucursal_id else None

    vendedores_qs = Usuario.objects.none()
    supervisores_qs = Usuario.objects.none()

    if sucursal:
        vendedores_qs = Usuario.objects.filter(
            sucursales__in=[sucursal],
            rango__in=["Vendedor", "Supervisor"],
            suspendido=False
        ).distinct()

        supervisores_qs = Usuario.objects.filter(
            sucursales__in=[sucursal],
            rango="Supervisor",
            suspendido=False
        ).distinct()

    # Armamos vendedores incluyendo, si existe, su supervisor
    vendedores = []
    for v in vendedores_qs:
        supervisor_a_cargo = v.supervisores_a_cargo
        sup_id = supervisor_a_cargo.id if supervisor_a_cargo else None
        sup_nombre = supervisor_a_cargo.nombre if supervisor_a_cargo else None
        
        vendedores.append({
            "id": v.id,
            "name": v.nombre,
            "supervisor_id": sup_id,
            "supervisor_name": sup_nombre
        })

    supervisores = [{"id": s.id, "name": s.nombre} for s in supervisores_qs]

    return JsonResponse({"vendedores": vendedores, "supervisores": supervisores})

#region Detalle de ventas y funciones de ventas
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
@method_decorator(vary_on_headers('Cookie'), name='dispatch')
class DetailSale(TestLogin, generic.DetailView):
    model = Ventas
    template_name = "detail_sale.html"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        
        response_dolar = cotizaciones_dolar(request)
        dolar_oficial = 0
        if response_dolar.status_code == 200:
            dolar_oficial = json.loads(response_dolar.content).get("oficial").get("compra",0)

        # 1) Primero recalculá/actualizá estados
        self.object.testVencimientoCuotas()
        # Si testVencimientoCuotas() hace .save() o persiste algo, refrescá
        try:
            self.object.refresh_from_db()
        except Exception:
            pass

        sale = self.object  # UNA sola instancia coherente

        request.session["ventaPK"] = sale.pk
        request.session["statusKeyPorcentajeBaja"] = False

        try:
            if sale.adjudicado["status"] == True:
                sale.addPorcentajeAdjudicacion()
            else:
                sale.removePorcentajeAdjudicacion()
        except Exception:
            if sale.adjudicado:
                sale.addPorcentajeAdjudicacion()
            else:
                sale.removePorcentajeAdjudicacion()

        context = {
            "object": sale,
            "changeTitularidad": list(reversed(sale.cambioTitularidadField)),
            "cuotas": sale.cuotas,
            "cobradores": CuentaCobranza.objects.all(),
            "metodosDePagos": MetodoPago.objects.all(),
            "nro_cuotas": sale.nro_cuotas,
            "dolar_oficial": dolar_oficial,
            "urlRedirectPDF": reverse("sales:bajaPDF", args=[sale.pk]),
            "urlUser": reverse("users:cuentaUser", args=[sale.pk]),
            "deleteSaleUrl": reverse("sales:delete_sale", args=[sale.pk]),
            "solicitudAnulacionCuotaUrl": reverse("sales:solicitudAnulacionCuota"),
            "confirmacionAnulacionCuotaUrl": reverse("sales:darBajaCuota"),
        }

        request.session["venta"] = model_to_dict(sale)

        try:
            context["porcetageDefault"] = 50 if len(sale.cuotas_pagadas()) >= 6 else 0
        except IndexError:
            context["porcetageDefault"] = 0

        return render(request, self.template_name, context)

# Eliminar una venta
def eliminarVenta(request,pk):
    form = json.loads(request.body)
    nro_operacion = form["nro_operacion_delete"]
    venta = Ventas.objects.get(pk=pk)
    if(venta.nro_operacion == int(nro_operacion)):
        try:
            venta.delete()
            return JsonResponse({"status": True,'urlRedirect': reverse('users:cuentaUser', args=[venta.nro_cliente.pk])}, safe=False)
        except Exception:
            return JsonResponse({"status": False,"message":"Error al eliminar la venta"}, safe=False)
    else:
        return JsonResponse({"status": False,"message":"Numero incorrecto"}, safe=False)


# Anula una cuota de una venta
def anularCuota(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            permisosUser = request.user.get_all_permissions()

            venta = Ventas.objects.get(pk=request.session["venta"]["id"])
            cuota = data.get('cuota')
            cuotasPagadas = venta.cuotas_pagadas()

            ultima_cuotaPagadaSinCredito = getCuotasPagadasSinCredito(cuotasPagadas)[-1]

            if  ultima_cuotaPagadaSinCredito["cuota"] == cuota and "sales.my_anular_cuotas" in permisosUser:            
                if(request.user.additional_passwords["anular_cuotas"]["password"] == data.get("password")):
                    venta.anularCuota(cuota)
                else:
                    return JsonResponse({"status": False,"password": False,"message":"Contraseña incorrecta"}, safe=False)
                
            return JsonResponse({"status": True,"password":True,"message":"Cuota anulada correctamente"}, safe=False)
        except Exception as error:
            print("Error")
            print(error)
            return JsonResponse({"status": False,"password":True,"message":"Error al anular la cuota","detalleError":str(error)}, safe=False)


# Aplica el descuento a una cuota
def aplicarDescuentoCuota(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            venta = Ventas.objects.get(pk=int(data["ventaID"]))
            cuota = data.get('cuota')
            descuento = data.get('descuento')
            autorizado = data.get('autorizado')

            venta.aplicarDescuento(cuota,int(descuento),autorizado)
            cuotaUpdate = getUnaCuotaDeUnaVenta(request) # Actualizamos la cuota en la session
            print()
            return JsonResponse({"status": True,"message":"Descuento aplicado correctamente","cuotaUpdate": json.loads(cuotaUpdate.content)}, safe=False)

        except Exception as error:   
            return JsonResponse({"status": False,"message":"Descuento fallido","detalleError":str(error)}, safe=False)


# Obtenemos una cuota
@never_cache
def getUnaCuotaDeUnaVenta(request):
    if request.method == 'POST':
        # try:
            data = json.loads(request.body)
            venta = Ventas.objects.get(pk=int(data.get("ventaID")))
            cuotas = venta.cuotas
            cuotaRequest = data.get("cuota")
            # print(data)
            bloqueado_path = static('images/icons/iconBloqueado.png')
            permisosUser = request.user.get_all_permissions()
            buttonDescuentoCuota = f'<button type="button" id="btnDescuentoCuota" onclick="openModalDescuento()" class="buttonDescuentoCuota">Aplicar descuento</button>'
            buttonAnularCuota = '<button type="button" id="anularButton" onclick="anulacionCuota()" class="delete-button-default">Anular cuota</button>'

            buttonSolicitudDeCancelacionPago = '<button type="button" onclick="solicitudBajaCuota()" class="buttonCancelarPago delete-button-default" id="btnBajaCuota">Cancelar pago</button>'
            buttonConfirmacionDeCancelacionPago = '<button type="button" onclick="anulacionCuota()" class="buttonAnularCuota delete-button-default" id="btnAnularCuota">Anular pago</button>'
            cuotasPagadas_parciales = venta.cuotas_pagadas() + venta.cuotas_parciales()
            lista_cuotasPagadasSinCredito = getCuotasPagadasSinCredito(cuotasPagadas_parciales)
            
            for cuota in cuotas:
                if cuota["cuota"] == cuotaRequest:
                    # print("CUOTA ENCONTRADA:", cuota)
                    pagos = PagoCannon.objects.filter(venta=venta, nro_cuota=int(cuotaRequest.split()[-1]))
                    cuota["pagos"] = [pago.json() for pago in pagos]

                    cuota["styleBloqueado"] = f"background-image: url('{bloqueado_path}')"
                    if len(lista_cuotasPagadasSinCredito) != 0 and lista_cuotasPagadasSinCredito[-1]["cuota"] == cuota["cuota"] and "sales.my_anular_cuotas" in permisosUser:
                        cuota["buttonAnularCuota"] = buttonAnularCuota
                    
                    # if (cuota["status"] == "Pagado" or cuota["status"] == "Parcial") and not cuota["autorizada_para_anular"]:
                    #     cuota["buttonCancelacionDePago"] = buttonSolicitudDeCancelacionPago
                    elif(cuota["status"] == "Pagado" or cuota["status"] =="Parcial") and cuota["autorizada_para_anular"]:
                        cuota["buttonAnularCuota"] = buttonConfirmacionDeCancelacionPago
                        

                    if ((cuota["cuota"] == "Cuota 0" or cuota["cuota"] == "Cuota 1") and cuota["status"] != "Pagado"):
                        cuota["buttonDescuentoCuota"] = buttonDescuentoCuota
                    
                    restante_por_divisa = get_dinero_restante_segun_divisa(cuota, data.get("dolar",0))
                    cuota["restante_por_divisa"] = restante_por_divisa
                    
                    resp = JsonResponse(cuota, safe=False)
                    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
                    resp["Pragma"] = "no-cache"
                    resp["Expires"] = "0"
                    return resp

        # except Exception as error:
        #     return JsonResponse({"status": False,"message":"Error al obtener la cuota","detalleError":str(error)}, safe=False)

# ===== helpers de normalización / formato =====
def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    return s

def _fmt_money(x) -> str:
    d = Decimal(x or 0)
    return f"{d:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ===== mapeo a 3 categorías =====
WALLETS_Y_TRANSFER = {
    "transferencia", "tranferencia", "transfer", "transferencias",
    "mercado pago", "mp", "mercadopago",
    "uala", "uála",
    "naranja x", "naranjax",
    "claro pay", "claropay",
    "personal pay", "personalpay",
    "prex",
    "galicia",  # si en tu flujo 'Galicia' representa transferencia bancaria
    "otros"     # si querés forzar 'otros' aquí; podés moverlo a donde prefieras
}

TARJETAS_Y_BANCOS = {
    "posnet", "tarjeta", "tarjetas", "visa", "mastercard", "amex",
    "debito", "crédito", "credito", "bapro", "nativa", "cabal"
}

EFECTIVO_SET = {"efectivo", "contado", "cash"}

def categorizar_metodo(nombre: str) -> str:
    n = _norm(nombre)
    if n in EFECTIVO_SET:
        return "efectivo"
    if n in TARJETAS_Y_BANCOS or "posnet" in n or "tarjeta" in n:
        return "bancos"
    if n in WALLETS_Y_TRANSFER or "transfer" in n or "mercado" in n or "uala" in n:
        return "transferencias"
    # default: ponelo donde prefieras; yo lo llevo a transferencias
    return "transferencias"
def to_local_date_str_compat(value, out_fmt="%Y-%m-%d") -> str:
    """Acepta str|date|datetime y devuelve string en la tz local."""
    if not value:
        return ""
    # date puro
    if isinstance(value, datetime) is False and hasattr(value, "strftime"):
        return value.strftime(out_fmt)

    # datetime o string
    if not isinstance(value, datetime):
        # 1) tu helper (DD/MM/YYYY[ HH:MM])
        dt = parse_fecha(str(value))
        if dt is None:
            # 2) ISO u otros
            dt = parse_datetime(str(value)) or (
                datetime.fromisoformat(str(value).replace("Z", "+00:00"))
                if "Z" in str(value)
                else None
            )
        if dt is None:
            return str(value)
    else:
        dt = value

    # asegurar aware
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return timezone.localtime(dt).strftime(out_fmt)

# ===== util para sacar metodo/fecha/monto robusto del pago =====
def extraer_info_pago(p):
    # método
    metodo = getattr(p, "metodo_pago", None) or getattr(p, "metodoPago", None) or getattr(p, "metodo", "") or ""
    # monto
    monto = getattr(p, "monto", None)
    if monto is None and hasattr(p, "importe"):
        monto = getattr(p, "importe", 0)
    # fecha
    dt = getattr(p, "fecha", None)
    fstr = to_local_date_str_compat(dt, "%Y-%m-%d")

    return str(metodo or ""), Decimal(monto or 0), fstr

@login_required
@require_GET
def recibo_pago_json(request, pk: int):
    """
    Ahora: dado un pago (pk), buscamos TODOS los pagos de la misma venta + misma cuota,
    armamos un solo recibo con la lista completa y el resumen por categoría (3 buckets).
    """

    pago = get_object_or_404(PagoCannon, pk=pk)

    venta = getattr(pago, "venta", None) or get_object_or_404(Ventas, id=pago.venta_id)

    # normalizar nro_cuota: intentar convertir a int, si no -> None
    nro_cuota_raw = getattr(pago, "nro_cuota", None)
    try:
        nro_cuota_val = int(nro_cuota_raw) if nro_cuota_raw not in (None, "") else None
    except (ValueError, TypeError):
        nro_cuota_val = None

    nro_cuota_display = str(nro_cuota_val) if nro_cuota_val is not None else ""

    print(f"Generando recibo para pago ID={pago.id}, venta ID={venta.id}, cuota={nro_cuota_display}")

    if nro_cuota_val is None:
        pagos_qs = PagoCannon.objects.filter(venta_id=venta.id).order_by("fecha")
    else:
        pagos_qs = PagoCannon.objects.filter(venta_id=venta.id, nro_cuota=nro_cuota_val).order_by("fecha")

    # Datos empresa / cliente
    cliente = venta.nro_cliente
    agencia = venta.agencia

    # Sumas por bucket
    sumas = {"transferencias": Decimal("0"), "bancos": Decimal("0"), "efectivo": Decimal("0")}
    pagos = []

    for p in pagos_qs:
        metodo_raw, monto_p, fecha_p = extraer_info_pago(p)
        cat = categorizar_metodo(metodo_raw)
        sumas[cat] += monto_p
        pagos.append({
            "id": p.pk,
            "fecha": fecha_p,
            "metodo_raw": metodo_raw or "-",
            "metodo_categoria": cat,  # transferencias | bancos | efectivo
            "monto": monto_p,
            "monto_display": _fmt_money(monto_p),
        })

    total = sum(sumas.values())
    # Tomamos número de recibo del último pago (o el actual) — ajustá si querés otro criterio.
    nro_recibo = getattr(pagos_qs.last(), "nro_recibo", "") or getattr(pago, "nro_recibo", "") or ""

    # Fecha del recibo (del último pago, o del actual)
    dt = getattr(pagos_qs.last(), "fecha", None) or getattr(pago, "fecha", None)
    fecha_str = to_local_date_str_compat(dt, "%d/%m/%Y") or datetime.now().strftime("%d/%m/%Y")
    dia_mes_ano = fecha_str.split("/")

    # Helpers HTML para tu template actual (inyectamos HTML seguro desde server)
    pagos_rows_html = "".join(
        f"""
        <div class="fila-pago">
          <span class="pago-cat">{item['metodo_categoria'].title()}</span>
          <span class="pago-detalle">{item['metodo_raw']}</span>
          <div class="linea-pago"></div>
          <span>${item['monto_display']}</span>
        </div>
        """
        for item in pagos
    )

    resumen_rows_html = f"""
      <div class="resumen-metodos">
        <div>Transferencias <strong>${_fmt_money(sumas['transferencias'])}</strong></div>
        <div>Bancos <strong>${_fmt_money(sumas['bancos'])}</strong></div>
        <div>Efectivo <strong>${_fmt_money(sumas['efectivo'])}</strong></div>
      </div>
    """

    ctx = {
        "empresa": {
            "nombre": "ELANEL",
            "sub": "Servicios S.R.L.",
            "direccion": getattr(agencia, "direccion", "") or "",
            "telefono": getattr(agencia, "tel_ref", "") or "",
            "pseudonimo": getattr(agencia, "pseudonimo", "") or "",
            "email": getattr(agencia, "email_ref", "") or "",
            "cuit": "C.U.I.T./D.G.R: 30-71804533-5",
            "leyenda": "DOCUMENTO NO VÁLIDO COMO FACTURA",
            "regimen": "CLIENTE",
        },
        "recibo": {
            "numero_formateado": nro_recibo,
            "fecha_dia": dia_mes_ano[0],
            "fecha_mes": dia_mes_ano[1],
            "fecha_anio": dia_mes_ano[2],
            "importe_letras": convertir_moneda_a_texto(total),
            "total": _fmt_money(total),
            "concepto": f"Pago de cuota N°{nro_cuota_val} ({len(pagos)} pago(s))",
        },
        "cliente": {
            "nombre": getattr(cliente, "nombre", "") or "",
            "domicilio": getattr(cliente, "domic", "") or "",
            "localidad": getattr(cliente, "loc", "") or "",
            "cuit": getattr(cliente, "cuit", "") or "",
        },
        "pagos": pagos,  # por si querés usar bucles en template
        "sumas": {
            "transferencias": _fmt_money(sumas["transferencias"]),
            "bancos": _fmt_money(sumas["bancos"]),
            "efectivo": _fmt_money(sumas["efectivo"]),
        },
        # html “inyectable” (compatible con tu template actual)
        "pagos_html": pagos_rows_html,
        "resumen_html": resumen_rows_html,
        "copias": ["ORIGINAL", "DUPLICADO"],
    }
    return ctx

@login_required
def recibo_preview(request, pk):
    context = recibo_pago_json(request, pk)
    return render(request, "recibo_preview.html", {"context": context})


# Pagar una cuota
def pagarCuota(request):
    print(request.method)
    if request.method == 'POST':
        try:
            # print(request.body)
            data = json.loads(request.body)
            venta = Ventas.objects.get(pk=int(data.get("ventaID")))

            cuotaRequest = data.get("cuota")
            metodoPago = data.get("metodoPago")
            formaPago = data.get("typePayment") # Si es parcial o total
            cobrador = data.get('cobrador')

            monto = 0
            if(formaPago =="total"):
                cuota = list(filter(lambda x:x["cuota"] == cuotaRequest,venta.cuotas))[0]
                print(cuota)
                monto = cuota["total"] - cuota["descuento"]["monto"]
            elif(formaPago =="parcial"):
                monto = data.get('valorParcial')

            cuota_digit = cuotaRequest.split()[-1]
            venta.pagarCuota(cuota_digit,int(monto),metodoPago,cobrador,request.user) #Funcion que paga parcialmente
                
            return JsonResponse({"status": True,"message":f"Pago de {cuotaRequest.lower()} exitosa"}, safe=False)
        except Exception as error:
            print(error)
            # print("sssssssssssssssss")
            return JsonResponse({"status": False,"message":f"Error en el pago de {cuotaRequest.lower()}","detalleError":str(error)}, safe=False)

def solicitudBajaCuota(request):
    if request.method == "POST":
        try: 
            data = json.loads(request.body)
            venta = Ventas.objects.get(pk=int(data.get("ventaID")))
            cuota = data.get("cuota")

            # Generar la URL de autorización
            url_autorizar = request.build_absolute_uri(
                reverse('sales:autorizar_baja_cuota', args=[venta.id, cuota])
            )  

            subject = "Solicitud de Cancelación de Pago"
            template = "email_autorizacion_baja_cuota.html"
            context = {
                "venta": venta.nro_operacion, 
                "cuota": cuota, 
                "url_autorizar": url_autorizar
            }
            from_email = settings.EMAIL_HOST_USER
            to_email = "lautaro.rodriguez553@gmail.com"

            send_html_email(subject, template, context, from_email, to_email)

            response_data = {
                "message": "Solicitud enviada exitosamente",
                "iconMessage": "/static/images/icons/checkMark.svg",
                "status": True
            }
            return JsonResponse(response_data)
        except Exception as error:
            print(error)
            response_data = {
                "message": "Error al solicitar la cancelacion de la cuota",
                "iconMessage": "/static/images/icons/error_icon.svg",
                "status": False
            }
            return JsonResponse(response_data)
   
def darAutorizacionBajaCuota(request, ventaID, cuota):
    venta = get_object_or_404(Ventas, id=ventaID)  # Obtiene la venta

    # Aquí puedes actualizar un campo o lista en la BD para indicar que la cuota está autorizada
    for c in venta.cuotas:  # Si `cuotas` es una lista
        if c["cuota"] == cuota:
            c["autorizada_para_anular"] = True  # Agregar un flag en la cuota
            venta.save()
            break  # Termina el loop una vez encontrada la cuota

    return redirect("sales:pagina_confirmacion_baja_cuota")  # Redirige a una página de confirmación

def darBajaCuota(request):
    try: 
        data = json.loads(request.body)
        venta = Ventas.objects.get(pk=int(data.get("ventaID")))
        cuota = data.get("cuota").strip()[-1]  # Extrae el número de la cuota

        venta.anularCuota(cuota)  # Anula la cuota

        response_data = {
                "message": "Cuota anulada exitosamente",
                "iconMessage": "/static/images/icons/checkMark.svg",
                "status": True
        } 
        return JsonResponse(response_data)

    except Exception as error:
        print(error)
        response_data = {
                "message": "Error al anular la cuota",
                "iconMessage": "/static/images/icons/error_icon.svg",
                "status": False
        }
        return JsonResponse(response_data)

def pagina_confirmacion(request):
    return render(request, "confirmacion_baja_cuota.html")
    

#Dar de baja una venta
def darBaja(request, pk):
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
            porcentage         = data.get("porcentage")
            motivoDetalle      = data.get("motivo")
            motivoDescripcion  = data.get("motivoDescripcion")
            responsable        = request.user.nombre

            venta = get_object_or_404(Ventas, id=pk)

            porcentajeEsperado = venta.porcentajeADevolver()
            puedeModificar     = request.session.get("statusKeyPorcentajeBaja", False)
            porcentajeFinal    = porcentage if puedeModificar else porcentajeEsperado

            venta.darBaja("cliente", porcentajeFinal, motivoDetalle, motivoDescripcion, responsable)
            venta.save()

            # invalidar la habilitación una vez usada
            request.session["statusKeyPorcentajeBaja"] = False

            return JsonResponse({
                'status': True,
                'urlPDF':  reverse("sales:bajaPDF", args=[pk]),
                'urlUser': reverse("users:cuentaUser", args=[venta.nro_cliente.pk])
            })

        except Exception as error:
            return JsonResponse({'status': False, 'message': str(error)}, status=400)
#endregion 


class PlanRecupero(generic.DetailView):
    model = Ventas
    template_name = "plan_recupero.html"

    def get(self,request,*args, **kwargs):
        self.object = self.get_object()
        url = request.path
        cuotasPagadas = self.object.cuotas_pagadas()

        sucursal = request.user.sucursales.all()[0]

        vendedores = Usuario.objects.filter(sucursales__in=[sucursal], rango__in=["Vendedor","Supervisor"])
        supervisores = Usuario.objects.filter(sucursales__in = [sucursal], rango="Supervisor")

        # Suma las cuotas pagadas para calcular el total a adjudicar
        valoresCuotasPagadas = [item["pagado"] for item in cuotasPagadas]
        sumaCuotasPagadas = sum(valoresCuotasPagadas)
        campaniaActual = getCampaniaActual()

        #region Para determinar si se habilita la campaña anterior
        fechaActual = datetime.datetime.now()
        ultimo_dia_mes_pasado = datetime.datetime.now().replace(day=1) - relativedelta(days=1)
        diferencia_dias = (fechaActual - ultimo_dia_mes_pasado).days

        context = {
            'venta': self.object,
            'agencias': request.user.sucursales.all(), 
            'agenciaActual': request.user.sucursales.all()[0],
            'campania': campaniaActual,
            'sumaCuotasPagadas' : int(sumaCuotasPagadas),
            'autorizado_por':  Sucursal.objects.get(pseudonimo = request.user.sucursales.all()[0].pseudonimo).gerente.nombre,
            'cantidad_cuotas_pagadas': len(cuotasPagadas),
            'idCliente': self.object.nro_cliente.nro_cliente,
            'vendedores': vendedores, 
            'supervisores': supervisores, 
        }
        
        return render(request,self.template_name,context)
    

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form =json.loads(request.body)
        errors ={}
        sale = Ventas()

        
        # Para guardar como objeto Producto
        producto = form["producto"]
        if producto and not Products.objects.filter(nombre=producto,activo=True).exists():
            errors['producto'] = 'Producto invalido.' 
        else:
            producto = Products.objects.get(nombre=producto)
            sale.producto = producto

        # Validar la sucursal
        agencia = form["agencia"] 

        if agencia and not Sucursal.objects.filter(pseudonimo=agencia).exists():
            errors['agencia'] = 'Agencia invalida.'
        else:
            agencia = Sucursal.objects.get(pseudonimo=agencia)
            sale.agencia = agencia

        
        # Comprobar el vendendor
        vendedor = form['vendedor']
        if  not Usuario.objects.filter(nombre__iexact=vendedor).exists():
            errors['vendedor'] = 'Vendedor invalido.' 
        else:
            vendedor_instance = Usuario.objects.get(nombre__iexact=form['vendedor'])
            sale.vendedor = vendedor_instance

        # Comprobar el supervisor
        supervisor = form['supervisor']
        if not Usuario.objects.filter(nombre__iexact=supervisor).exists():
            errors['supervisor'] = 'Supervisor invalido.' 
        else:
            supervisor_instance = Usuario.objects.get(nombre__iexact=form['supervisor'])
            sale.supervisor = supervisor_instance




        sale.nro_cliente = Cliente.objects.get(nro_cliente__iexact=self.get_object().nro_cliente.nro_cliente)
        
        sale.nro_contrato = form['nro_contrato']
        sale.modalidad = form['modalidad'] if form['modalidad'] else ""
        sale.importe = int(form['importe'])
        sale.primer_cuota = int(form['primer_cuota']) if form['primer_cuota'] else 0
        sale.anticipo = int(form['anticipo']) if form['anticipo'] else 0
        sale.tasa_interes = float(form['tasa_interes']) if form['tasa_interes'] else 0
        sale.intereses_generados = int(form['intereses_generados']) if form['intereses_generados'] else 0
        sale.importe_x_cuota = int(form['importe_x_cuota']) if form['importe_x_cuota'] else 0
        sale.nro_cuotas = int(form['nro_cuotas'])
        sale.total_a_pagar = int(form['total_a_pagar']) if form['total_a_pagar'] else 0
        sale.fecha = form['fecha']
        sale.tipo_producto = form['tipo_producto']
        sale.paquete = form['paquete']
        sale.nro_orden = form['nro_orden']
        sale.campania = form['campania']
        sale.observaciones = form['observaciones']
        
        try:
            sale.full_clean()
        except ValidationError as e:
            errors.update(e.message_dict)
       
        if len(errors) != 0:
            print(errors)
            return JsonResponse({'success': False, 'errors': errors}, safe=False)
        else:
            sale.fecha = form['fecha'] + " 00:00"
            sale.crearCuotas()
            sale.save()

            # Dar de baja la venta a la que se le aplico el plan de recupero
            self.object.planRecupero("plan recupero",request.user.nombre, sale.observaciones,sale.pk)
            self.object.save()
            return JsonResponse({'success': True,'urlRedirect': reverse('users:cuentaUser',args=[sale.nro_cliente.pk])}, safe=False)

@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch') # Para no guardar el cache 
class CreateAdjudicacion(TestLogin,generic.DetailView):
    model = Ventas
    template_name = "create_adjudicacion.html"
    
    def get(self,request,*args, **kwargs):
        self.object = self.get_object()
        url = request.path
        sumaDePagos = 0
        pagos_de_venta = PagoCannon.objects.filter(venta = self.object)

        for pago in pagos_de_venta:
            if pago.nro_cuota != 0:
                sumaDePagos += pago.monto


        tipoDeAdjudicacion = "NEGOCIACIÓN" if "negociacion" in url else "SORTEO"
        
        intereses = CoeficientesListadePrecios.objects.all()
        
        context = {
            'venta': self.object,
            'intereses' : intereses,
            'agencias': Sucursal.objects.all(),
            'sumaCuotasPagadas' : int(sumaDePagos),
            'autorizado_por':  Sucursal.objects.get(pseudonimo = request.user.sucursales.all()[0].pseudonimo).gerente.nombre,
            'cantidad_cuotas_pagadas': len(self.object.cuotas_pagadas()),
            'tipoDeAdjudicacion' : tipoDeAdjudicacion,
            'idCliente': self.object.nro_cliente.nro_cliente,
        }
        
        return render(request,self.template_name,context)
    

    def post(self, request, *args, **kwargs):

        form =json.loads(request.body)
        self.object = self.get_object()
        numeroOperacion = self.object.nro_operacion
        errors ={}
        cuotas_consideradas = PagoCannon.objects.filter(venta=self.object).exclude(nro_cuota=0)


        # Suma las cuotas pagadas para calcular el total a adjudicar
        sumaDePagos = sum([cuota.monto for cuota in cuotas_consideradas])

        sale = Ventas()

        # Obtenemos el tipo de adjudicacion - - - - - - - - - - - - - - - - - - - -
        url = request.path
        tipo_adjudicacion = "sorteo" if "sorteo" in url else "negociacion"

        # Validar el producto
        producto = form['producto']
        print("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
        print(producto)
        if producto and not Products.objects.filter(id=producto,activo=True).exists():
            errors['producto'] = 'Producto invalido.' 

        elif producto:
            print(Products.objects.filter(id=producto,activo=True).first().nombre)
            producto = Products.objects.get(id=int(producto))
            sale.producto = producto



        # Validar la sucursal
        agencia = form['agencia']
        if agencia and not Sucursal.objects.filter(id=agencia).exists():
            errors['agencia'] = 'Agencia invalida.' 

        elif agencia:
            agencia = Sucursal.objects.get(id=agencia)
            sale.agencia = agencia

        contratoAdjudicado = form["nro_contrato"]
        sale.adjudicado["status"] = True
    
        sale.nro_cliente = Cliente.objects.get(pk=request.session["venta"]["nro_cliente"])
        sale.nro_operacion = request.session["venta"]["nro_operacion"]
        sale.vendedor = Usuario.objects.get(pk = request.session["venta"]["vendedor"])
        sale.supervisor = Usuario.objects.get(pk = request.session["venta"]["supervisor"])
        sale.cantidadContratos = request.session["venta"]["cantidadContratos"]
        sale.paquete = str(request.session["venta"]["paquete"])
        sale.campania = getCampaniaActual()
        sale.importe = int(form['importe'])
        sale.modalidad = form['modalidad']
        sale.nro_cuotas = int(form['nro_cuotas'])
        sale.tasa_interes = int(form['tasa_interes'])
        sale.intereses_generados = int(form['intereses_generados'])
        sale.importe_x_cuota = int(form['importe_x_cuota'])
        sale.total_a_pagar = int(form['total_a_pagar'])
        sale.tipo_producto = form['tipo_producto']
        sale.fecha = form['fecha']

        sale.observaciones = form['observaciones']
        sale.fecha = form['fecha'] + " 00:00"

        sale.setDefaultFields()
        sale.crearCuotas() # Crea las cuotas
        if(tipo_adjudicacion == "sorteo"):
            sale.acreditarCuotasPorAnticipo(sumaDePagos,request.user)
        sale.crearAdjudicacion(contratoAdjudicado,numeroOperacion,tipo_adjudicacion) # Crea la adjudicacion eliminando la cuota 0
        


        try:
            sale.full_clean()
        except ValidationError as e:
            errors.update(e.message_dict)
       
        if len(errors) != 0:
            print(errors)
            return JsonResponse({'success': False, 'errors': errors}, safe=False)  
        else:
            sale.save()
            
            self.object.darBaja("adjudicacion",0,"","",request.user.nombre) # Da de baja la venta que fue adjudicada
            self.object.save()

            
            # #region Para enviar el correo
            # subject = 'Se envio una adjudicacion'
            # template = 'adjudicacion_correo.html'
            # context = {'nombre': 'Usuario'}  # Contexto para renderizar el template
            # from_email = 'lautaror@elanelsys.com'
            # to_email = 'lautaro.rodriguez553@gmail.com'

            # send_html_email(subject, template, context, from_email, to_email)
            #endregion
            return JsonResponse({'success': True,'urlRedirect':reverse_lazy('users:cuentaUser',args=[request.session["venta"]["nro_cliente"]])}, safe=False)          
        

class ChangePack(TestLogin,generic.DetailView):
    model = Ventas
    template_name = "change_pack.html"

    def get(self,request,*args, **kwargs):
        self.object = self.get_object()

        customers = Cliente.objects.all()
        products = Products.objects.all()
        sucursal = request.user.sucursales.all()[0]

        vendedores = Usuario.objects.filter(sucursales__in=[sucursal], rango__in=["Vendedor","Supervisor"])
        supervisores = Usuario.objects.filter(sucursales__in = [sucursal], rango="Supervisor")


        cuotasPagadas_parciales = self.object.cuotas_pagadas() + self.object.cuotas_parciales()
        sumaDePagos = 0

        # Suma las cuotas pagadas para calcular el total a adjudicar
        for cuota in cuotasPagadas_parciales:
            print("Cuota:")
            print(cuota)
            for pago_id in cuota["pagos"]:
                pago_instance = PagoCannon.objects.get(id=pago_id)
                sumaDePagos += pago_instance.monto

        contratosAsociados = ""
        for contrato in self.object.cantidadContratos:
            contratosAsociados += f" {contrato['nro_contrato']} -"
        contratosAsociados = contratosAsociados[:-1] # Elimina el ultimo guion
        
        context = {
            'venta': self.object,
            'agencias': Sucursal.objects.all(),
            "contratosAsociados": contratosAsociados,
            'sumaCuotasPagadas' : int(sumaDePagos),
            'cantidad_cuotas_pagadas': len(self.object.cuotas_pagadas()),
            'idCliente': self.object.nro_cliente.nro_cliente,
            'customers': customers,
            'vendedores': vendedores, 
            'supervisores': supervisores, 
            'products': products, 
        }


        return render(request,self.template_name,context)
    

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form =json.loads(request.body)
        errors ={}
        cuotasPagadas_parciales = self.object.cuotas_pagadas() + self.object.cuotas_parciales()

        sumaDePagos = 0
        for cuota in cuotasPagadas_parciales:
            pagos = cuota["pagos"]
            sumaDePagos += sum([pago["monto"] for pago in pagos])

        sale = Ventas()

        # Validar el producto
        producto = form['producto']

        if producto and not Products.objects.filter(nombre=producto,activo=True).exists():
            errors['producto'] = 'Producto invalido.' 

        elif producto:
            producto = Products.objects.get(nombre=producto)
            sale.producto = producto
    
        # Para guardar la cantidad de contratos que se haga
        chances = []
        chance_counter = 1
        while f'nro_contrato_{chance_counter}' in form:

            # Obtenemos y validamos el nro de contrato
            nro_contrato = form.get(f'nro_contrato_{chance_counter}')
            if not re.match(r'^\d+$', nro_contrato):
                raise ValidationError({f'nro_contrato_{chance_counter}': 'Debe contener solo números.'})
            
            # Obtenemos y validamos el nro de orden
            nro_orden = form.get(f'nro_orden_{chance_counter}')
            if not re.match(r'^\d+$', nro_orden):
                raise ValidationError({f'nro_orden_{chance_counter}': 'Debe contener solo números.'})
            
            # Si ambos campos son validos, los añadimos a la lista de chances
            if nro_contrato and nro_orden:
                chances.append({
                    'nro_contrato': nro_contrato,
                    'nro_orden': nro_orden
                })
            chance_counter += 1

        # Guardar las chances en el campo JSONField
        sale.cantidadContratos = chances

        sale.nro_cliente = Cliente.objects.get(pk=request.session["venta"]["nro_cliente"])
        sale.nro_operacion = request.session["venta"]["nro_operacion"]
        sale.agencia = self.object.agencia
        sale.vendedor = Usuario.objects.get(pk = request.session["venta"]["vendedor"])
        sale.supervisor = Usuario.objects.get(pk = request.session["venta"]["supervisor"])
        sale.paquete = str(request.session["venta"]["paquete"])
        sale.campania = getCampaniaActual()
        sale.importe = int(form['importe'])
        sale.modalidad = form['modalidad']
        sale.nro_cuotas = int(form['nro_cuotas'])
        sale.tasa_interes = float(form['tasa_interes'])
        sale.intereses_generados = int(form['intereses_generados'])
        sale.importe_x_cuota = int(form['importe_x_cuota'])
        sale.total_a_pagar = int(form['total_a_pagar'])
        sale.primer_cuota = int(form['primer_cuota'])
        sale.anticipo = int(form['anticipo'])
        sale.fecha = form['fecha']
        sale.tipo_producto = form['tipo_producto']
        sale.observaciones = form['observaciones']
        
        try:
            sale.full_clean()
        except ValidationError as e:
            errors.update(e.message_dict)
       
        if len(errors) != 0:
            return JsonResponse({'success': False, 'errors': errors}, safe=False) 
        else:
            sale.fecha = form['fecha'] + " 00:00"
            sale.crearCuotas()
            sale.acreditarCuotasPorAnticipo(sumaDePagos,request.user.nombre)
            sale.save()

            self.object.darBaja("cambio de pack",0,"","",request.user.nombre) # Da de baja la venta que fue cambiada de pack
            self.object.save()
            return JsonResponse({'success': True,'urlRedirect':reverse_lazy('sales:detail_sale',args=[sale.pk])}, safe=False)


class ChangeTitularidad(TestLogin,generic.DetailView):
    template_name = 'changeTitularidad.html'
    model = Ventas


    def get(self,request,*args, **kwargs):
        self.object = self.get_object()
        customers = Cliente.objects.all()
        context = {
            "customers": customers,
            "object": self.object,
        }
        return render(request, self.template_name,context)

    def post(self,request,*args,**kwargs):
        self.object = self.get_object()
        form = json.loads(request.body)
        newCustomer = form["customer"]

        dniNewCustomer = Cliente.objects.all().filter(dni=newCustomer)[0]
        if(dniNewCustomer == self.object.nro_cliente):
            return JsonResponse({'success': False, 'errors': "EL CLIENTE NUEVO NO PUEDE SER IGUAL AL ANTIGUO"}, safe=False)
        else:
            # Coloca los datos del cambio de titularidad
            cuotasPagadas_parciales = self.object.cuotas_pagadas() + self.object.cuotas_parciales()
            lastCuota = getCuotasPagadasSinCredito(cuotasPagadas_parciales)[-1]

            self.object.createCambioTitularidad(lastCuota,request.user.nombre,self.object.nro_cliente.nombre,dniNewCustomer.nombre,self.object.nro_cliente.pk,dniNewCustomer.pk)
            
            # Actualiza el dueño de la venta
            self.object.nro_cliente = dniNewCustomer
            self.object.save()
            return JsonResponse({'success': True,'urlRedirect': reverse("sales:detail_sale",args = [self.get_object().pk])}, safe=False) 

# ========= Helpers =========

def _parse_ddmmyyyy(s: str):
    """Devuelve el string normalizado dd/mm/yyyy si luce como fecha, o None."""
    s = (s or "").strip()
    if not s:
        return None
    # aceptamos dd/mm/yyyy o dd-mm-yyyy
    import re
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if not m:
        return None
    d, mth, y = m.groups()
    d = d.zfill(2); mth = mth.zfill(2)
    return f"{d}/{mth}/{y}"

def _build_search_q(search: str) -> Q:
    """Búsqueda flexible: nombre, DNI, telé, loc/prov, nro_op, producto, fecha (string)."""
    search = (search or "").strip()
    if not search:
        return Q()
    q = Q()
    # texto directo
    q |= Q(nro_cliente__nombre__icontains=search)
    q |= Q(producto__nombre__icontains=search)
    q |= Q(nro_cliente__prov__icontains=search) | Q(nro_cliente__loc__icontains=search) | Q(nro_cliente__domic__icontains=search)
    # dígitos (dni/tel/nro operación)
    digits = "".join(ch for ch in search if ch.isdigit())
    if digits:
        q |= Q(nro_cliente__dni__icontains=digits)
        q |= Q(nro_cliente__tel__icontains=digits)
        q |= Q(nro_operacion__icontains=digits)
    # fecha como string (dd/mm/yyyy)
    norm = _parse_ddmmyyyy(search)
    if norm:
        q |= Q(fecha__icontains=norm)
    else:
        # también dejamos substring por si vino "05/2024" o similar
        q |= Q(fecha__icontains=search)
    return q

def _estado_map(value: str):
    """Mapea lo que venga ('p','pendientes', etc.) a una clave uniforme."""
    v = (value or "").strip().lower()
    if v in ("p", "pendientes", "pendiente"): return "pendientes"
    if v in ("r", "realizadas", "realizada"): return "realizadas"
    if v in ("a", "aprobadas", "aprobada"):   return "aprobadas"
    if v in ("d", "desaprobadas", "desaprobada"): return "desaprobadas"
    return ""

def _status_for(aud_count: int, last_grade):
    """Arma statusText/Icon a partir de cantidad y última grade (True/False/None)."""
    if not aud_count:
        return {"statusClean": "p","statusText": "Pendiente", "statusIcon": "/static/images/icons/pending.svg"}
    if last_grade is True:
        return {"statusClean": "a","statusText": "Aprobada", "statusIcon": "/static/images/icons/checkMark.svg"}
    return {"statusClean": "d", "statusText": "Desaprobada", "statusIcon": "/static/images/icons/error_icon.svg"}

def _venta_dict(v, aud_count=None, last_grade=None):
    c = v.nro_cliente

    # status
    if not aud_count:
        st = {"statusClean": "p","statusText": "Pendiente", "statusIcon": "/static/images/icons/pending.svg"}
    elif last_grade is True:
        st = {"statusClean": "a","statusText": "Aprobada", "statusIcon": "/static/images/icons/checkMark.svg"}
    else:
        st = {"statusClean": "d","statusText": "Desaprobada", "statusIcon": "/static/images/icons/error_icon.svg"}

    # historial (usar prefetch si está disponible)
    aud_list = getattr(v, "auditorias_pref", None)
    if aud_list is not None:
        auds = [
            {"version": a.version, "grade": a.grade, "comentarios": a.comentarios, "fecha_hora": a.fecha_hora}
            for a in aud_list
        ]
    else:
        auds = list(
            v.auditorias.order_by("version").values("version", "grade", "comentarios", "fecha_hora")
        )

    return {
        "id": v.id,
        "statusText": st["statusText"],
        "statusClean": st["statusClean"],
        "statusIcon": st["statusIcon"],
        "nombre": getattr(c, "nombre", ""),
        "dni": str(getattr(c, "dni", "") or ""),
        "nro_operacion": v.nro_operacion,
        "fecha": v.fecha or "",
        "tel": str(getattr(c, "tel", "") or ""),
        "loc": getattr(c, "loc", "") or "",
        "cod_postal": str(getattr(c, "cod_postal", "") or ""),
        "prov": getattr(c, "prov", "") or "",
        "domic": getattr(c, "domic", "") or "",
        "vendedor": getattr(getattr(v, "vendedor", None), "nombre", "") or "",
        "supervisor": getattr(getattr(v, "supervisor", None), "nombre", "") or "",
        "producto": getattr(getattr(v, "producto", None), "nombre", "") or "",
        "campania": v.campania or "",
        "auditorias": auds,
    }

def _extract_params(request):
    """Lee de GET o JSON-POST. Acepta: page, page_size, search, campania, sucursal_id, agencia_id, estado"""
    if request.method == "POST":
        import json
        try:
            data = json.loads(request.body or "{}")
        except Exception:
            data = {}
        getv = lambda k, d=None: data.get(k, d)
    else:
        getv = lambda k, d=None: request.GET.get(k, d)

    page      = int(getv("page", 1) or 1)
    page_size = int(getv("page_size", 20) or 20)
    search    = (getv("search", "") or "").strip()
    campania  = (getv("campania", "") or "").strip()
    suc_id    = getv("sucursal_id") or getv("sucursal")  # alias
    ag_id     = getv("agencia_id")
    estado    = _estado_map(getv("estado", ""))

    return {
        "page": page, "page_size": page_size, "search": search, "campania": campania,
        "sucursal_id": suc_id, "agencia_id": ag_id, "estado": estado
    }

# ========= API de auditorías (listado/filtrado + contadores) =========

@login_required
def auditorias_api(request):
    """
    JSON: results, total, page, page_size, has_next, has_previous, counts{...}
    Filtros: campania, sucursal_id/agencia_id, estado, search.
    """
    p = _extract_params(request)

    base = (
        Ventas.objects
        .select_related("nro_cliente", "vendedor", "supervisor", "producto", "agencia")
    )

    if p["campania"]:
        base = base.filter(campania=p["campania"])

    # sucursal/agencia
    for key in ("sucursal_id", "agencia_id"):
        if p[key]:
            try:
                base = base.filter(agencia_id=int(p[key]))
            except ValueError:
                pass

    if p["search"]:
        base = base.filter(_build_search_q(p["search"]))

    # Última grade por subquery (versión más alta)
    latest_grade_sq = Subquery(
        Auditoria.objects.filter(venta=OuterRef("pk"))
        .order_by("-version").values("grade")[:1]
    )

    # len(cantidadContratos) en SQL (PostgreSQL): jsonb_array_length(...)
    contratos_len_sql = Coalesce(
        # jsonb_array_length(cantidadContratos)
        Func(F("cantidadContratos"), function="jsonb_array_length", output_field=IntegerField()),
        Value(0),
        output_field=IntegerField()
    )

    # Anotaciones base para listado + contadores
    ann = base.annotate(
        aud_count=Count("auditorias"),
        last_grade=latest_grade_sq,
        contratos_len=contratos_len_sql,
    )

    # ---- Contadores 100% en DB (un solo query) ----
    agg = ann.aggregate(
        pendientes=Coalesce(Sum("contratos_len", filter=Q(aud_count=0)), Value(0)),
        realizadas=Coalesce(Sum("contratos_len", filter=Q(aud_count__gt=0)), Value(0)),
        aprobadas=Coalesce(Sum("contratos_len", filter=Q(aud_count__gt=0, last_grade=True)), Value(0)),
        desaprobadas=Coalesce(Sum("contratos_len", filter=Q(aud_count__gt=0, last_grade=False)), Value(0)),
    )
    counts = {
        "pendientes": int(agg["pendientes"] or 0),
        "realizadas": int(agg["realizadas"] or 0),
        "aprobadas": int(agg["aprobadas"] or 0),
        "desaprobadas": int(agg["desaprobadas"] or 0),
    }

    # ---- Filtro por estado para el listado ----
    qs = ann
    if p["estado"] == "pendientes":
        qs = qs.filter(aud_count=0)
    elif p["estado"] == "realizadas":
        qs = qs.filter(aud_count__gt=0)
    elif p["estado"] == "aprobadas":
        qs = qs.filter(aud_count__gt=0, last_grade=True)
    elif p["estado"] == "desaprobadas":
        qs = qs.filter(aud_count__gt=0, last_grade=False)

    # Orden: si querés más rapidez y 'fecha' es string, usá -id. Si necesitás por fecha string, dejá -fecha.
    qs = qs.order_by("-id")

    # Prefetch SOLO para la página actual (evita N+1 del historial)
    auditorias_qs = (
        Auditoria.objects
        .only("venta_id", "version", "grade", "comentarios", "fecha_hora")
        .order_by("version")
    )
    qs = qs.prefetch_related(Prefetch("auditorias", queryset=auditorias_qs, to_attr="auditorias_pref"))

    paginator = Paginator(qs, p["page_size"])
    page_obj = paginator.get_page(p["page"])

    results = [
        _venta_dict(
            v,
            aud_count=getattr(v, "aud_count", 0),
            last_grade=getattr(v, "last_grade", None)
        )
        for v in page_obj.object_list
    ]

    return JsonResponse({
        "results": results,
        "total": paginator.count,
        "page": page_obj.number,
        "page_size": p["page_size"],
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
        "counts": counts,
    })

# ========= Alta/edición de auditoría (crea nueva versión) =========

@login_required
def crear_auditoria(request):
    """POST JSON: { venta_id, grade: 'a'|'d', comentarios } -> crea nueva versión de Auditoria."""
    if request.method != "POST":
        return JsonResponse({"status": False, "message": "Método no permitido"}, status=405)

    import json
    try:
        data = json.loads(request.body or "{}")
    except Exception:
        data = {}
    print(f"POST {request}")
    venta_id   = data.get("venta_id") or data.get("idVenta")
    grade_code = (data.get("grade") or "").strip().lower()
    comentarios = data.get("comentarios") or data.get("comentarioInput") or ""

    if not venta_id or grade_code not in ("a", "d"):
        return JsonResponse({"status": False, "message": "Parámetros inválidos"}, status=400)

    venta = get_object_or_404(Ventas, pk=int(venta_id))

    # próxima versión
    last_version = Auditoria.objects.filter(venta=venta).aggregate(m=Max("version"))["m"] or 0
    next_version = last_version + 1

    audit = Auditoria.objects.create(
        venta=venta,
        version=next_version,
        grade=(grade_code == "a"),
        comentarios=comentarios,
        # fecha_hora: lo pone el default now_formatted()
    )

    # armo respuesta para refrescar card
    # recalcular status con última grade
    latest_grade = Auditoria.objects.filter(venta=venta).order_by("-version").values_list("grade", flat=True).first()
    aud_count = Auditoria.objects.filter(venta=venta).count()
    st = _status_for(aud_count, latest_grade)

    return JsonResponse({
        "status": True,
        "message": f"Auditoría de la venta {venta.id} registrada",
        "ventaUpdated_id": str(venta.id),
        "auditorias": list(venta.auditorias.order_by("version").values("version", "grade", "comentarios", "fecha_hora")),
        "statusIcon": st["statusIcon"],
        "statusText": st["statusText"],
        "iconMessage": "/static/images/icons/checkMark.svg",
    })

# ========= Página =========

@method_decorator(login_required, name="dispatch")
class PostVenta(generic.View):
    template_name = "postVenta.html"

    def get(self, request, *args, **kwargs):
        campania_actual = getCampaniaActual()  # tu helper
        suc_default = Sucursal.objects.filter(pseudonimo="Resistencia, Chaco").first()
        context = {
            "campania_actual": campania_actual,
            "sucursales": Sucursal.objects.all(),
            "sucursalDefault": suc_default,
            "campanias": getTodasCampaniasDesdeInicio(),
            # La lista y contadores ahora los trae el front via auditorias_api
        }
        return render(request, self.template_name, context)

class PanelVentasSuspendidas(TestLogin,generic.View):
    template_name = 'panel_ventas_suspendidas.html'

    def get(self,request,*args,**kwargs):
        context = {}
        return render(request, self.template_name, context)
    
    def post(self,request,*args,**kwargs):
        data = json.loads(request.body)
        sucursal = request.user.sucursales.all()[0]
        saldo_Afavor = 0
        try:
            venta = Ventas.objects.get(nro_operacion=data["nro_operacion"], suspendida=True, agencia=sucursal)
            for cuotaPagada in venta.cuotas_pagadas():
                pagosDeCuota = cuotaPagada["pagos"]
                saldo_Afavor += sum(pago["monto"] for pago in pagosDeCuota)

            cuotas_pagadas = len(venta.cuotas_pagadas())
            cuotas_atrasadas = len([cuota for cuota in venta.cuotas if cuota["status"] == "Atrasado"])

            context = {
                "cliente": venta.nro_cliente.nombre,
                "tipo_producto": venta.producto.tipo_de_producto,
                "producto": venta.producto.nombre,
                "importe": venta.producto.plan.valor_nominal,
                "nro_orden": 1,
                "fecha_inscripcion": venta.fecha,
                "nro_operacion": venta.nro_operacion,
                "cuotas_atrasadas": cuotas_atrasadas,
                "saldo_Afavor": saldo_Afavor,
                "cuotas_pagadas": cuotas_pagadas,
                "urlSimularPlanRecupero": reverse("sales:simuladorPlanrecupero",args=[venta.pk]),

            }
            return JsonResponse({"status": True,"venta":context}, safe=False)
        except Exception as e:
            print(e)
            return JsonResponse({"status": True,"venta":None}, safe=False)


class SimuladorPlanRecupero(TestLogin,generic.DetailView):
    template_name = "simulador_plan_recupero.html"
    model = Ventas

    def get(self,request,*args,**kwargs):
        self.object = self.get_object()

        cuotasPagadas = self.object.cuotas_pagadas()

        # Suma las cuotas pagadas para calcular el total a adjudicar
        valoresCuotasPagadas = [item["pagado"] for item in cuotasPagadas]
        sumaCuotasPagadas = sum(valoresCuotasPagadas)

        context = {
            'venta' : self.object,
            'sumaCuotasPagadas' : int(sumaCuotasPagadas),
            'cantidad_cuotas_pagadas': len(cuotasPagadas),
        }
        return render(request, self.template_name, context)


class VentasComisionables(generic.View):
    """
    Muestra todas las Ventas de la sucursal y campaña de sesión, 
    o con pagos en esa campaña, y permite toggle AJAX de is_commissionable.
    """
    def get(self, request, *args, **kwargs):
        campania = request.session.get('campania_notCommissionable')
        agencia_id = request.session.get('sucursal_notCommissionable')
        agenciaObject = Sucursal.objects.get(id=agencia_id) if agencia_id else None
        if not campania or not agencia_id:
            return render(request, 'error.html', {
                'msg': "Debes seleccionar antes campaña y sucursal."
            })
        print(f"Agencia: {agenciaObject.pseudonimo} - Campania: {campania}")

        qs = Ventas.objects.filter(agencia=agenciaObject).filter((
                Q(campania=campania) |
                Q(pagos_cannon__campana_de_pago=campania)
            )).distinct().select_related(
            'nro_cliente', 'vendedor', 'supervisor'
        ).prefetch_related('auditorias').order_by('-nro_operacion')
        
        contextVentas = []
        for venta in qs:
            ventaStatus = getEstadoVenta2(venta)
            auditoria = venta.auditorias.last()
            grado_auditoria =""
            comentario_auditoria=""

            if (auditoria):
                grado_auditoria =  "Aprobada" if auditoria.grade else "Desaprobada"
                comentario_auditoria = auditoria.comentarios if auditoria.comentarios else "-"

            contextVentas.append(
                {
                    'id': venta.id,
                    'is_commissionable': venta.is_commissionable,
                    'nro_operacion': venta.nro_operacion,
                    'cliente': venta.nro_cliente,
                    'agencia': venta.agencia.pseudonimo,
                    'campania': venta.campania,
                    'estado': ventaStatus["status"],
                    'motivo': ventaStatus["motivo"] if ventaStatus["motivo"] else "-",
                    'auditoria_grado': grado_auditoria,
                    'auditoria_comentarios':comentario_auditoria,
                }
            )
        return render(request, 'ventas_comisionables.html', {
            'ventas': contextVentas,
            'cantidadVentas_involucradas': len(contextVentas),
            'campania': campania,
            'sucursal': agenciaObject.pseudonimo,
        })


def toggle_comisionable(request):
    """
    Recibe JSON {"id": <venta_id>, "value": true|false}
    y actualiza is_commissionable de esa venta.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            vid = int(data['id'])
            val = bool(data['value'])
        except Exception as e:
            print(e)
            return JsonResponse({'status':False}, status=404)

        updated = Ventas.objects.filter(pk=vid).update(is_commissionable=val)
        if updated:
            return JsonResponse({'status':True, 'id': vid, 'value': val})
        return JsonResponse({'status':False}, status=404)
#endregion - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

#region PDFs - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
def viewsPDFBaja(request,pk):
    operacionBaja = Ventas.objects.get(id=pk)
    context ={
                "nroContrato":operacionBaja.nro_contrato,
                "cliente": operacionBaja.nro_cliente.nombre,
                "domicilio": operacionBaja.nro_cliente.domic,
                "localidad": operacionBaja.nro_cliente.loc,
                "pack": operacionBaja.producto.plan.tipodePlan,
                "producto": operacionBaja.producto.nombre,
                "cantCuotasPagadas" : len(operacionBaja.cuotas_pagadas()),
                "cuotas" : operacionBaja.nro_cuotas,
                "agencia" : f'{operacionBaja.agencia.localidad}, {operacionBaja.agencia.provincia}',
                "motivo" : operacionBaja.deBaja["detalleMotivo"],
                "observacion" : operacionBaja.deBaja["observacion"],
                "dineroDevolver" : operacionBaja.calcularDineroADevolver(),
                "fecha" : operacionBaja.deBaja["fecha"],
            }
            
    bajaName = f'baja_venta_nro_contrato_{str(context["nroContrato"])}'
    urlPDF= os.path.join(settings.PDF_STORAGE_DIR, "baja.pdf")
    
    printPDFBaja(context,request.build_absolute_uri(),urlPDF)

    with open(urlPDF, 'rb') as pdf_file:
        response = HttpResponse(pdf_file,content_type="application/pdf")
        response['Content-Disposition'] = 'inline; filename='+bajaName+'.pdf'
        return response

def viewPDFTitularidad(request,pk,idCambio):
    operacionTitu = Ventas.objects.get(id=pk)
    newCustomer = operacionTitu.cambioTitularidadField[idCambio]["pkNewCustomer"]
    oldCustomer = operacionTitu.cambioTitularidadField[idCambio]["oldCustomer"]

    # Establece el idioma local en español
    locale.setlocale(locale.LC_TIME, 'es_AR.utf8')

    dateNow = datetime.date.today().strftime("%d de %B de %Y")
    context ={
                "fechaNow": dateNow,
                "oldCustomer": oldCustomer,
                "nroOperacion":operacionTitu.nro_operacion,
                "cliente": Cliente.objects.get(id=newCustomer).nombre,
                "domicilio": Cliente.objects.get(id=newCustomer).domic,
                "dni": Cliente.objects.get(id=newCustomer).dni,
                "localidad": Cliente.objects.get(id=newCustomer).loc,
                "provincia": Cliente.objects.get(id=newCustomer).prov,
                "estado_civil" : Cliente.objects.get(id=newCustomer).estado_civil,
                "fecha_nac" : Cliente.objects.get(id=newCustomer).fec_nacimiento,
                "ocupacion" : Cliente.objects.get(id=newCustomer).ocupacion,
                "telefono" : Cliente.objects.get(id=newCustomer).tel,
            }
            
    titularName = "Cambio de titular: " + str(Cliente.objects.get(id=newCustomer).nombre) + str(operacionTitu.nro_orden)
    urlPDF= os.path.join(settings.PDF_STORAGE_DIR, "titularidad.pdf")
    
    printPDFtitularidad(context,request.build_absolute_uri(),urlPDF)

    with open(urlPDF, 'rb') as pdf_file:
        response = HttpResponse(pdf_file,content_type="application/pdf")
        response['Content-Disposition'] = 'inline; filename='+titularName+'.pdf'
        return response

@login_required
def viewPDFArqueo(request, pk):
    arqueo = get_object_or_404(ArqueoCaja, pk=pk)

    # Seguridad: que el usuario pueda ver esa agencia
    if arqueo.agencia_id:
        suc = validate_agencia_allowed(request, int(arqueo.agencia_id))
        if not suc:
            return HttpResponseForbidden("Sin permisos para ver este arqueo.")

    return export_arqueo_pdf(request, arqueo, timeout_sec=300)

@require_GET
@login_required
def viewsPDFInforme(request):
    # usa los mismos filtros del querystring
    return export_caja(request, "pdf", timeout_sec=300)


@require_GET
@login_required
def viewsExcelInforme(request):
    # usa los mismos filtros del querystring
    return export_caja(request, "xlsx", timeout_sec=300)


@login_required
def viewsPDFInformePostVenta(request):
    """
    Genera el PDF de Post-Venta consultando la BD con filtros por querystring:
    ?search=&campania=&sucursal_id=&agencia_id=&estado=
    """
    # ----- leer filtros -----
    search    = (request.GET.get("search", "") or "").strip()
    campania  = (request.GET.get("campania", "") or "").strip()
    suc_id    = request.GET.get("sucursal_id") or request.GET.get("sucursal")
    ag_id     = request.GET.get("agencia_id")
    estado    = _estado_map(request.GET.get("estado", ""))  # <- corregido

    # ----- base queryset -----
    qs = (
        Ventas.objects
        .select_related("nro_cliente", "vendedor", "supervisor", "producto", "agencia")
        .only(
            "id", "nro_operacion", "campania", "fecha",
            "agencia_id",
            "nro_cliente__nombre", "nro_cliente__dni", "nro_cliente__tel",
            "nro_cliente__cod_postal", "nro_cliente__prov", "nro_cliente__loc",
            "nro_cliente__domic",
            "vendedor__nombre", "supervisor__nombre", "producto__nombre",
        )
    )

    if campania:
        qs = qs.filter(campania=campania)

    # sucursal/agencia (es lo mismo: agencia_id)
    for key in (suc_id, ag_id):
        if key:
            try:
                qs = qs.filter(agencia_id=int(key))
            except (TypeError, ValueError):
                pass

    if search:
        qs = qs.filter(_build_search_q(search))

    # ----- estado: solo anotar si hace falta -----
    if estado:
        latest_grade_sq = Subquery(
            Auditoria.objects.filter(venta=OuterRef("pk"))
            .order_by("-version").values("grade")[:1]
        )
        qs = qs.annotate(
            aud_count=Count("auditorias"),
            last_grade=latest_grade_sq,
        )
        if estado == "pendientes":
            qs = qs.filter(aud_count=0)
        elif estado == "realizadas":
            qs = qs.filter(aud_count__gt=0)
        elif estado == "aprobadas":
            qs = qs.filter(aud_count__gt=0, last_grade=True)
        elif estado == "desaprobadas":
            qs = qs.filter(aud_count__gt=0, last_grade=False)

    # ----- prefetch auditorías (evita N+1) -----
    aud_qs = (
        Auditoria.objects
        .only("venta_id", "version", "grade", "comentarios", "fecha_hora")
        .order_by("version")
    )
    qs = qs.prefetch_related(Prefetch("auditorias", queryset=aud_qs))

    # orden (tu "fecha" es char; si prefieres por id desc, cambia a .order_by("-id"))
    qs = qs.order_by("-fecha")

    # ----- armar estructura para el template del PDF -----
    datos_modificado = []
    for v in qs:
        c = v.nro_cliente
        info = {
            "Camp": v.campania or "---",
            "Cliente": (getattr(c, "nombre", "") or "---"),
            "DNI": str(getattr(c, "dni", "") or "---"),
            "Fec insc": (v.fecha or "---"),
            "Tel": str(getattr(c, "tel", "") or "---"),
            "CP": str(getattr(c, "cod_postal", "") or "---"),
            "Prov": (getattr(c, "prov", "") or "---"),
            "Loc": (getattr(c, "loc", "") or "---"),
            "Direc": (getattr(c, "domic", "") or "---"),
            "Vendedor": (getattr(getattr(v, "vendedor", None), "nombre", "") or "---"),
            "Supervisor": (getattr(getattr(v, "supervisor", None), "nombre", "") or "---"),
        }
        auditorias = [
            {
                "grade": a.grade,
                "fecha_hora": a.fecha_hora,
                "comentarios": a.comentarios or "",
            }
            for a in v.auditorias.all()  # ← ya viene prefetcheado y ordenado
        ]

        datos_modificado.append({
            "operacion": v.nro_operacion or "---",
            "info": info,
            "auditorias": auditorias,
        })

    # ----- generar PDF -----
    informeName = "Informe Post-Venta"
    urlPDF = os.path.join(settings.PDF_STORAGE_DIR, "postVentaInforme.pdf")

    printPDFinformePostVenta({"data": datos_modificado}, request.build_absolute_uri(), urlPDF)

    with open(urlPDF, "rb") as pdf_file:
        resp = HttpResponse(pdf_file, content_type="application/pdf")
        resp["Content-Disposition"] = f"inline; filename={informeName}.pdf"
        return resp

def _parse_fecha(fecha_str: str):
    """
    Acepta 'YYYY-MM-DD' o 'DD/MM/YYYY' (con o sin hora).
    Devuelve datetime (naive o aware según settings).
    """
    raw = (fecha_str or "").strip()[:10]
    fmt_candidates = ("%Y-%m-%d", "%d/%m/%Y")
    for fmt in fmt_candidates:
        try:
            dt = datetime.strptime(raw, fmt)
            # si usás TZ, la volvemos aware
            try:
                return make_aware(dt)
            except Exception:
                return dt
        except ValueError:
            pass
    # fallback: ahora
    try:
        return make_aware(datetime.now())
    except Exception:
        return datetime.now()


def _format_nro_recibo(nro):
    """
    Si viene como int o string simple, lo deja como string.
    Si ya viene formateado '0007-00000000', lo respeta.
    """
    if nro is None:
        return ""
    s = str(nro)
    return s


def viewPDFReciboCuota(request, pk):
    pago = PagoCannon.objects.get(id=pk)
    pagoData = pago.json()  # asumes que te devuelve dict con campos usados abajo

    # Venta / cliente / agencia
    venta = Ventas.objects.select_related("agencia", "nro_cliente").get(id=int(pagoData["venta_id"]))
    cliente = venta.nro_cliente  # tu FK a Cliente
    agencia = venta.agencia

    # Datos básicos
    numero_recibo = _format_nro_recibo(pagoData.get("nro_recibo"))
    fecha_dt = _parse_fecha(pagoData.get("fecha", ""))

    # Concepto e importes
    monto = pagoData.get("monto", 0)
    concepto = f"Pago de cuota N°{pagoData.get('nro_cuota')}"
    monto_letras = convertir_moneda_a_texto(monto)

    # Contexto que espera el template A4 DOBLE
    context = {
        "recibo": {
            "numero_formateado": numero_recibo,
            "condicion": "contado",             # o "cc" si tenés esa info
            "importe_letras": monto_letras,     # se muestra en mayúsculas en CSS si querés
            "efectivo": f"{monto:,.2f}".replace(",", "."),
            "cheque1_numero": "",
            "cheque1_banco": "",
            "cheque1_importe": "",
            "cheque2_numero": "",
            "cheque2_banco": "",
            "cheque2_importe": "",
            "cheque3_numero": "",
            "cheque3_banco": "",
            "cheque3_importe": "",
            "total": f"{monto:,.2f}".replace(",", "."),
            "concepto": concepto,
        },
        "cliente": {
            "nombre": getattr(cliente, "nombre", "") or "",
            "domicilio": getattr(cliente, "domic", "") or "",
            "localidad": getattr(cliente, "loc", "") or "",
            "cuit": getattr(cliente, "cuit", "") or "",  # si tu modelo no tiene cuit, quedará vacío
            "iva": "RM",  # "RI" | "RM" | "CF" | "MS"  (ajustá según tu dato real)
        },
        "fecha": fecha_dt,
        # Si querés inyectar datos de agencia en el header, podés
        # adaptar el template y usar estas claves:
        "agencia": {
            "direccion": getattr(agencia, "direccion", "") or "",
            "tel": getattr(agencia, "tel_ref", "") or "",
            "email": getattr(agencia, "email_ref", "") or "",
        },
    }

    # Archivo de salida
    informeName = "Recibo"
    urlPDF = os.path.join(settings.PDF_STORAGE_DIR, "pdf_recibo_cuota.pdf")

    # Render PDF con el template de doble copia A4
    # Asegurate de tener:
    # - templates/recibos/recibo_a4_doble.html
    # - templates/recibos/_recibo_contenido.html
    # - static/css/recibo_a4_doble.css
    printPDF(
        context,
        request.build_absolute_uri(),
        urlPDF,
        "pdf_recibo_doble.html",
        "static/css/pdfReciboCuota.css",
    )

    with open(urlPDF, "rb") as pdf_file:
        response = HttpResponse(pdf_file, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{informeName}.pdf"'
        return response
#endregion - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -


#region Caja - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
class Caja(TestLogin,generic.View):
    template_name = "caja.html"
    FILTROS_EXISTENTES = (
        ("tipo_mov","Tipo de movimiento"),
        ("metodoPago", "Metodo de pago"),
        ("fecha", "Fecha"),
        ("cobrador","Cobrador"),
        ("agencia","Agencia"),
        ("campania","Campaña"),
    )

    def get(self,request,*args, **kwargs):
        context ={}
        context["agencias_permitidas"] = [{"id":agencia.id,"pseudonimo":agencia.pseudonimo} for agencia in request.user.sucursales.all()]
        context["cuentas_de_cobro"] = [{"id":cuenta.id,"nombre":cuenta.alias} for cuenta in CuentaCobranza.objects.all()]
        context["metodos_de_pago"] = [{"id":metodo.id,"nombre":metodo.alias} for metodo in MetodoPago.objects.all()]
        context["campanias"] = getTodasCampaniasDesdeInicio()
        context["campaniasDisponibles"] = getCampanasDisponibles()


        
        # print(os.path.join(settings.BASE_DIR, "templates/mailPlantilla.html"))
        paramsDict = (request.GET).dict()
        clearContext = {key: value for key, value in paramsDict.items() if value != '' and key != 'page'}

        
        # Extrae las tuplas segun los querys filtrados en clearContext
        filtros_activados = list(filter(lambda x: x[0] in clearContext, self.FILTROS_EXISTENTES))

        # Por cada tupla se coloca de llave el valor 1 y se extrae el valor mediante su key de clearContext ( Por eso es [x[0]] )
        # Es lo mismo que decir clearContext["metodoPago"], etc, etc
        context["filtros"] = list(map(lambda x: {x[1], clearContext[x[0]]}, filtros_activados))
        
        return render(request, self.template_name, context)
        

class CierreCaja(TestLogin, generic.View):
    template_name = "cierreDeCaja.html"

    def get(self, request, *args, **kwargs):
        agencias = get_allowed_agencias(request)
        selected = agencias[0] if agencias else None

        fecha_hoy = today_fragment_ddmmyyyy()
        saldo = compute_saldo_diario_caja(selected.id, fecha_hoy) if selected else 0.0

        context = {
            "selected_agencia_id": selected.id if selected else "",
            "selected_agencia_label": str(selected) if selected else "Seleccionar",
            "agencias": agencias,
            "fecha": fecha_hoy,
            "admin": request.user,
            "saldo_diario": float(saldo),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        # ✅ Agencia por ID (no por "localidad, provincia")
        agencia_id = request.POST.get("agencia_id") or ""
        try:
            agencia_id_int = int(agencia_id)
        except Exception:
            return JsonResponse({"success": False, "error": "Agencia inválida."}, status=400)

        sucursal = validate_agencia_allowed(request, agencia_id_int)
        if not sucursal:
            return JsonResponse({"success": False, "error": "No tenés permisos para esa agencia."}, status=403)

        responsable = (request.POST.get("responsable") or "").strip()
        observaciones = (request.POST.get("observaciones") or "").strip()

        # ✅ saldo del día: NO confiar en lo que manda el front
        fecha_hoy = today_fragment_ddmmyyyy()
        total_segun_diario = compute_saldo_diario_caja(sucursal.id, fecha_hoy)

        arqueo = ArqueoCaja()
        arqueo.agencia = sucursal
        arqueo.fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        arqueo.admin = str(request.user)
        arqueo.responsable = responsable
        arqueo.observaciones = observaciones

        detalle = {}
        total_planilla = 0

        for b in BILLETES:
            raw = request.POST.get(f"p{b}", "")
            try:
                cantidad = int(raw) if str(raw).strip() != "" else 0
            except Exception:
                cantidad = 0

            importe = cantidad * b
            total_planilla += importe
            detalle[f"p{b}"] = {"cantidad": cantidad, "importeTotal": importe}

        arqueo.detalle = detalle
        arqueo.totalPlanilla = float(total_planilla)
        arqueo.totalSegunDiarioCaja = float(total_segun_diario)
        arqueo.diferencia = float(total_planilla) - float(total_segun_diario)

        arqueo.save()

        return JsonResponse(
            {
                "success": True,
                "urlPDF": reverse("sales:arqueoPDF", args=[arqueo.pk]),
                "urlCaja": reverse("sales:caja"),
            },
            safe=False,
        )


class OldArqueosView(TestLogin, generic.View):
    model = ArqueoCaja
    template_name = "listaDeArqueos.html"

    def get(self, request, *args, **kwargs):
        sucursales = get_allowed_agencias(request)
        ids = [s.id for s in sucursales]

        arqueos = ArqueoCaja.objects.filter(agencia_id__in=ids).order_by("-pk")

        context = {
            "arqueos": arqueos,
            "sucursales": sucursales,  
        }

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            arqueos_list = []
            for a in arqueos:
                arqueos_list.append(
                    {
                        "pk": a.pk,
                        "agencia_id": a.agencia_id or 0,
                        "sucursal": str(a.agencia) if a.agencia else "",
                        "fecha": a.fecha,
                        "admin": a.admin,
                        "responsable": a.responsable,
                        "totalPlanilla": float(a.totalPlanilla or 0),
                        "pdf_url": reverse("sales:arqueoPDF", args=[a.pk]),  # ✅ link directo
                    }
                )
            return JsonResponse(arqueos_list, safe=False)

        return render(request, self.template_name, context)


# ✅ Endpoint AJAX: saldo según diario de caja (cash) para agencia seleccionada
class ArqueoSaldoAPI(TestLogin, generic.View):
    def get(self, request, *args, **kwargs):
        agencia_id = request.GET.get("agencia_id") or ""
        try:
            agencia_id_int = int(agencia_id)
        except Exception:
            return JsonResponse({"success": False, "error": "Agencia inválida."}, status=400)

        sucursal = validate_agencia_allowed(request, agencia_id_int)
        if not sucursal:
            return JsonResponse({"success": False, "error": "Sin permisos."}, status=403)

        fecha = request.GET.get("fecha") or today_fragment_ddmmyyyy()
        saldo = compute_saldo_diario_caja(sucursal.id, fecha)

        return JsonResponse(
            {
                "success": True,
                "agencia_id": sucursal.id,
                "fecha": fecha,
                "saldo": float(saldo),
            },
            safe=False,
        )


@require_GET
@login_required
def movs_pagos_list_api(request):
    page = max(1, _parse_int(request.GET.get("page"), 1) or 1)
    page_size = max(1, min(200, _parse_int(request.GET.get("page_size"), 20) or 20))
    search = (request.GET.get("search") or "").strip()

    # Filtros esperados
    origen = (request.GET.get("origen") or "").strip().lower()
    fecha = _normalize_fecha_fragment(request.GET.get("fecha") or "")
    concepto = (request.GET.get("concepto") or "").strip()
    tipo_mov = (request.GET.get("tipo_mov") or "").strip().lower()  # ingreso|egreso
    nro_cuota = _parse_int(request.GET.get("nro_cuota"), None)
    monto_min = request.GET.get("monto_min")
    monto_max = request.GET.get("monto_max")
    metodo_pago_ids = _parse_int_list(request.GET.get("metodo_pago") or "")
    cobrador_ids = _parse_int_list(request.GET.get("cobrador") or "")

    # Seguridad: sólo sucursales del usuario (o todas si superuser)
    allowed_agencias = set(_get_allowed_agencia_ids(request))
    requested_agencias = set(_parse_int_list(request.GET.get("agencia") or ""))
    agencia_ids = list(requested_agencias.intersection(allowed_agencias)) if requested_agencias else list(allowed_agencias)

    pagos_qs = (
        PagoCannon.objects
        .select_related(
            "venta",
            "venta__nro_cliente",
            "venta__agencia",
            "metodo_pago",
            "cobrador",
            "responsable_pago",
        )
        .filter(venta__agencia__id__in=agencia_ids)
    )

    externos_qs = (
        MovimientoExterno.objects
        .select_related("metodoPago", "agencia", "ente")
        .filter(agencia__id__in=agencia_ids)
    )
    if origen in ("cannon", "cannons", "pago", "pagos"):
        externos_qs = externos_qs.none()
    elif origen in ("externo", "externos"):
        pagos_qs = pagos_qs.none()

    # --- Filtro tipo_mov (ingreso/egreso)
    if tipo_mov == "egreso":
        pagos_qs = pagos_qs.none()  # PagoCannon es ingreso
        externos_qs = externos_qs.filter(movimiento__iexact="egreso")
    elif tipo_mov == "ingreso":
        externos_qs = externos_qs.filter(movimiento__iexact="ingreso")

    # --- Filtro fecha (contiene el día)
    if fecha:
        pagos_qs = pagos_qs.filter(fecha__icontains=fecha)
        externos_qs = externos_qs.filter(fecha__icontains=fecha)

    # --- nro_cuota solo aplica a pagos
    if nro_cuota is not None:
        pagos_qs = pagos_qs.filter(nro_cuota=nro_cuota)

    # --- concepto (externos) y/o texto de cliente
    if concepto:
        pagos_qs = pagos_qs.filter(
            Q(venta__nro_cliente__nombre__icontains=concepto)
            | Q(nro_recibo__icontains=concepto)
            | Q(venta__nro_operacion__icontains=concepto)
        )
        externos_qs = externos_qs.filter(
            Q(concepto__icontains=concepto)
            | Q(denominacion__icontains=concepto)
            | Q(nroComprobante__icontains=concepto)
            | Q(nroIdentificacion__icontains=concepto)
        )

    # --- búsqueda general
    if search:
        pagos_qs = pagos_qs.filter(
            Q(venta__nro_cliente__nombre__icontains=search)
            | Q(venta__nro_operacion__icontains=search)
            | Q(nro_recibo__icontains=search)
        )
        externos_qs = externos_qs.filter(
            Q(concepto__icontains=search)
            | Q(denominacion__icontains=search)
            | Q(nroComprobante__icontains=search)
            | Q(nroIdentificacion__icontains=search)
        )

    # --- método de pago
    if metodo_pago_ids:
        pagos_qs = pagos_qs.filter(metodo_pago_id__in=metodo_pago_ids)
        externos_qs = externos_qs.filter(metodoPago_id__in=metodo_pago_ids)

    # --- cobrador (PagoCannon.cobrador) / ente recaudador (MovimientoExterno.ente)
    if cobrador_ids:
        pagos_qs = pagos_qs.filter(cobrador_id__in=cobrador_ids)
        externos_qs = externos_qs.filter(ente_id__in=cobrador_ids)

    # --- monto min/max
    try:
        if monto_min not in (None, ""):
            vmin = float(monto_min)
            pagos_qs = pagos_qs.filter(monto__gte=vmin)
            externos_qs = externos_qs.filter(dinero__gte=vmin)
    except Exception:
        pass
    try:
        if monto_max not in (None, ""):
            vmax = float(monto_max)
            pagos_qs = pagos_qs.filter(monto__lte=vmax)
            externos_qs = externos_qs.filter(dinero__lte=vmax)
    except Exception:
        pass

    # Total (para paginación). Nota: la ordenación se hace en Python (fechas string).
    total = int(pagos_qs.count() + externos_qs.count())

    # Para armar página N sin traer TODO: overfetch razonable y luego sort + slice
    offset = (page - 1) * page_size
    take = min(total, offset + (page_size * 3)) if total else (page_size * 3)

    pagos = list(pagos_qs.order_by("-id")[:take])
    externos = list(externos_qs.order_by("-id")[:take])

    merged = []
    for p in pagos:
        merged.append(_row_from_pago(p))
    for m in externos:
        merged.append(_row_from_externo(m))

    merged.sort(key=lambda x: x[0], reverse=True)
    page_rows = [d for _, d in merged[offset: offset + page_size]]

    # Resumen por método de pago (neto = ingreso - egreso)
    resumen_map = {}

    # Pagos: sólo ingreso
    for r in pagos_qs.values("metodo_pago_id", "metodo_pago__alias").annotate(
        ingreso=Coalesce(Sum("monto"), 0)
    ):
        mid = r.get("metodo_pago_id")
        alias = r.get("metodo_pago__alias") or "Sin método"
        ingreso = float(r.get("ingreso") or 0)
        key = f"mp-{mid}" if mid else "mp-none"
        resumen_map[key] = {
            "metodo_pago_id": mid,
            "metodo_pago": alias,
            "ingreso": ingreso,
            "egreso": 0.0,
        }

    # Externos: ingreso/egreso
    for r in externos_qs.values("metodoPago_id", "metodoPago__alias").annotate(
        ingreso=Coalesce(Sum("dinero", filter=Q(movimiento__iexact="ingreso")), 0.0),
        egreso=Coalesce(Sum("dinero", filter=Q(movimiento__iexact="egreso")), 0.0),
    ):
        mid = r.get("metodoPago_id")
        alias = r.get("metodoPago__alias") or "Sin método"
        ingreso = float(r.get("ingreso") or 0)
        egreso = float(r.get("egreso") or 0)
        key = f"mp-{mid}" if mid else "mp-none"
        if key not in resumen_map:
            resumen_map[key] = {
                "metodo_pago_id": mid,
                "metodo_pago": alias,
                "ingreso": 0.0,
                "egreso": 0.0,
            }
        resumen_map[key]["ingreso"] += ingreso
        resumen_map[key]["egreso"] += egreso

    resumen = []
    total_neto = 0.0
    # Ordenar por alias para que sea estable
    for item in sorted(resumen_map.values(), key=lambda x: (x.get("metodo_pago") or "")):
        neto = float(item["ingreso"] or 0) - float(item["egreso"] or 0)
        total_neto += neto
        resumen.append({
            "metodo_pago_id": item["metodo_pago_id"],
            "metodo_pago": item["metodo_pago"],
            "ingreso": _format_money(item["ingreso"]),
            "egreso": _format_money(item["egreso"]),
            "neto": _format_money(neto),
            "neto_value": neto,
        })

    resumen.append({
        "metodo_pago_id": None,
        "metodo_pago": "Total",
        "ingreso": "-",
        "egreso": "-",
        "neto": _format_money(total_neto),
        "neto_value": total_neto,
    })

    return JsonResponse(
        {
            "results": page_rows,
            "count": total,
            "page": page,
            "page_size": page_size,
            "summary": resumen,
        },
        safe=False,
    )


@require_GET
@login_required
def movs_pagos_detail_api(request, kind: str, pk: int):
    kind = (kind or "").strip().lower()
    allowed_agencias = set(_get_allowed_agencia_ids(request))

    if kind == "pago":
        obj = get_object_or_404(
            PagoCannon.objects.select_related(
                "venta",
                "venta__nro_cliente",
                "venta__agencia",
                "metodo_pago",
                "cobrador",
                "responsable_pago",
            ),
            pk=pk,
        )
        agencia_id = getattr(getattr(getattr(obj, "venta", None), "agencia", None), "id", None)
        if agencia_id not in allowed_agencias and not getattr(request.user, "is_superuser", False):
            return JsonResponse({"detail": "No autorizado"}, status=403)

        venta = obj.venta
        cliente = getattr(venta, "nro_cliente", None)
        data = {
            "id": obj.id,
            "fecha": obj.fecha,
            "monto": obj.monto,
            "monto_por_chance": obj.monto / len(getattr(venta, "cantidadContratos", "") or []) if venta else 0,
            "monto_formatted": _format_money(obj.monto or 0),
            "monto_por_chance_formatted": _format_money(
                (obj.monto or 0) / len(getattr(venta, "cantidadContratos", "") or []) if venta else 0
            ),
            "nro_cuota": obj.nro_cuota,
            "nro_recibo": obj.nro_recibo,
            "campana_de_pago": obj.campana_de_pago,
            "metodo_pago": getattr(getattr(obj, "metodo_pago", None), "alias", ""),
            "cobrador": getattr(getattr(obj, "cobrador", None), "alias", ""),
            "responsable": getattr(getattr(obj, "responsable_pago", None), "username", ""),
            "venta": {
                "id": getattr(venta, "id", None),
                "nro_operacion": getattr(venta, "nro_operacion", ""),
                "chances": len(getattr(venta, "cantidadContratos", "") or []),
                "campania": getattr(venta, "campania", ""),
                "agencia": getattr(getattr(venta, "agencia", None), "pseudonimo", ""),
                "cliente": {
                    "id": getattr(cliente, "id", None),
                    "nombre": getattr(cliente, "nombre", ""),
                    "nro_cliente": getattr(cliente, "nro_cliente", ""),
                    "dni": getattr(cliente, "dni", ""),
                },
            },
        }
        return JsonResponse({"kind": "pago", "data": data}, safe=False)

    if kind == "externo":
        obj = get_object_or_404(
            MovimientoExterno.objects.select_related("metodoPago", "agencia", "ente"),
            pk=pk,
        )
        agencia_id = getattr(getattr(obj, "agencia", None), "id", None)
        if agencia_id not in allowed_agencias and not getattr(request.user, "is_superuser", False):
            return JsonResponse({"detail": "No autorizado"}, status=403)

        data = {
            "id": obj.id,
            "fecha": obj.fecha,
            "movimiento": obj.movimiento,
            "monto": obj.dinero,
            "monto_formatted": _format_money(obj.dinero or 0),
            "concepto": obj.concepto,
            "metodo_pago": getattr(getattr(obj, "metodoPago", None), "alias", ""),
            "agencia": getattr(getattr(obj, "agencia", None), "pseudonimo", ""),
            "ente": getattr(getattr(obj, "ente", None), "alias", ""),
            "campania": obj.campania,
            "tipoIdentificacion": obj.tipoIdentificacion,
            "nroIdentificacion": obj.nroIdentificacion,
            "tipoComprobante": obj.tipoComprobante,
            "nroComprobante": obj.nroComprobante,
            "denominacion": obj.denominacion,
            "tipoMoneda": obj.tipoMoneda,
            "premio": bool(obj.premio),
            "adelanto": bool(obj.adelanto),
            "observaciones": obj.observaciones,
        }
        return JsonResponse({"kind": "externo", "data": data}, safe=False)

    return JsonResponse({"detail": "Tipo inválido"}, status=400)


#endregion - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -


#region Specifics Functions - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
def requestMovimientos(request):
    try:
        #region Logica para obtener los movimientos segun los filtros aplicados 
        agencia = request.user.sucursales.first().pseudonimo if not request.GET.get("agencia") else request.GET.get("agencia") # Si no hay agencia seleccionada, se coloca la primera sucursal del usuario
        # print(request)
        cannons = dataStructureCannons(agencia)
        # print(cannons)
        cannons = list(filter(lambda x: x["estado"]["data"].lower() in ["pagado", "parcial"], cannons))
        # allMovimientos = cannons
        allMovimientos = dataStructureMovimientosExternos(agencia) + cannons
        allMovimientosTidy = sorted(allMovimientos, key=lambda x: datetime.strptime(formatear_dd_mm_yyyy_h_m(x['fecha']["data"]), '%d/%m/%Y %H:%M'),reverse=True) # Ordenar de mas nuevo a mas viejo los movimientos
        # Agregar a cada diccionario del movimiento el campo id_cont para poder identificarlo en el template 
        for i, mov in enumerate(allMovimientosTidy):
            mov["id_cont"] = i
        #endregion
        

        response_data ={
            "request": request.GET,
            "movs": allMovimientosTidy
        }
        print(f"request - - - - - \n {response_data['request']}")

        

        movs = filterMainManage(response_data["request"], response_data["movs"])
        
        request.session["informe_data"] = movs # Por si se quiere imprimir el informe

        # region Logica para obtener el resumen de cuenta de los diferentes tipos de pagos
        resumenEstadoCuenta=[]
        total_money = 0

        metodosPagos = MetodoPago.objects.all()
        for metodo in metodosPagos:
            estado_cuenta_by_metodo = {}
            estado_cuenta_by_metodo["verbose_name"] = metodo.alias

            metodoClean = metodo.alias.replace(" ","_").lower() + "_total_money"
            estado_cuenta_by_metodo["name_clean"] = metodoClean

            movs_by_metodo = list(filter(lambda mov:mov["metodoPago"]["data"] == str(metodo.id), movs))
            money_by_metodo_ingreso = sum([mov["monto"]["data"] for mov in movs_by_metodo if mov["tipo_mov"]["data"] == "ingreso"])
            money_by_metodo_egreso = sum([mov["monto"]["data"] for mov in movs_by_metodo if mov["tipo_mov"]["data"] == "egreso"])
            print(f"Metodo: {metodo.alias} - Ingreso: {money_by_metodo_ingreso} - Egreso: {money_by_metodo_egreso}")
            money_by_metodo = money_by_metodo_ingreso - money_by_metodo_egreso
            estado_cuenta_by_metodo["money"] = formatear_moneda_sin_centavos(money_by_metodo)
            resumenEstadoCuenta.append(estado_cuenta_by_metodo)
            
            total_money += money_by_metodo

        dictTotal = {
            "verbose_name": "Total",
            "name_clean": "total_money",
            "money": formatear_moneda_sin_centavos(total_money)
        }
        resumenEstadoCuenta.append(dictTotal)
        
        #endregion

        #region Paginación
        page = request.GET.get('page')
        items_per_page = 20  # Número de elementos por página
        paginator = Paginator(movs, items_per_page)

        try:
            movs = paginator.page(page)
        except PageNotAnInteger:
            movs = paginator.page(1)
        except EmptyPage:
            movs = paginator.page(paginator.num_pages)
        #endregion -----------------------------------------------------

        return JsonResponse({"data": list(movs), "numbers_pages": paginator.num_pages,"estadoCuenta":resumenEstadoCuenta,"status": True}, safe=False)
    except Exception as e:
        print(e)
        return JsonResponse({"data": [], "numbers_pages": 0,"filtros":[],"estadoCuenta":{},"status": False}, safe=False)


# Funcion para crear un nuevo movimiento externo
def createNewMov(request):
    if request.method == 'POST':
        newMov = MovimientoExterno()
        movimiento = request.POST.get("movimiento")
        print(movimiento)
        newMov.movimiento=movimiento
        newMov.agencia = Sucursal.objects.get(id = request.POST.get('agencia'))
        newMov.ente= CuentaCobranza.objects.filter(id=request.POST.get('ente')).first() 
        newMov.fecha=request.POST.get("fecha")
        newMov.concepto= request.POST.get('concepto')
        newMov.metodoPago= MetodoPago.objects.filter(id=request.POST.get('tipoPago')).first() 
        newMov.dinero= float(request.POST.get('dinero'))
        newMov.campania = request.POST.get("campania")

        if movimiento == 'ingreso':
            print("Entro a ingreso")

            newMov.tipoMoneda = request.POST.get('tipoMoneda')
        elif movimiento == 'egreso':
            print("Entro a egreso")
            newMov.tipoIdentificacion=request.POST.get('tipoID')
            newMov.nroIdentificacion=request.POST.get('nroIdentificacion')
            newMov.tipoComprobante=request.POST.get('tipoComprobante')
            newMov.nroComprobante=request.POST.get('nroComprobante')
            newMov.denominacion=request.POST.get('denominacion')
            if(request.POST.get('adelanto_premio') == "premio"):
                newMov.premio= True
            elif(request.POST.get('adelanto_premio') == "adelanto"):
                newMov.adelanto = True
        else:
            return HttpResponseBadRequest('Fallo en el servidor', status=405)
                 
        newMov.save()
        return JsonResponse({'status': True, 'message': 'Movimiento creado exitosamente'})
        

    return JsonResponse({'status': False, 'message': ' Ha sucedido un error y no se pudo completar la operacion'})


# Funcion para devolver las ventas (Utilizada en el sector de auditorias)
def requestVentasAuditoria(request):
    if(request.method == "GET"):
        sucursal = request.GET.get('sucursal')
        campania = request.GET.get('campania')
        responseData = {"ventas": [],"resumenAuditorias":{"pendientes": 0, "realizadas":0, "aprobadas":0, "desaprobadas":0}}
        allVentas = Ventas.objects.filter(campania=campania, agencia=sucursal)

        auditorias_realidas = allVentas.filter(auditoria__0__realizada=True)
        responseData["resumenAuditorias"]["realizadas"] = len(auditorias_realidas)

        auditorias_pendientes = allVentas.filter(auditoria__0__realizada=False)
        responseData["resumenAuditorias"]["pendientes"] = len(auditorias_pendientes)

        auditorias_realidas_list = list(auditorias_realidas.values())
        auditorias_aprobadas = len(list(filter(lambda x: x["auditoria"][-1]["grade"] == True,auditorias_realidas_list)))
        responseData["resumenAuditorias"]["aprobadas"] = auditorias_aprobadas

        auditorias_desaprobadas = len(list(filter(lambda x: x["auditoria"][-1]["grade"] == False,auditorias_realidas_list)))
        responseData["resumenAuditorias"]["desaprobadas"] = auditorias_desaprobadas

        for v in allVentas:
            ventaToDict = {
                "id": v.pk,
                "nroOrden": v.nro_orden,
                "campania": v.campania,
                "cliente": v.nro_cliente.nombre,
                "dni": v.nro_cliente.dni,
                "fec_insc": v.fecha,
                "tel": v.nro_cliente.tel,
                "cp": v.nro_cliente.cod_postal,
                "prov": v.nro_cliente.prov,
                "loc": v.nro_cliente.loc,
                "direc": v.nro_cliente.domic,
                "vendedor": str(v.vendedor),
                "supervisor": str(v.supervisor),
                "auditoria": v.auditoria,
            }
            responseData["ventas"].append(ventaToDict)
        request.session["postVenta_info"] = responseData

        return JsonResponse(responseData, safe=False)
   
#endregion - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

#region MANEJO DE DOLAR ------------------------------------------------------------------
CACHE_KEY = "cotizaciones:dolar:blue_oficial:v1"
CACHE_TTL_SECONDS = 60

def cotizaciones_dolar(request):
    cached = cache.get(CACHE_KEY)
    if cached:
        return JsonResponse({"cached": True, **cached}, status=200)

    try:
        data = asyncio.run(get_dolar_blue_y_oficial_async())
        cache.set(CACHE_KEY, data, CACHE_TTL_SECONDS)
        return JsonResponse({"cached": False, **data}, status=200)
    except Exception as e:
        # fallback: si falla y tenemos algo viejo en cache, lo devolvemos igual
        stale = cache.get(CACHE_KEY)
        if stale:
            return JsonResponse({"cached": True, "stale": True, **stale}, status=200)

        return JsonResponse({"error": str(e)}, status=502)
#endregion -----------------------------------------------------------------------------